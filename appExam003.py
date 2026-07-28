# app.py
from datetime import datetime, timedelta
import os, csv, io, json, random, string, bisect, secrets, hashlib, sys
import zipfile, math, re, shutil, subprocess
from functools import wraps
from flask import (Flask, request, Response, redirect, url_for,
                   render_template_string, session, send_file, abort, flash,
                   render_template, jsonify)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from flask_migrate import Migrate 
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer
from flask_socketio import SocketIO, emit, join_room, leave_room
from sqlalchemy import func, desc, or_, case, inspect, text
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, Image)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from scheduler import start_scheduler
from collections import defaultdict
from cryptography.fernet import Fernet
from math import ceil
import sqlite3, socket, smtplib, traceback
from scheduler import scheduler

RECYCLE_RETENTION_DAYS = 730
#from flask_login import login_required
try:
    import openpyxl
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

def resource_path(relative_path):
    """ Get absolute path for PyInstaller bundled files"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".") #Uses current project directory
    #Combines base + file paths and returns absolute path
    return os.path.join(base_path, relative_path)

# Initialize Flask app with instance path
app = Flask(__name__, instance_path=resource_path("instance"))

# =========================================================
# ENCRYPTION KEY
# =========================================================
KEY_FILE = resource_path("instance/secret.key")
def load_key():
    if not os.path.exists(KEY_FILE):
        key = Fernet.generate_key()
        with open(KEY_FILE, "wb") as f:
            f.write(key)
    with open(KEY_FILE, "rb") as f:
        return f.read()

cipher = Fernet(load_key())
# Secret key for sessions
app.config["SECRET_KEY"] = os.urandom(24)
# ------Generate Email Reset Tokens------
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])

### SQLAlchemy config
# =========================================================
# POSTGRESQL CONFIG
# =========================================================
DB_USER = "postgres"
DB_PASSWORD = "ABCD1234"
DB_HOST = "localhost"
DB_PORT = "5432"
DB_NAME = "examdb"
app.config['SQLALCHEMY_DATABASE_URI'] = (
    "postgresql+psycopg2://postgres:ABCD1234@localhost/examdb" )
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Initialize database and migration
db = SQLAlchemy(app)
migrate = Migrate(app, db)
def backup_postgres():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_folder = resource_path("backups")
    os.makedirs(backup_folder, exist_ok=True)
    backup_file = os.path.join(
        backup_folder,
        f"{DB_NAME}_backup_{timestamp}.sql"
    )
    env = os.environ.copy()
    env["PGPASSWORD"] = DB_PASSWORD
    pg_dump_path = r"C:\Program Files\PostgreSQL\18\bin\pg_dump.exe"
    subprocess.run(
        [
            pg_dump_path,
            "-U", DB_USER,
            "-h", DB_HOST,
            "-p", str(DB_PORT),
            "-f", backup_file,
            DB_NAME
        ],
        env=env,
        check=True
    )
    print(f"✅ Database backup created: {backup_file}")
    
def sync_database_schema():
    print("\n========== DATABASE SYNC ==========")
    backup_postgres()
    db.create_all()
    inspector = inspect(db.engine)
    model_tables = {
        table.name: table
        for table in db.metadata.sorted_tables
    }
    db_tables = set(inspector.get_table_names())
    # ===================================
    # CREATE MISSING COLUMNS
    # ===================================
    for table_name, table in model_tables.items():
        db_columns = {
            c["name"]
            for c in inspector.get_columns(table_name)
        }
        for column in table.columns:
            if column.name not in db_columns:
                sql_type = str(
                    column.type.compile(db.engine.dialect)
                )
                sql = (
                    f'ALTER TABLE "{table_name}" '
                    f'ADD COLUMN "{column.name}" '
                    f'{sql_type}'
                )
                db.session.execute(text(sql))
                print(
                    f"➕ Added column: "
                    f"{table_name}.{column.name}"
                )
    db.session.commit()
    inspector = inspect(db.engine)
    # ===================================
    # DELETE EXTRA COLUMNS
    # ===================================
    for table_name, table in model_tables.items():
        db_columns = {
            c["name"]
            for c in inspector.get_columns(table_name)
        }
        model_columns = {
            c.name
            for c in table.columns
        }
        extra_columns = db_columns - model_columns
        for col in extra_columns:
            sql = (
                f'ALTER TABLE "{table_name}" '
                f'DROP COLUMN "{col}" CASCADE'
            )
            db.session.execute(text(sql))
            print(
                f"🗑 Removed column: "
                f"{table_name}.{col}"
            )
    db.session.commit()
    # ===================================
    # SYNC UNIQUE CONSTRAINTS
    # ===================================
    inspector = inspect(db.engine)
    for table_name, table in model_tables.items():
        db_uniques = {
            uc["name"]: tuple(uc["column_names"])
            for uc in inspector.get_unique_constraints(table_name)
            if (
                uc["name"]
                and uc["name"].startswith("uq_")
            )
        }
        model_uniques = {}
        if table.constraints:
            for constraint in table.constraints:
                if (
                    isinstance(constraint, db.UniqueConstraint)
                    and constraint.name
                    and constraint.name.startswith("uq_")
                ):
                    model_uniques[
                        constraint.name
                    ] = tuple(
                        c.name
                        for c in constraint.columns
                    )
        # -----------------------------------
        # Remove obsolete UNIQUE constraints
        # -----------------------------------
        for db_name in list(db_uniques.keys()):
            # Never touch PostgreSQL-generated constraints.
            if not db_name.startswith("uq_"):
                continue
            if db_name not in model_uniques:
                try:
                    sql = (
                        f'ALTER TABLE "{table_name}" '
                        f'DROP CONSTRAINT "{db_name}"'
                    )
                    db.session.execute(text(sql))
                    print(
                        f"🗑 Removed UNIQUE constraint: "
                        f"{table_name}.{db_name}"
                    )
                except Exception as e:
                    db.session.rollback()
                    import traceback
                    print("=" * 80)
                    print(f"Failed to remove constraint: {db_name}")
                    traceback.print_exc()
                    print("=" * 80)
        # -----------------------------------
        # Add missing UNIQUE constraints
        # -----------------------------------
        for model_name, cols in model_uniques.items():
            if (
                model_name not in db_uniques
                or db_uniques.get(model_name) != cols
            ):
                if model_name in db_uniques:
                    try:
                        sql = (
                            f'ALTER TABLE "{table_name}" '
                            f'DROP CONSTRAINT "{model_name}"'
                        )
                        db.session.execute(text(sql))
                    except Exception:
                        pass
                cols_sql = ", ".join(
                    f'"{c}"'
                    for c in cols
                )
                sql = (
                    f'ALTER TABLE "{table_name}" '
                    f'ADD CONSTRAINT "{model_name}" '
                    f'UNIQUE ({cols_sql})'
                )
                db.session.execute(text(sql))
                print(
                    f"➕ Added UNIQUE constraint: "
                    f"{table_name}.{model_name}"
                )
    db.session.commit()
    inspector = inspect(db.engine)
    db_tables = set(
        inspector.get_table_names()
    )
    model_table_names = set(
        model_tables.keys()
    )
    # ===================================
    # DELETE EXTRA TABLES
    # ===================================
    protected_tables = {
        "alembic_version"
    }
    extra_tables = (
        db_tables
        - model_table_names
        - protected_tables
    )
    for table in extra_tables:
        sql = (
            f'DROP TABLE "{table}" CASCADE'
        )
        db.session.execute(text(sql))
        print(
            f"🗑 Removed table: {table}"
        )
    db.session.commit()
    print("✅ Database sync complete.")

# Support chat
##socketio = SocketIO(app)
socketio = SocketIO( app, cors_allowed_origins="*", async_mode="threading" )

# =========================================================
# FLASK MAIL CONFIG
# =========================================================
#sensitive app password is not hardcoded in your code or pushed to GitHub accidentally.
#set it as environment variable in CMD: setx MAIL_PASSWORD "abcd efgh ijkl mnop"  ,, import os
#Restart terminal
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465 #587
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USERNAME'] = ''
app.config['MAIL_PASSWORD'] = ''
app.config['MAIL_DEFAULT_SENDER'] = ''
##mail = Mail(app)
mail = Mail()
##def init_mail(app):
mail.init_app(app)

# =========================================================
# SERIALIZER
# =========================================================
serializer = URLSafeTimedSerializer( app.config['SECRET_KEY'] )
# Hardcoded licenser email
##LICENSER_EMAIL = "awise303@gmail.com"
####password(self, pw): return check_password_hash(self.pw_hash, pw)
class Institution(db.Model):
    __tablename__ = "institution"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column( db.String(200), unique=True, nullable=False, index=True )
    code = db.Column( db.String(50), unique=True, nullable=True, index=True )
    description = db.Column( db.Text, nullable=True )
    active = db.Column( db.Boolean, default=True )
    is_system = db.Column( db.Boolean, default=False, index=True )    
    created_at = db.Column( db.DateTime, default=datetime.utcnow )
    # recycle bin support
    deleted = db.Column( db.Boolean, default=False, index=True )
    deleted_at = db.Column( db.DateTime, nullable=True )
    owner_admin_id = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=True, index=True )    
    admin_id = db.Column( db.Integer, db.ForeignKey("user.id"), nullable=True,
                          index=True )
    approved = db.Column(db.Boolean, default=False)
    approved_by = db.Column(db.Integer, db.ForeignKey("user.id"))
    approved_at = db.Column(db.DateTime)
    owner_admin = db.relationship( "User", foreign_keys=[owner_admin_id],
        backref=db.backref( "owned_institutions", lazy=True ) )
    admin = db.relationship( "User", foreign_keys=[admin_id],
        backref=db.backref( "administered_institutions", lazy=True ) )
    approver = db.relationship( "User", foreign_keys=[approved_by],
        backref=db.backref( "approved_institutions", lazy=True ) )

    def __repr__(self):
        return f"<Institution {self.name}>"

class AcademicClass(db.Model):
    __tablename__ = "academic_class"
    id = db.Column( db.Integer, primary_key=True )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=False, index=True )
    class_code = db.Column( db.Integer, nullable=False )
    class_name = db.Column( db.String(100), nullable=False )
    education_level = db.Column( db.String(30), nullable=True )    
    display_order = db.Column( db.Integer, default=0, nullable=False )
    active = db.Column( db.Boolean, default=True, nullable=False )
    created_at = db.Column( db.DateTime, default=datetime.utcnow )
    updated_at = db.Column( db.DateTime, default=datetime.utcnow,
        onupdate=datetime.utcnow )
    institution = db.relationship( "Institution",
       foreign_keys=[institution_id], backref=db.backref( "academic_classes", lazy=True,
        cascade="all, delete-orphan" ) )
    __table_args__ = (
        db.UniqueConstraint( "institution_id", "class_name",
            name="uq_institution_class_name" ),
        db.UniqueConstraint( "institution_id", "class_code",
            name="uq_institution_class_code" ), )
    def __repr__(self):
        return ( f"<AcademicClass " f"{self.class_name}>" )

class StudentClassHistory(db.Model):
    __tablename__ = "student_class_history"
    id = db.Column( db.Integer, primary_key=True )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=False, index=True )
    student_id = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=False, index=True )
    academic_class_id = db.Column( db.Integer,
        db.ForeignKey("academic_class.id"), nullable=False, index=True )
    session = db.Column( db.String(30), nullable=False )
    term = db.Column( db.String(30), nullable=False )
    change_type = db.Column( db.String(30), default="Promotion",
        nullable=False )
    remarks = db.Column( db.Text, nullable=True )
    changed_by = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=True )
    changed_at = db.Column( db.DateTime, default=datetime.utcnow )
    institution = db.relationship("Institution",
                                  foreign_keys=[institution_id],)
    student = db.relationship( "User", foreign_keys=[student_id] )
    academic_class = db.relationship( "AcademicClass" )
    admin = db.relationship( "User", foreign_keys=[changed_by] )    

class InstitutionTransfer(db.Model):
    __tablename__ = "institution_transfer"
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column( db.Integer,  db.ForeignKey("user.id"),
        nullable=False, index=True )
    from_institution_id = db.Column( db.Integer,
        db.ForeignKey("institution.id"), nullable=False, index=True )
    to_institution_id = db.Column( db.Integer,
        db.ForeignKey("institution.id"), nullable=False, index=True )
    status = db.Column( db.String(20), default="pending", index=True )
    # pending
    # approved
    # rejected
    # cancelled
    requested_at = db.Column( db.DateTime, default=datetime.utcnow )
    processed_at = db.Column( db.DateTime, nullable=True )
    processed_by = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=True )
    reason = db.Column( db.Text, nullable=True )
    student = db.relationship( "User", foreign_keys=[student_id]
    )
    from_institution = db.relationship( "Institution",
        foreign_keys=[from_institution_id] )
    to_institution = db.relationship( "Institution",
        foreign_keys=[to_institution_id] )
    processor = db.relationship( "User", foreign_keys=[processed_by] )    

class User(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    username    = db.Column(db.Text, unique=True, nullable=False)
    pw_hash     = db.Column(db.Text, nullable=False)
    role        = db.Column(db.String(20), nullable=False, default='student')
    allowed_subject_ids_json = db.Column(db.Text, default='[]')
    allowed_group_ids_json   = db.Column(db.Text, default='[]')
    first_name  = db.Column(db.Text)
    last_name   = db.Column(db.Text)
    email       = db.Column(db.String(255), nullable=True)
    # ===============================
    # REPORT CARD / STUDENT PROFILE
    # ===============================
##    class_name = db.Column(db.String(120), nullable=True)
    student_class = db.Column(db.Integer, nullable=True, index=True)
    academic_class_id = db.Column( db.Integer,
        db.ForeignKey("academic_class.id"), nullable=True, index=True )
    academic_class = db.relationship( "AcademicClass",
        foreign_keys=[academic_class_id] )    
    age = db.Column(db.Integer, nullable=True)
    passport_path = db.Column( db.String(300), nullable=True )
    days_open = db.Column( db.Integer, default=0 )
    days_present = db.Column( db.Integer, default=0 )
    registration_number = db.Column( db.String(100), nullable=True,
        index=True )  

    # ===============================
    # RESET
    # ===============================
    reset_token = db.Column(db.String(255), nullable=True)
    reset_used  = db.Column( db.Boolean, default=False )
    def set_password(self, pw):
        self.pw_hash = generate_password_hash(pw)

    def check_password(self, pw):
        return check_password_hash(self.pw_hash, pw)

    deleted = db.Column( db.Boolean, default=False )
    deleted_at = db.Column( db.DateTime, nullable=True )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],
        backref=db.backref( "users", lazy=True,
            foreign_keys="User.institution_id" ) )
    # -------- Pending Institution Request --------
    requested_institution_name = db.Column( db.String(200), nullable=True )
    requested_institution_code = db.Column( db.String(50), nullable=True )
    requested_existing_institution_id = db.Column( db.Integer,
        db.ForeignKey("institution.id"), nullable=True )
    requested_existing_institution = db.relationship( "Institution",
        foreign_keys=[requested_existing_institution_id] )    
    phone = db.Column( db.String(50), nullable=True )
    is_global = db.Column( db.Boolean, default=False )
    rejected_by = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=True, index=True )
    rejected_at = db.Column( db.DateTime, nullable=True )
    approval_comment = db.Column( db.String(255), nullable=True )
    account_status = db.Column( db.String(20), default="active",
                            nullable=False, index=True )
    approved_by = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=True, index=True )
    approved_at = db.Column( db.DateTime, nullable=True )
    approved = db.Column( db.Boolean, default=False, index=True )
    approval_status = db.Column( db.String(20), default="Pending",
        nullable=False, index=True )
    # -----------------------------------------
    # Administrator Type
    # -----------------------------------------
    is_institution_admin = db.Column( db.Boolean, default=False,
        nullable=False )
    is_temporary_admin = db.Column( db.Boolean, default=False, nullable=False )
    replaced_admin_id = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=True )    
    

class SubjectGroup(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    # Group name
    name = db.Column(db.String(120), nullable=False)
    hide_scores_for_subadmins = db.Column( db.Boolean, default=False )
    # Parent group (NULL = top-level group)
    parent_id = db.Column( db.Integer, db.ForeignKey("subject_group.id") )
    parent = db.relationship( "SubjectGroup", remote_side=[id],
        backref=db.backref( "children", cascade="all, delete-orphan" ) )
    subjects = db.relationship( "Subject", backref="group", lazy=True )
    deleted = db.Column(db.Boolean, default=False)
    deleted_at = db.Column(db.DateTime, nullable=True)
    # Allow same names under different parents,
    # but prevent duplicates inside one parent.
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
       foreign_keys=[institution_id], backref=db.backref( "groups", lazy=True))    

class ChatMessage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sender_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    receiver_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    message = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    read = db.Column(db.Boolean, default=False)   # ✅ NEW
    # Relationships (optional)
    sender = db.relationship("User", foreign_keys=[sender_id])
    receiver = db.relationship("User", foreign_keys=[receiver_id])
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],
        backref=db.backref( "chat_messages", lazy=True ) )
    deleted = db.Column(db.Boolean, default=False, nullable=False)
    deleted_at = db.Column(db.DateTime, nullable=True)    

class License(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    key         = db.Column(db.String(64), unique=True, nullable=False)
    issued_at   = db.Column(db.DateTime, default=datetime.utcnow)  # when licenser made it
    activated_at= db.Column(db.DateTime, nullable=True)            # when student activates
    expires_at  = db.Column(db.DateTime, nullable=True)            # set on activation
    valid_days  = db.Column(db.Integer, default=365)               # duration from activation
    active      = db.Column(db.Boolean, default=True)
    device_id = db.Column(db.String(128))   # 🔒 lock to a device
    reset_count = db.Column(db.Integer, default=0)  # track resets
    email = db.Column(db.String(120), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # NEW field to track who generated the license
    generated_by = db.Column(db.String(20), nullable=False, default="licenser")
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    user        = db.relationship('User', backref='license', uselist=False)
    def __repr__(self):
        return f"<License {self.key} for {self.email}>"
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship("Institution", foreign_keys=[institution_id])    

class AutoLicenseConfig(db.Model):
    __tablename__ = "auto_license_config"
    id = db.Column(db.Integer, primary_key=True)
    remaining_count = db.Column(db.Integer, default=0)
    valid_days = db.Column(db.Integer, default=365)
    email = db.Column(db.String(120), nullable=False)
    enabled = db.Column(db.Boolean, default=False)
    created_at = db.Column( db.DateTime, default=datetime.utcnow )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )    

class AutoLicenseRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    request_key = db.Column( db.String(64), unique=True, nullable=False )
    count = db.Column(db.Integer, nullable=False)
    valid_days = db.Column( db.Integer, default=365 )
    email = db.Column( db.String(120), nullable=False )
    admin_email = db.Column( db.String(120), nullable=True )
    activated = db.Column( db.Boolean, default=False )
    created_at = db.Column( db.DateTime, default=datetime.utcnow )
    is_auto_license = db.Column( db.Boolean, default=False )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )    

# =========================================================
# SUBJECT MODEL
# =========================================================
class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    class_level = db.Column(db.Integer, nullable=True, index=True)
    # =========================================
    # LABEL SYSTEM
    # =========================================
    label = db.Column(db.String(120), nullable=True)
    # Examples:
    # 2025/2026 EXAM A 1ST
    # 2025/2026 CA B 2ND
    # 2025/2026 PRACTICE C 3RD
    result_type = db.Column( db.String(20), nullable=True )
    # exam | ca | practice
    duration_minutes = db.Column( db.Integer, default=60 )
    # NEW THEORY
    theory_duration_minutes = db.Column( db.Integer, default=30 )    
    hide_scores_for_subadmins = db.Column( db.Boolean, default=False )
    hidden_from_students = db.Column( db.Boolean, default=False )
    # =========================================
    # AUTO ANSWER REVIEW
    # =========================================
    # Practice exams automatically enable this.
    show_answers_to_students = db.Column( db.Boolean, default=False )
    group_id = db.Column( db.Integer, db.ForeignKey('subject_group.id'),
        nullable=True )
    series_group = db.Column(db.String(120), nullable=True)
    #Whenever a Subject is deleted, automatically delete all its Questions
    questions = db.relationship( 'Question', backref='subject', lazy=True,
        cascade="all, delete-orphan" )
    # =========================================
    # RECYCLE BIN
    # =========================================
    deleted = db.Column( db.Boolean, default=False )
    deleted_at = db.Column( db.DateTime, nullable=True )
    credit_unit = db.Column( db.Integer, default=1 ) #gpa
    session = db.Column( db.String(50), nullable=True )
    term = db.Column( db.String(20), nullable=True )
    semester = db.Column( db.String(20), nullable=True )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],
        backref=db.backref( "subjects", lazy=True ) )    

class Question(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    subject_id    = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    prompt        = db.Column(db.Text, nullable=False)
    choices_json  = db.Column(db.Text, default='[]')  # list of choices; empty list means open-ended
    answer        = db.Column(db.Text, nullable=True) # can be exact text, choice index (0-based), or free-text answer
    image         = db.Column(db.String(255), nullable=True)  # the uploaded image
    #ALTER TABLE question ADD COLUMN image VARCHAR(200); in SQLite
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],                                   
        backref=db.backref( "questions", lazy=True ) )    

class ExamSession(db.Model):
    id                 = db.Column(db.Integer, primary_key=True)
    user_id            = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subject_id         = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    started_at         = db.Column(db.DateTime, default=datetime.utcnow)
    submitted_at = db.Column(db.DateTime)
    completed          = db.Column(db.Boolean, default=False)
    paused             = db.Column(db.Boolean, default=False)
    remaining_seconds  = db.Column(db.Integer, default=0)
    flagged_json       = db.Column(db.Text, default="[]")
    current_index      = db.Column(db.Integer, default=0)
    progress_count     = db.Column(db.Integer, default=0)
    # Persist *fixed* question order for this session, and answers
    question_order_json = db.Column(db.Text, default='[]')   # list of question IDs
    answers_json        = db.Column(db.Text, default='{}')   # {question_id_str: "answer"}
    # Use JSON for history instead of PickleType (better change tracking)
    navigation_history_json = db.Column(db.Text, default='[]')  # list of visited indices, in order
    user    = db.relationship('User', lazy=True)
    subject = db.relationship('Subject', lazy=True)
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],
        backref=db.backref( "exam_sessions", lazy=True ) )    

class Result(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'),
        nullable=False)
    # ======================
    # ORIGINAL SCORES
    # ======================
    score = db.Column(db.Integer, nullable=False)
    theory_score = db.Column(db.Integer, nullable=True, default=0)
    total = db.Column(db.Integer, nullable=False)
    # ======================
    # NEW LABEL SYSTEM
    # ======================
    label = db.Column(db.String(120), nullable=True)
    # Exam or CA
    result_type = db.Column(db.String(20), nullable=True)
    # ======================
    # RECALCULATED SCORES
    # ======================
    converted_objective = db.Column(db.Float, default=0)
    converted_ca = db.Column(db.Float, default=0)
    final_score = db.Column(db.Float, default=0)
    # ======================
    # EXISTING FEATURES
    # ======================
    total_score = db.Column(db.Integer, nullable=True)
    taken_at = db.Column(db.DateTime, default=datetime.utcnow)
    session_id = db.Column(db.Integer, db.ForeignKey('exam_session.id'))
    hidden = db.Column(db.Boolean, default=False)
    deleted = db.Column(db.Boolean, default=False)
    deleted_at = db.Column(db.DateTime, nullable=True)
    grade = db.Column(db.String(2), nullable=True)
    remark = db.Column(db.String(255), nullable=True)
    teacher_name = db.Column(db.String(100), nullable=True)
    percentage = db.Column(db.Float, default=0)
    # ======================
    # RELATIONSHIPS
    # ======================
##    user = db.relationship('User', backref='results', lazy=True)
##    subject = db.relationship('Subject', lazy=True)
    user = db.relationship( "User", foreign_keys=[user_id], backref="results",
        lazy=True )
    subject = db.relationship( "Subject", lazy=True )
    
    __table_args__ = (db.UniqueConstraint( 'session_id', 
            name='uniq_result_per_session' ), )
    # ======================
    # Retake
    # ======================
    can_retake = db.Column( db.Boolean, default=False )
    # ==========================================
    # RETAKE AUDIT
    # ==========================================    
##    retake_count = db.Column( db.Integer, default=0 )
    retake_count = db.Column( db.Integer, nullable=False, default=0,
        server_default="0" )    
    retake_granted_by = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=True )
    retake_granted_at = db.Column( db.DateTime, nullable=True )
    retake_admin = db.relationship(
        "User",
        foreign_keys=[retake_granted_by],
        backref="results_retakes_granted" )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        backref=db.backref( "results", lazy=True ) )
    registration_type = db.Column( db.String(30), default="new_institution",
        nullable=False, index=True )    

class RetakeHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column( db.Integer, db.ForeignKey("user.id"), nullable=False )
    subject_id = db.Column( db.Integer, db.ForeignKey("subject.id"),
        nullable=False )
    result_id = db.Column( db.Integer, db.ForeignKey("result.id"),
                           nullable=True )
    old_score = db.Column(db.Integer)
    old_total = db.Column(db.Integer)
    new_score = db.Column(db.Integer)
    new_total = db.Column(db.Integer)
    old_theory_score = db.Column(db.Integer)
    new_theory_score = db.Column(db.Integer)
    granted_by = db.Column( db.Integer, db.ForeignKey("user.id"), nullable=True
    )
    granted_at = db.Column(db.DateTime)
    completed_at = db.Column(db.DateTime)
    attempt_number = db.Column( db.Integer, default=1 )
    user = db.relationship( "User", foreign_keys=[user_id] )
    admin = db.relationship( "User", foreign_keys=[granted_by] )
    subject = db.relationship("Subject")
    result = db.relationship("Result")
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        backref=db.backref( "retake_history", lazy=True ) )    

class SchoolSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    # =================================
    # SCHOOL INFO
    # =================================
    name = db.Column(db.String(200))
    address = db.Column(db.String(300))
    phone = db.Column(db.String(50))
    term = db.Column(db.String(50))
    session = db.Column(db.String(50))
    next_term_begins = db.Column( db.String(100) )
    # =================================
    # REPORT CARD FILES
    # =================================
    logo_path = db.Column( db.String(300), nullable=True )
    background_path = db.Column( db.String(300), nullable=True )
    principal_signature_path = db.Column( db.String(300), nullable=True )
    # =====================================
    # REPORT CONFIGURATION
    # =====================================
    # Institution type:
    # secondary
    # university
    # training
    academic_mode = db.Column( db.String(50), default="secondary" )
    # Grading scale:
    # percentage
    # gpa4
    # cgpa5
    grading_scale = db.Column( db.String(30), default="percentage" )
    default_report_design = db.Column( db.String(100), default="standard.html")    
    cumulative_across_sessions = db.Column( db.Boolean, default=True )
    # =====================================
    # INSTITUTION
    # =====================================
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],
        backref=db.backref( "school_settings", lazy=True ) )
    # ==========================
    # REPORT CARD SETTINGS
    # ==========================
    ca1_percentage = db.Column(db.Float, default=10)
    ca2_percentage = db.Column(db.Float, default=10)
    ca3_percentage = db.Column(db.Float, default=10)
    objective_percentage = db.Column(db.Float, default=40)
    theory_percentage = db.Column(db.Float, default=30)
    # =====================================
    # REPORT DISPLAY OPTIONS
    # =====================================
    show_student_position = db.Column( db.Boolean, default=True, nullable=False)
    show_class_average = db.Column( db.Boolean, default=True, nullable=False )
    show_gpa = db.Column( db.Boolean, default=False, nullable=False )
    show_cgpa = db.Column( db.Boolean, default=False, nullable=False )
    show_class_size = db.Column( db.Boolean, default=True, nullable=False )
    show_attendance = db.Column( db.Boolean, default=True, nullable=False )
    show_psychomotor = db.Column( db.Boolean, default=True, nullable=False )
    show_teacher_comment = db.Column( db.Boolean, default=True, nullable=False)
    show_principal_comment = db.Column( db.Boolean, default=True,
        nullable=False )
    show_next_term = db.Column( db.Boolean, default=True, nullable=False )    
    #====Publication status for the current term/session====
    report_cards_published = db.Column( db.Boolean, default=False )
    report_cards_published_at = db.Column( db.DateTime )
    report_cards_published_by = db.Column( db.Integer, db.ForeignKey("user.id"))

class ReportCardAccessLink(db.Model):
    __tablename__ = "report_card_access_link"
    id = db.Column(db.Integer, primary_key=True)
    institution_id = db.Column( db.Integer,
        db.ForeignKey("institution.id"), nullable=False, index=True )
    token = db.Column( db.String(128), unique=True, nullable=False, index=True )
    role_type = db.Column( db.String(20), nullable=False )#teacher / principal
    description = db.Column( db.String(200) )
                # e.g.
                # "JSS Teachers"
                # "Primary Section"
                # "Science Department"    
    created_by = db.Column( db.Integer, db.ForeignKey("user.id") )
    expires_at = db.Column( db.DateTime, nullable=False )
    is_active = db.Column( db.Boolean, default=True )
    created_at = db.Column( db.DateTime, default=datetime.utcnow )

class ReportCardRemark(db.Model):
    __tablename__ = "report_card_remark"
    id = db.Column(db.Integer, primary_key=True)
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=False, index=True )
    student_id = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=False, index=True )
    term = db.Column( db.String(30), nullable=False )
    session = db.Column( db.String(30), nullable=False )
    teacher_name = db.Column( db.String(150) )
    days_open = db.Column( db.Integer, default=0 )
    days_present = db.Column( db.Integer, default=0 )
    punctuality = db.Column( db.String(100) )
    honesty = db.Column( db.String(100) )
    neatness = db.Column( db.String(100) )
    leadership = db.Column( db.String(100) )
    handwriting = db.Column( db.String(100) )
    sports = db.Column( db.String(100) )
    class_teacher_comment = db.Column( db.Text )
    principal_comment = db.Column( db.Text )
    teacher_link_id = db.Column( db.Integer,
        db.ForeignKey("report_card_access_link.id") )
    principal_link_id = db.Column( db.Integer,
        db.ForeignKey("report_card_access_link.id") )
    updated_at = db.Column( db.DateTime, default=datetime.utcnow,
        onupdate=datetime.utcnow )
    report_batch = db.Column( db.String(50), nullable=False )
    academic_class_id = db.Column( db.Integer,
        db.ForeignKey("academic_class.id"), index=True )
    report_design = db.Column( db.String(100) )
    completed_by = db.Column( db.String(150) )
    teacher_completed_at = db.Column( db.DateTime )
    teacher_updated_at = db.Column(db.DateTime)
    principal_updated_at = db.Column(db.DateTime)
    teacher_locked = db.Column( db.Boolean, default=False )
    principal_locked = db.Column( db.Boolean, default=False )
    teacher_locked_at = db.Column( db.DateTime )
    principal_locked_at = db.Column( db.DateTime )
    teacher_unlocked_at = db.Column( db.DateTime )
    principal_unlocked_at = db.Column( db.DateTime )
    teacher_unlocked_by = db.Column( db.Integer, db.ForeignKey("user.id") )
    principal_unlocked_by = db.Column( db.Integer, db.ForeignKey("user.id") )
    frozen_data = db.Column(db.Text)

class ReportCardAutoComment(db.Model):
    __tablename__ = "report_card_auto_comment"
    id = db.Column(db.Integer, primary_key=True)
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=False, index=True )
    role_type = db.Column( db.String(20), nullable=False )# teacher / principal
    minimum_score = db.Column( db.Float, nullable=False )
    maximum_score = db.Column( db.Float, nullable=False )
    comment = db.Column( db.Text, nullable=False )
    created_at = db.Column( db.DateTime, default=datetime.utcnow )

class ReportCardAudit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=False )
    report_card_remark_id = db.Column( db.Integer,
        db.ForeignKey("report_card_remark.id"), nullable=False )
    student_id = db.Column( db.Integer, db.ForeignKey("user.id"),
        nullable=False )
    role_type = db.Column( db.String(20) )
    actor_name = db.Column( db.String(150) )
    action = db.Column( db.String(50) )
    created_at = db.Column( db.DateTime, default=datetime.utcnow )
    details = db.Column( db.Text )    

#==========GPA HISTORY MODEL==========
class StudentTermRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column( db.Integer, db.ForeignKey('user.id') )
    session = db.Column(db.String(50))
    term = db.Column(db.String(50))
    # Example:
    # A / B / C
    # First Semester
    # Second Semester
    gpa = db.Column(db.Float, default=0)
    cgpa = db.Column(db.Float, default=0)
    total_points = db.Column(db.Float, default=0)
    total_units = db.Column(db.Float, default=0)
    grading_scale = db.Column( db.String(50) )
    created_at = db.Column( db.DateTime, default=datetime.utcnow )
    student = db.relationship( 'User', backref='term_records' )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],
        backref=db.backref( "student_term_records", lazy=True ) )

class StudentRemark(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column( db.Integer, db.ForeignKey('user.id') )
    session = db.Column(db.String(50))
    term = db.Column(db.String(50))
    class_teacher_comment = db.Column( db.Text )
    principal_comment = db.Column( db.Text )
    punctuality = db.Column(db.Integer, default=0)
    honesty = db.Column(db.Integer, default=0)
    neatness = db.Column(db.Integer, default=0)
    leadership = db.Column(db.Integer, default=0)
    handwriting = db.Column(db.Integer, default=0)
    sports = db.Column(db.Integer, default=0)
    student = db.relationship( "User", foreign_keys=[student_id],
        backref=db.backref( "remarks", lazy=True ) )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],
        backref=db.backref( "student_remarks", lazy=True ) )    

#============PRINCIPAL COMMENT MODEL===============
class AutoComment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    role_type = db.Column(db.String(20))
    # principal
    # teacher
    owner_id = db.Column( db.Integer, db.ForeignKey('user.id'),
        nullable=True )
    owner = db.relationship( "User", foreign_keys=[owner_id] )    
    term = db.Column(db.String(50))
    # A / B / C
    # First Semester
    min_gpa = db.Column(db.Float)
    max_gpa = db.Column(db.Float)
    comments = db.Column(db.Text)
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],
        backref=db.backref( "auto_comments", lazy=True ) )    
    # Excellent work:Keep it up:Outstanding
    created_at = db.Column( db.DateTime, default=datetime.utcnow )

class TheorySubmission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'))
    file_path = db.Column(db.String(255))  # uploaded answer file
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    score = db.Column(db.Integer, nullable=True)  # admin will fill this
    marked_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    # Relationships
    user = db.relationship("User", foreign_keys=[user_id], backref="submissions")
    subject = db.relationship("Subject")
    marker = db.relationship("User", foreign_keys=[marked_by], backref="marked_submissions")
    answers_json = db.Column(db.Text)  # store submitted answers as JSON string
    # ✅ ADD THESE
    deleted = db.Column(db.Boolean, default=False)
    deleted_at = db.Column(db.DateTime, nullable=True)
    started_at = db.Column(db.DateTime, nullable=True)
    remaining_seconds = db.Column(db.Integer, default=0)
    completed = db.Column(db.Boolean, default=False)
    # ==========================================
    # RETAKE
    # ==========================================    
    can_retake = db.Column( db.Boolean, default=False )
    # ==========================================
    # RETAKE AUDIT
    # ==========================================
    retake_granted_by = db.Column( db.Integer, db.ForeignKey('user.id'),
        nullable=True )
    retake_granted_at = db.Column( db.DateTime, nullable=True )
##    retake_count = db.Column( db.Integer, default=0 )
    retake_count = db.Column( db.Integer, nullable=False, default=0,
        server_default="0" )
    retake_admin = db.relationship(
        "User",
        foreign_keys=[retake_granted_by],
        backref="theory_submissions_retakes_granted" )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        backref=db.backref( "theory_submissions", lazy=True ) )    

class TheoryQuestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    prompt = db.Column(db.Text, nullable=False)
    image = db.Column(db.String(255), nullable=True)
    answer = db.Column(db.Text, nullable=True)
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True, index=True )
    institution = db.relationship( "Institution",
        foreign_keys=[institution_id],
        backref=db.backref( "theory_questions", lazy=True ) )    

# =========================================================
# APP SETTINGS MODEL
# =========================================================
class InstitutionSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    institution_id = db.Column(
        db.Integer,
        db.ForeignKey("institution.id"),
        nullable=False,
        unique=True,
        index=True,
    )

    smtp_provider = db.Column(db.String(30), default="custom")
    smtp_server = db.Column(db.String(255))
    smtp_port = db.Column(db.String(10), default="587")
    smtp_username = db.Column(db.Text)
    smtp_password = db.Column(db.Text)
    smtp_use_tls = db.Column(db.Boolean, default=True)

    auto_forward_scores = db.Column(db.Boolean, default=False)

    institution = db.relationship(
        "Institution",
        backref=db.backref(
            "settings",
            uselist=False,
            cascade="all, delete-orphan",
        ),
    )
    admin_email = db.Column(db.Text, nullable=True)

class AppSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True)
    value = db.Column(db.Text)
    
#=============EMAIL QUEUE TABLE=============
class EmailQueue(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    recipient = db.Column(db.String(255))
    subject = db.Column(db.String(255))
    body = db.Column(db.Text)
    status = db.Column(db.String(20), default="pending")  # pending, sent, failed, dead
    error = db.Column(db.Text, nullable=True)
    attempts = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    sent_at = db.Column(db.DateTime, nullable=True)
    mail_profile = db.Column( db.String(20), default="admin" )
    institution_id = db.Column( db.Integer, db.ForeignKey("institution.id"),
        nullable=True )    

# --------Model (store global pagination in DB) --------#
class AppConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.String(50), nullable=False)

    @staticmethod
    def get(key, default=None):
        rec = AppConfig.query.filter_by(key=key).first()
        return rec.value if rec else default

    @staticmethod
    def set(key, value):
        rec = AppConfig.query.filter_by(key=key).first()
        if not rec:
            rec = AppConfig(key=key, value=str(value))
            db.session.add(rec)
        else:
            rec.value = str(value)
        db.session.commit()
#=====Admin/Licenser Account Reset and Restore=====
class EmergencyRecoveryToken(db.Model):
    id = db.Column( db.Integer, primary_key=True )
    token = db.Column( db.String(128), unique=True, nullable=False )
    created_at = db.Column( db.DateTime, default=datetime.utcnow )
    expires_at = db.Column( db.DateTime, nullable=False )
    used = db.Column( db.Boolean, default=False )
    used_at = db.Column( db.DateTime )
    requested_ip = db.Column( db.String(100) )
    requested_browser = db.Column( db.Text ) 

# ---------- Helpers / Decorators ----------
#===Generate Secure Access Token===
def generate_report_card_token():
    """Generate a secure URL token."""
    return secrets.token_urlsafe(48)

#===Create Teacher/Principal Link===
def create_report_card_link(role_type, description=""):
    """
    Creates a 24-hour report card access link.
    role_type = 'teacher' or 'principal'
    """
    token = generate_report_card_token()
    row = ReportCardAccessLink(
        institution_id=current_institution_id(),
        token=token,
        role_type=role_type,
        description=description,
        created_by=get_current_user().id,
        expires_at=datetime.utcnow() + timedelta(hours=24),
        is_active=True
    )
    db.session.add(row)
    db.session.commit()

    return row

#====Validate Token====
def get_valid_report_card_link(token):
    row = (
        filter_by_institution(
            ReportCardAccessLink.query,
            ReportCardAccessLink
        )
        .filter_by(
            token=token,
            is_active=True
        )
        .first()
    )
    if not row:
        return None
    if row.expires_at < datetime.utcnow():
        return None

    return row

#===Deactivate Link===
def deactivate_report_card_link(link_id):
    row = get_record_in_scope(
        ReportCardAccessLink,
        link_id
    )
    row.is_active = False
    db.session.commit()







def create_emergency_recovery_token(request):
    """
    Generates a one-time emergency recovery token.
    """
    token = secrets.token_urlsafe(48)
    recovery = EmergencyRecoveryToken(
        token=token,
        expires_at=datetime.utcnow() + timedelta(hours=24),
        requested_ip=request.remote_addr,
        requested_browser=request.headers.get("User-Agent")
    )
    db.session.add(recovery)
    db.session.commit()
    return recovery

def record_student_class_history(
    student,
    academic_class,
    session,
    term,
    change_type="Promotion",
    remarks=None,
    changed_by=None
):
    history = StudentClassHistory(
        institution_id=student.institution_id,
        student_id=student.id,
        academic_class_id=academic_class.id,
        session=session,
        term=term,
        change_type=change_type,
        remarks=remarks,
        changed_by=changed_by
    )

    db.session.add(history)

def can_login(user):
    """
    Returns:
        (True, None)
            User may log in.
        (False, message)
            User cannot log in.
    """
    if user is None:
        return ( False, "Invalid username or password." )
    if user.deleted:
        return ( False, "This account no longer exists." )
    if user.account_status != "active":
        return ( False, "This account is currently inactive." )
    # Global Admin and Global Licenser
    # always bypass approval.
    if (
        user.is_global
        and
        user.role in ( "admin", "licensing" ) ):
        return ( True, None )
    if not user.approved:
        if user.approval_status == "Pending":
            return (
                False,
                "Your account is awaiting approval. "
                "You may contact support using the channels below."
            )
        if user.approval_status == "Rejected":
            return (
                False,
                "Your registration request was not approved. "
                "Please contact support."
            )
        if user.approval_status == "Suspended":
            return (
                False,
                "Your account has been suspended. "
                "Please contact support."
            )
        return ( False, "Your account cannot be used." )
##        if institution is disabled:
##            pass
##        if license expired:
##            pass
##        if server disabled account:
##            pass
    return (
        True,
        None
    )

def get_next_academic_class(current_class):
    if not current_class:
        return None

    return (
        filter_by_institution( AcademicClass.query, AcademicClass )
        .filter( AcademicClass.display_order > current_class.display_order )
        .order_by( AcademicClass.display_order )
        .first() )

def get_previous_academic_class(current_class):
    if not current_class:
        return None

    return (
        filter_by_institution( AcademicClass.query, AcademicClass )
        .filter( AcademicClass.display_order < current_class.display_order )
        .order_by( AcademicClass.display_order.desc() )
        .first() )

def get_academic_class_by_code(class_code):
    if class_code is None:
        return None

    return ( filter_by_institution( AcademicClass.query, AcademicClass )
        .filter_by( class_code=class_code )
        .first() )

#===============Promotion Helper==============
def promote_student( student, session, term, changed_by=None, remarks=None ):
    """
    Promote a student to the next academic class.
    Returns:
        (True, message) on success
        (False, reason) if promotion is not possible.
    """
    if not student.academic_class:
        return False, "Student has no academic class."
    next_class = get_next_academic_class( student.academic_class )
    if next_class is None:
        return False, "Student is already in the highest class."
    previous_class = student.academic_class
    student.academic_class_id = next_class.id
    # Keep legacy field synchronized
    student.student_class = next_class.class_code
    record_student_class_history(
        student=student,
        academic_class=next_class,
        session=session,
        term=term,
        change_type="Promotion",
        remarks=remarks,
        changed_by=changed_by
    )
    return ( True, f"{previous_class.class_name} → {next_class.class_name}" )

#==============Transfer Helper===============
def transfer_student( student, destination_class, session, term,
    changed_by=None, remarks=None ):
    """
    Transfer a student to another academic class.
    """
    if not student:
        return False, "Invalid student."
    if not destination_class:
        return False, "Destination class not found."
    if student.institution_id != destination_class.institution_id:
        return False, "Cross-institution transfer is not allowed."    
    if student.academic_class_id == destination_class.id:
        return False, "Student is already in that class."
    student.academic_class_id = destination_class.id
    # Keep legacy field synchronized
    student.student_class = destination_class.class_code
    record_student_class_history( student=student,
        academic_class=destination_class, session=session, term=term,
        change_type="Transfer", remarks=remarks, changed_by=changed_by )
    return True, "Transferred"

#=====Keep both class fields in sync, can remove student_class later=====
def sync_student_academic_classes():
    """
    Populate academic_class_id for existing students
    using their current student_class value.
    Safe to run multiple times.
    """
    students = User.query.filter( User.student_class.isnot(None) ).all()
    updated = 0
    for student in students:
        if student.academic_class_id:
            continue
        academic_class = ( AcademicClass.query
            .filter_by(
                institution_id=student.institution_id,
                class_code=student.student_class )
            .first() )
        if academic_class:
            student.academic_class_id = academic_class.id
            updated += 1
    if updated:
        db.session.commit()
    return updated

def sync_student_class_history():
    students = User.query.filter(
        User.academic_class_id.isnot(None)
    ).all()
    added = 0
    for student in students:
        exists = (
            StudentClassHistory.query.filter_by(
                student_id=student.id
            ).first()
        )
        if exists:
            continue
        settings = (
            SchoolSettings.query.filter_by(
                institution_id=student.institution_id
            ).first()
        )
        if not student.academic_class:
            continue
        record_student_class_history(
            student=student,
            academic_class=student.academic_class,
            session=settings.current_session if settings else "",
            term=settings.current_term if settings else "",
            change_type="Migration",
            remarks="Automatically migrated",
            changed_by=None
        )
        added += 1
    if added:
        db.session.commit()
    return added

def create_default_academic_classes(institution_id):
    defaults = [
        (1, "JSS1"), (2, "JSS2"), (3, "JSS3"), (4, "SSS1"), (5, "SSS2"),
        (6, "SSS3"), ]
    for code, name in defaults:
        exists = (
            AcademicClass.query.filter_by(
                institution_id=institution_id, class_code=code ).first() )
        if exists:
            continue
        db.session.add(
            AcademicClass(
                institution_id=institution_id, class_code=code,
                class_name=name, education_level="Secondary",
                display_order=code, active=True ) )
    db.session.commit()

def get_academic_classes( active_only=True ):
    query = filter_by_institution( AcademicClass.query, AcademicClass )
    if active_only:
        query = query.filter_by( active=True )
    return ( query
        .order_by( AcademicClass.display_order, AcademicClass.class_name )
        .all() )

def get_students_in_academic_class( academic_class_id, include_deleted=False):
    """
    Returns students in an academic class
    belonging to the current institution.
    """
    query = ( filter_by_institution( User.query, User )
        .filter( User.role == "student",
            User.academic_class_id == academic_class_id ) )
    if not include_deleted:
        query = query.filter( User.deleted == False )
    return ( query
        .order_by( User.last_name, User.first_name, User.username )
        .all() )

def get_student_class_history(student_id):
    return (
        filter_by_institution( StudentClassHistory.query,
            StudentClassHistory )
        .filter_by( student_id=student_id )
        .order_by( StudentClassHistory.changed_at.desc() )
        .all() )

def get_academic_class_summary(academic_class_id):
    """
    Returns summary information for an academic class.
    """
    students = get_students_in_academic_class( academic_class_id )
    males = 0
    females = 0
    for student in students:
        gender = getattr( student, "gender", None )
        if gender == "Male":
            males += 1
        elif gender == "Female":
            females += 1
    return {
        "student_count": len(students),
        "male_count": males,
        "female_count": females }

def get_current_institution_id():
    user = get_current_user()

    if not user:
        return None

    if user.role == "global_admin":
        free = Institution.query.filter_by(is_system=True).first()
        return free.id if free else None

    return user.institution_id

#======Institutionsettings helper=====
def get_institution_settings():
    institution_id = current_settings_institution_id()
    if institution_id is None:
        raise RuntimeError( "Unable to determine institution settings." )
    settings = InstitutionSettings.query.filter_by(
        institution_id=institution_id ).first()
    if settings is None:
        settings = InstitutionSettings( institution_id=institution_id )
        db.session.add(settings)
        db.session.commit()

    return settings

#===InstitutionSettings helper for ownership of newly created records===
def current_settings_institution_id():
    """
    Returns the institution whose settings should be edited.
    Global admin edits the Free Institution settings.
    """
    user = get_current_user()
    if not user:
        return None
    if is_global_admin(user):
        free = Institution.query.filter_by(is_system=True).first()
        return free.id if free else None

    return user.institution_id

# =========================================================
# GET RECORD WITH INSTITUTION CHECK
# =========================================================
def get_record_in_scope(model, record_id):

    user = get_current_user()

    query = model.query.filter_by(id=record_id)

    # Ignore deleted records where applicable
    if hasattr(model, "deleted"):
        query = query.filter(model.deleted == False)

    record = query.first_or_404()

    # Super Admin can access everything
    if is_super_admin(user):
        return record

    # Direct institution ownership
    if hasattr(record, "institution_id"):
        if record.institution_id != user.institution_id:
            abort(404)
        return record

    # Models that belong to a Subject
    if hasattr(record, "subject"):
        if (
            record.subject
            and record.subject.institution_id != user.institution_id
        ):
            abort(404)
        return record

    # Models that belong to an Exam Session
    if hasattr(record, "exam_session"):
        if (
            record.exam_session
            and record.exam_session.institution_id != user.institution_id
        ):
            abort(404)
        return record

    # Models that belong to a User
    if hasattr(record, "user"):
        if (
            record.user
            and record.user.institution_id != user.institution_id
        ):
            abort(404)
        return record

    return record

def is_super_admin(user):
    return user and user.role == "superadmin"

def filter_by_institution(query, model):
    """
    Restrict queries to the current institution.
    Global Admin -> Free institution only
    Institution Admin -> Own institution
    Licenser -> All licenses only
    """
    user = get_current_user()
    if not user:
        return query.filter(False)
    # ----------------------------------
    # Global Licenser
    # ----------------------------------
    if user.role == "licensing" and user.is_global:
        if model == License:
            return query          # Can see every license
        return query.filter(False)
    # ----------------------------------
    # Determine institution
    # ----------------------------------
    institution_id = get_current_institution_id()
    if institution_id is None:
        return query.filter(False)
    if hasattr(model, "institution_id"):
        return query.filter(
            model.institution_id == institution_id
        )
    return query

def institution_filter(model):
    """
    Returns a filter expression that can be used inside
    query.filter(...)
    """

    user = get_current_user()

    if not user:
        return True

    if is_super_admin(user):
        return True

    if hasattr(model, "institution_id"):
        return model.institution_id == user.institution_id

    return True

# =========================================================
# CURRENT USER INSTITUTION
# =========================================================
def get_current_institution():
    user = get_current_user()
    if not user:
        return None
    if not user.institution_id:
        return None
    return Institution.query.get(
        user.institution_id
    )

# =========================================================
# FILTER HELPERS
# =========================================================
def institution_users_query():
    user = get_current_user()
    return User.query.filter(
        User.deleted == False,
        User.institution_id == user.institution_id )

def institution_subjects_query():
    user = get_current_user()
    return Subject.query.filter(
        Subject.deleted == False,
        Subject.institution_id == user.institution_id )

def institution_groups_query():
    user = get_current_user()
    return SubjectGroup.query.filter(
        SubjectGroup.deleted == False,
        SubjectGroup.institution_id == user.institution_id )

#=============Format Institution Name normally===========
def normalize_institution_name(name):
    """
    Produces a consistent institution name.
    Examples
    olivet    baptist academy, oyo, ibadan
    becomes
    Olivet Baptist Academy, Oyo, Ibadan
    """
    if not name:
        return None
    # Remove leading/trailing spaces
    name = name.strip()
    # Collapse multiple whitespace
    name = re.sub(r"\s+", " ", name)
    # Remove spaces before commas
    name = re.sub(r"\s+,", ",", name)
    # Ensure one space after commas
    name = re.sub(r",\s*", ", ", name)
    # Capitalize every word
    name = name.title()
    return name

def find_existing_institution(name):
    """
    Returns an existing institution after normalization,
    or None if not found.
    """
    normalized = normalize_institution_name(name)
    if not normalized:
        return None
    return (
        Institution.query
        .filter(func.lower(Institution.name) == normalized.lower()
        ).first())

#=========Generate Temporary Administrators=======
##activate_institution_admin()
##resolve_requested_institution()
# =========================================================
# CREATE TEMPORARY INSTITUTION ADMINISTRATOR
# =========================================================
def create_temporary_institution_admin(original_admin):
    """
    Creates a temporary Institution Administrator when the
    original administrator is deleted.
    The temporary administrator keeps the institution
    operational until the original administrator is restored
    or another administrator is assigned.
    """
    if original_admin is None:
        raise ValueError(
            "Original administrator cannot be None."
        )
    if not original_admin.is_institution_admin:
        raise ValueError(
            "User is not an Institution Administrator."
        )
    # -----------------------------------------
    # Extract first name
    # -----------------------------------------
    first_name = (
        original_admin.first_name or "Admin"
    )
    # -----------------------------------------
    # Generate unique username
    # -----------------------------------------
    username = (
        f"temp_admin_{original_admin.id}"
    )
    counter = 1
    while User.query.filter_by(
        username=username
    ).first():
        username = (
            f"temp_admin_{original_admin.id}_{counter}"
        )
        counter += 1
    # -----------------------------------------
    # Create temporary administrator
    # -----------------------------------------
    temp_admin = User(
        first_name="Temporary",
        last_name=f"Admin {first_name}",
        username=username,
        password=original_admin.password,
        email=None,
        phone=None,
        role="admin",
        approved=True,
        approval_status="Approved",
        account_status="active",
        institution_id=original_admin.institution_id,
        approved_by=original_admin.approved_by,
        approved_at=datetime.utcnow(),
        is_institution_admin=True,
        is_temporary_admin=True,
        replaced_admin_id=original_admin.id
    )
    db.session.add(temp_admin)
    db.session.flush()
    return temp_admin

# =========================================================
# ENSURE DEFAULT INSTITUTION EXISTS
# =========================================================
def ensure_default_institution():
    free = Institution.query.filter_by( code="FREE" ).first()
    if free:
        return free
    # fallback in case code is missing
    free = Institution.query.filter_by( name="Free" ).first()
    if free:
        free.code = "FREE"
        free.active = True
        free.deleted = False
        free.deleted_at = None
        db.session.commit()
        print("✅ Default Free institution repaired.")
        return free
    # create if missing
    free = Institution(
        name="Free",
        code="FREE",
        description="Default institution for users that are not attached to any school.",
        active=True,
        deleted=False
    )
    db.session.add(free)
    db.session.commit()
    print("✅ Default Free institution created.")
    return free

def ensure_admin_created():
    # ==========================================
    # Ensure default admin exists
    # ==========================================
    free = Institution.query.filter_by( code="FREE" ).first()
    admin = User.query.filter_by( username="admin" ).first()
    if admin is None:
        admin = User(
            username="admin",
            password=generate_password_hash("admin123"),
            role="admin",
            institution_id=free.id,
            active=True,
            account_status="active",
            approved=True,
            is_global=True,
            approval_status = "Approved",
            approved_at = datetime.utcnow()
        )
        db.session.add(admin)
        db.session.commit()
        print("✅ Default admin created.")
    else:
        changed = False
        if admin.institution_id is None:
            admin.institution_id = free.id
            changed = True
        if changed:
            db.session.commit()
            print("✅ Default admin updated.")

# =====================================================
# Ensure Free Institution Exists
# =====================================================
def ensure_free_institution():
    institution = Institution.query.filter_by( code="FREE" ).first()
    if institution:
        if not institution.is_system:
            institution.is_system = True
            db.session.commit()
        return institution
    institution = Institution(
        name="Free Institution",
        code="FREE",
        description="Default institution for students not attached to any school.",
        active=True,
        is_system=True )
    db.session.add(institution)
    db.session.commit()
    print("✅ Free Institution created.")
    return institution

def migrate_null_records_to_free():
    free = Institution.query.filter_by(code="FREE").first()
    if not free:
        return
    free_id = free.id
    MODELS = [
        User,
        Subject,
        SubjectGroup,
        Question,
        TheoryQuestion,
        Result,
        TheorySubmission,
        ExamSession,
        ChatMessage,
        AutoComment,
        StudentRemark,
        StudentTermRecord,
        AcademicClass,
        StudentClassHistory,
        SchoolSettings,
        RetakeHistory,
        License,
        AutoLicenseConfig,
    ]
    for model in MODELS:
        ( model.query
            .filter(model.institution_id == None)
            .update( {"institution_id": free_id}, synchronize_session=False))
    db.session.commit()


# =========================================================
# RESOLVE REQUESTED INSTITUTION
# =========================================================
def resolve_requested_institution( user, manual_institution_id=None):
    """
    Determines which institution an Institution Administrator
    should belong to.
    Supports three registration cases:
        1. Existing Institution selected
        2. Requested New Institution
        3. No Institution specified
    """
    # ------------------------------------------------------
    # CASE 1
    # Existing Institution selected
    # ------------------------------------------------------
    if user.requested_existing_institution_id:
        institution = Institution.query.filter_by(
            id=user.requested_existing_institution_id,
            deleted=False
        ).first()
        if institution is None:
            raise ValueError(
                "The selected institution no longer exists."
            )
        return institution
    # ------------------------------------------------------
    # CASE 2
    # Requested New Institution
    # ------------------------------------------------------
    if user.requested_institution_name:
        normalized_name = normalize_institution_name(
            user.requested_institution_name
        )
        institution = find_existing_institution(
            normalized_name
        )
        if institution:
            return institution
        institution = Institution(
            name=normalized_name,
            code=user.requested_institution_code,
            active=True
        )
        db.session.add(institution)
        db.session.flush()
        create_default_academic_classes(
            institution.id
        )
        return institution
    # ------------------------------------------------------
    # CASE 3
    # Manual Assignment
    # ------------------------------------------------------
    if not manual_institution_id:
        raise ValueError(
            "Please select an institution."
        )
    institution = Institution.query.filter_by(
        id=manual_institution_id,
        deleted=False
    ).first()
    if institution is None:
        raise ValueError(
            "Institution not found."
        )
    return institution

# =========================================================
# ACTIVATE INSTITUTION ADMINISTRATOR
# =========================================================
def activate_institution_admin( user, institution, approved_by=None):
    """
    Activates an Institution Administrator after approval.
    This helper centralizes all Institution Admin activation
    logic so every workflow (approval, restore, replacement,
    transfer, etc.) behaves consistently.
    """
    if user is None:
        raise ValueError("User cannot be None.")
    if institution is None:
        raise ValueError("Institution cannot be None.")
    # -----------------------------------------
    # Assign Institution
    # -----------------------------------------
    user.institution_id = institution.id
    # -----------------------------------------
    # Convert pending Institution Administrator
    # into an active Institution Administrator.
    # (Current permissions still use 'admin')
    # -----------------------------------------
    user.role = "admin"
    user.is_institution_admin = True
    user.is_temporary_admin = False
    user.replaced_admin_id = None    
    # -----------------------------------------
    # Approval Information
    # -----------------------------------------
    user.approved = True
    user.approval_status = "Approved"
    user.account_status = "active"
    user.approved_by = (
        approved_by.id
        if hasattr(approved_by, "id")
        else approved_by
    )
    user.approved_at = datetime.utcnow()
    # -----------------------------------------
    # Clear Pending Request Information
    # -----------------------------------------
    user.requested_institution_name = None
    user.requested_institution_code = None
    user.requested_existing_institution_id = None
    return user


# =====================================================
# Institution Helpers
# =====================================================
def current_institution_id():
    """
    Returns the institution that should own newly-created records.
    Global admin returns None so their records belong to the
    Free/global area.
    """
    user = get_current_user()
    if not user:
        return None
    if is_global_admin(user):
        return None
    return user.institution_id

##def is_global_admin():
##    user = get_current_user()
##    return bool(user and user.role == "global_admin")
def is_global_admin(user=None):
    if user is None:
        user = get_current_user()
    return bool( user and user.is_global )

def restore_default_admin():
    """
    Restore the built-in admin account to its default configuration.
    Existing admin ID is preserved whenever possible.
    """

    admin = User.query.filter_by(
        username="admin",
        role="admin"
    ).first()

    if admin is None:

        admin = User(
            username="admin",
            role="admin"
        )

        db.session.add(admin)

    # Restore default values

    admin.username = "admin"
    admin.role = "admin"
    admin.first_name = None
    admin.last_name = None
    admin.email = None

    admin.reset_token = None
    admin.reset_used = False

    admin.deleted = False
    admin.deleted_at = None

    admin.set_password("admin123")

##    db.session.commit()

    return admin


def restore_default_licenser():
    """
    Restore the built-in licenser account to its default configuration.
    Existing licenser ID is preserved whenever possible.
    """

    licenser = User.query.filter_by( username="licenser", role="licensing"
    ).first()
    if licenser is None:
        licenser = User(
            username="licenser",
            role="licensing",
            account_status = "active",
            approved=True,
            is_global=True,
            approval_status = "Approved",
            approved_at = datetime.utcnow()
        )

        db.session.add(licenser)

    # =====================================
    # Restore default values
    # =====================================

    licenser.username = "licenser"
    licenser.role = "licensing"

    licenser.first_name = None
    licenser.last_name = None
    licenser.email = None

    licenser.reset_token = None
    licenser.reset_used = False

    licenser.deleted = False
    licenser.deleted_at = None

    licenser.set_password("mysecretlicensepwd")

##    db.session.commit()

    return licenser


# -----check the license only when the student starts an exam ------
def check_license(user: User):
    """Check if a student's license is valid."""
##    lic = License.query.filter_by(user_id=user.id, active=True).first()
    lic = ( filter_by_institution( License.query, License )
        .filter_by( user_id=user.id, active=True )
        .first() )    

    if not lic:
        config = AutoLicenseConfig.query.first()
        if ( config and config.enabled and config.remaining_count > 0 ):
            # Create automatic license
            key = secrets.token_hex(16)
            lic = License(
                key=key,
                email=config.email,
                institution_id=current_institution_id(),
                valid_days=config.valid_days,
                generated_by="auto-system",
                user_id=user.id,
                activated_at=datetime.utcnow(),
                expires_at=datetime.utcnow() + timedelta(
                    days=config.valid_days
                )
            )
            db.session.add(lic)
            config.remaining_count -= 1
            if config.remaining_count <= 0:
                config.enabled = False
            db.session.commit()
            # Reload the newly created license
            lic = License.query.filter_by( user_id=user.id, active=True
            ).first()
            lic = ( filter_by_institution( License.query, License )
                .filter_by( user_id=user.id, active=True )
                .first() )            
        if not lic:
            return False, "No license found. Please activate."
    if not lic.activated_at or not lic.expires_at:
        return False, "License not activated properly."
    if datetime.utcnow() > lic.expires_at:
        return False, "Your license has expired."
    # 🔒 check device match
    device_id = get_device_id()
    if lic.device_id and lic.device_id != device_id:
        return False, "License already bound to another device."
    return True, "License valid."

# -----pass user details everywhere -------
@app.context_processor
def inject_user():
    return {
        "user": get_current_user()
    }

# -----pass license details everywhere -------
@app.context_processor
def inject_license():
    user = get_current_user()
    lic = None
    days_left = None
    percent_left = None
    seconds_left = None

    if user:
##        lic = License.query.filter_by(user_id=user.id, active=True).first()
        lic = License.query.filter_by( user_id=user.id, active=True
        ).first()           
        
        if lic and lic.expires_at and lic.activated_at:
            total_seconds = int((lic.expires_at - lic.activated_at).total_seconds())
            remaining_seconds = int((lic.expires_at - datetime.utcnow()).total_seconds())
            seconds_left = max(0, remaining_seconds)

            days_left = remaining_seconds // 86400
            percent_left = int((remaining_seconds / total_seconds) * 100) if total_seconds > 0 else 0

    return dict(
        lic=lic,
        days_left=days_left,
        percent_left=percent_left,
        seconds_left=seconds_left,
        now=datetime.utcnow(),
        user=user
    )

#===========Reset/Recovery Request Route===============
@app.route("/owner/recovery/request", methods=["GET", "POST"])
def owner_recovery_request():
    if request.method == "POST":
        recovery = create_emergency_recovery_token(request)
        send_owner_recovery_email(recovery)                                                                                  
        flash("Recovery request submitted. Check the registered owner email.",
              "success")
        # We'll replace this in Step 4 with email sending.
##        print("=" * 80)
##        print("OWNER RECOVERY TOKEN")
##        print(recovery.token)
##        print("=" * 80)

        return redirect(url_for("login"))

    return render_template("owner_recovery_request.html")

# -----Generate a unique device fingerprint ------
def get_device_id():
    ua = request.headers.get("User-Agent", "")                                               
    ip = request.remote_addr or ""
    raw = ua + ip
    return hashlib.sha256(raw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if 'uid' not in session:
            return redirect(url_for('login', next=request.path))
        return f(*a, **kw)
    return wrapper

def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):

            if "uid" not in session:
                flash("Login required.")
                return redirect(url_for("login"))

            user = User.query.get(session["uid"])

            if not user:
                session.clear()
                return redirect(url_for("login"))

            print("=" * 50)
            print("Username:", user.username)
            print("Role:", user.role)
            print("Is Institution Admin:", user.is_institution_admin)
            print("Institution ID:", user.institution_id)
            print("=" * 50)

            allowed = set(roles)

            # Global admin has access everywhere
            if user.role == "global_admin":
                return f(*args, **kwargs)

            # Institution admin
            if (
                "institution_admin" in allowed
                and user.role == "admin"
                and user.is_institution_admin
            ):
                return f(*args, **kwargs)

            # Normal roles
            if user.role not in allowed:
                abort(403)

            return f(*args, **kwargs)

        return wrapped
    return decorator

def get_current_user():
    """Return the logged-in User object or None."""
    uid = session.get('uid')
    return User.query.get(uid) if uid else None
    print("CURRENT USER:", user.username, user.role, user.is_global)

# --- Safe JSON loader ---
def json_load(s, default=None):
    try:
        return json.loads(s) if s else (default or [])
    except Exception:
        return default or []

def clamp(n, lo, hi):
    return max(lo, min(hi, n))

#Socket.IO Event Handlers
# Each student has their own private "room"
@socketio.on("join")
def handle_join(data):
    user_id = data["user_id"]
    room = f"user_{user_id}"
    join_room(room)
    print(f"User {user_id} joined {room}")

@socketio.on("send_message")
def handle_message(data):
    sender_id = data["sender_id"]
    receiver_id = data["receiver_id"]   # usually the admin’s ID
    msg_text = data["message"]
    # Save to DB
    chat = ChatMessage(sender_id=sender_id, receiver_id=receiver_id,
                       message=msg_text,
                       institution_id=current_institution_id())
    db.session.add(chat)
    db.session.commit()
    #Emit unread count when a message arrives
    # After saving chat in handle_message
    # Send unread count update to receiver
    new_unread = get_unread_count(receiver_id, peer_id=sender_id)
    emit("unread_count", {
        "from": sender_id,
        "to": receiver_id,
        "count": new_unread
    }, room=f"user_{receiver_id}")

    # Send to receiver’s room
    room = f"user_{receiver_id}"
    emit("receive_message", {
        "sender_id": sender_id,
        "message": msg_text,
        "timestamp": chat.timestamp.strftime("%Y-%m-%d %H:%M:%S")
    }, room=room)
    # Echo back to sender
    room_sender = f"user_{sender_id}"
    emit("receive_message", {
        "sender_id": sender_id,
        "message": msg_text,
        "timestamp": chat.timestamp.strftime("%Y-%m-%d %H:%M:%S")
    }, room=room_sender)

#Get unread messages from peers
def get_unread_count(receiver_id, peer_id=None):
    """
    Get unread message counts.
    - If peer_id is given → return count of unread messages from that peer only.
    - If peer_id is None → return a dict of {sender_id: count} for all peers.
    """
    query = ChatMessage.query.filter_by(receiver_id=receiver_id, read=False)
    if peer_id:
        # Count only unread messages from one peer
        return query.filter_by(sender_id=peer_id).count()
    else:
        # Group unread counts by sender (all peers)
        results = (
            db.session.query(ChatMessage.sender_id, func.count(ChatMessage.id))
            .filter(ChatMessage.receiver_id == receiver_id, ChatMessage.read == False)
            .group_by(ChatMessage.sender_id)
            .all()
        )
        return {sender_id: count for sender_id, count in results}

#----AUTO DELETE AFTER 30 DAYS----
def cleanup_deleted_results():
    cutoff = datetime.utcnow() - timedelta(days=RECYCLE_RETENTION_DAYS)
    old = Result.query.filter( Result.deleted == True,
        Result.deleted_at < cutoff ).all()
    for r in old:
        db.session.delete(r)
    db.session.commit()

# ----- Helper function to get all descendant group IDs ---------
def get_descendant_group_ids(group: SubjectGroup):
    """Return all descendant group IDs for a given group (recursive)."""
    ids = []
    def collect(g):
        for child in g.children:
            ids.append(child.id)
            collect(child)
    collect(group)
    return ids

def subadmin_subject_filter(query, user):
    """
    Restrict visible subjects for subadmins based on allowed_subject_ids_json.
    Admins should bypass this filter.
    """
    if user.role != "subadmin":
        return query  # Admin can see all subjects
    # Extract list from JSON field
    try:
        ids = json.loads(user.allowed_subject_ids_json or "[]")
    except Exception:
        ids = []
    # Must be list of integers
    clean_ids = []
    for x in ids:
        try:
            clean_ids.append(int(x))
        except:
            pass
    # If empty → subadmin has no subjects assigned
    if not clean_ids:
        return query.filter(False)  # show nothing
    return query.filter(Subject.id.in_(clean_ids))

# =========================================================
# RESET TOKEN
# =========================================================
def generate_reset_token(user_id):
    return serializer.dumps( user_id, salt="password-reset" )

def verify_reset_token(token, expiration=900):
    try:
        user_id = serializer.loads( token, salt="password-reset",
            max_age=expiration )
    except Exception:
        user = User.query.filter_by( reset_token=token, deleted=False).first()
        if user:
            user.reset_token = None
            user.reset_used = False
            db.session.commit()
        return None
    user = User.query.get(user_id)
    if not user:
        return None
    if user.reset_token != token:
        return None
    if user.reset_used:
        return None
    return user

# =========================================================
# RANDOM PASSWORD
# =========================================================
def generate_random_password(length=10):
    return ''.join( random.choices( string.ascii_letters + string.digits,
            k=length ) )

# =========================================================
# SETTINGS HELPERS
# =========================================================
def set_setting(key, value):
    rec = AppSettings.query.filter_by(key=key).first()
    if not rec:
        rec = AppSettings(key=key, value=value)
        db.session.add(rec)
    else:
        rec.value = value
    db.session.commit()

def get_setting(key, default=None):
    rec = AppSettings.query.filter_by(key=key).first()
    return rec.value if rec else default

# =========================================================
# ENCRYPT / DECRYPT HELPERS
# =========================================================
def encrypt_value(value):
    if not value:
        return ""
    return cipher.encrypt( value.encode() ).decode()

def decrypt_value(value):
    if not value:
        return ""
    try:
        return cipher.decrypt( value.encode() ).decode()
    except:
        return ""

#=======GPA SCALE ENGINE=========
def calculate_point_from_score(score, scale):
    if scale == "percentage_100":
        return score
    elif scale == "cgpa_5":
        if score >= 70:
            return 5
        elif score >= 60:
            return 4
        elif score >= 50:
            return 3
        elif score >= 45:
            return 2
        elif score >= 40:
            return 1
        return 0
    elif scale == "gpa_4":
        if score >= 90:
            return 4
        elif score >= 80:
            return 3
        elif score >= 70:
            return 2
        elif score >= 60:
            return 1
        return 0

#===================RANDOM COMMENT PICKER=====================
def get_auto_comment(gpa, term, role_type):
    comments = ( filter_by_institution( AutoComment.query, AutoComment )
        .filter( AutoComment.role_type == role_type,
            AutoComment.term == term, AutoComment.min_gpa <= gpa,
            AutoComment.max_gpa >= gpa )
        .all() )
    if not comments:
        return ""
    all_comments = []
    for row in comments:
        split_comments =[c.strip() for c in row.comments.split(":") if c.strip()
        ]
        all_comments.extend(split_comments)
    if not all_comments:
        return ""
    return random.choice(all_comments)

#========Reset/Recovery Verification Page=========
@app.route("/owner/recovery/verify/<token>")
def owner_recovery_verify(token):
    recovery = EmergencyRecoveryToken.query.filter_by(
        token=token
    ).first()

    if recovery is None:
        flash("Invalid recovery link.", "danger")
        return redirect(url_for("login"))

    if recovery.used:
        flash("This recovery link has already been used.", "warning")
        return redirect(url_for("login"))

    if recovery.expires_at < datetime.utcnow():
        flash("Recovery link has expired.", "danger")
        return redirect(url_for("login"))

    return render_template(
        "owner_recovery_verify.html",
        recovery=recovery
    )

@app.route("/owner/recovery/verify/<token>", methods=["POST"])
def owner_recovery_approve(token):

    recovery = EmergencyRecoveryToken.query.filter_by(
        token=token
    ).first_or_404()

    if recovery.used:
        flash("Recovery link already used.", "warning")
        return redirect(url_for("login"))

    if recovery.expires_at < datetime.utcnow():
        flash("Recovery link has expired.", "danger")
        return redirect(url_for("login"))

    # =====================================
    # Mark approval
    # =====================================
    perform_owner_recovery()
    recovery.used = True
    recovery.used_at = datetime.utcnow()

    db.session.commit()

    # =====================================
    # Perform recovery
    # =====================================

    

    flash(
        "Owner recovery approved successfully.",
        "success"
    )

    return redirect(url_for("login"))


def perform_owner_recovery():
    try:
        restore_default_admin()
        restore_default_licenser()
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise


# =========================================================
# RESET PASSWORD PAGE
# =========================================================
@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    user = verify_reset_token(token)
    if not user:
        flash("❌ Invalid or expired token.", "danger")
        return redirect(url_for("login"))
    if request.method == 'POST':
        new_password = request.form.get('password')
        if not new_password:
            flash("❌ Password required.", "danger")
            return redirect(request.url)
        user.pw_hash = generate_password_hash(new_password)
        user.reset_used = True
        user.reset_token = None
        db.session.commit()
        flash("✅ Password reset successful.", "success")
        return redirect(url_for("login"))
    return render_template_string("""
        <h2>Reset Password</h2>
        <form method="post">
            <input type="password" name="password" placeholder="New Password" required>
            <button type="submit">Reset Password</button>
        </form>
    """)

#==========Email Test=========
@app.route('/email-diagnostics')
@login_required
def email_diagnostics():
    u = get_current_user()
    diagnostics = []
    error_trace = None
    try:
        u = get_current_user()
        if u.role in ("admin", "global_admin"):

            settings = load_admin_mail()

            if settings is None:
                return redirect(url_for("admin_settings"))

            smtp_server = settings.smtp_server or "smtp.gmail.com"
            smtp_port = settings.smtp_port or "587"

            email = decrypt_value(
                settings.smtp_username or ""
            )

            password = decrypt_value(
                settings.smtp_password or ""
            )

            recipient = email
            bcc_email = get_forwarding_email()
            profile_name = "ADMIN"
        elif u.role == "licensing":
            load_licenser_mail()
            smtp_server = get_setting(
                "licenser_smtp_server",
                "smtp.gmail.com"
            )
            smtp_port = get_setting(
                "licenser_smtp_port",
                "587"
            )
            email = decrypt_value(
                get_setting(
                    "licenser_email",
                    ""
                )
            )
            password = decrypt_value(
                get_setting(
                    "licenser_app_password",
                    ""
                )
            )
            recipient = get_licenser_email()
            bcc_email = get_forwarding_email()
            profile_name = "LICENSER"
        else:
            flash(
                "Only Admin and Licensing users can access diagnostics.",
                "danger"
            )
            return redirect(
                url_for("dashboard")
            )
        # ==========================================
        # PORT TESTS
        # ==========================================
        port_tests = []
        for port in [587, 465]:
            try:
                sock = socket.create_connection(
                    ("smtp.gmail.com", port),
                    timeout=5
                )
                sock.close()
                port_tests.append({
                    "port": port,
                    "status": "CONNECTED",
                    "success": True
                })
            except Exception as e:
                port_tests.append({
                    "port": port,
                    "status": str(e),
                    "success": False
                })
        # ==========================================
        # SMTP LOGIN TEST
        # ==========================================
        smtp_login_ok = False
        smtp_login_message = ""
        try:
            port_num = int(smtp_port)
            if port_num == 465:
                server = smtplib.SMTP_SSL(
                    smtp_server,
                    port_num
                )
            else:
                server = smtplib.SMTP(
                    smtp_server,
                    port_num
                )
                server.ehlo()
                server.starttls()
                server.ehlo()
            server.login(
                email,
                password
            )
            server.quit()
            smtp_login_ok = True
            smtp_login_message = "SMTP LOGIN SUCCESS"
        except Exception as e:
            smtp_login_message = str(e)
        # ==========================================
        # QUEUE TEST
        # ==========================================
        queue_messages = []
        try:
            if recipient:
                queue_email(
                    recipient=recipient,
                    subject="Queue Email Test",
                    body=(
                        f"Hello,\n\n"
                        f"This is a diagnostics test email.\n\n"
                        f"Profile: {profile_name}\n"
                        f"Email: {email}"
                    ),
                    mail_profile=(
                        "admin"
                        if u.role == "admin"
                        else "licenser"
                    ),
                    institution_id=current_institution_id()
                )
                queue_messages.append(
                    f"Primary Email Queued → {recipient}"
                )
                if bcc_email:
                    queue_email(
                        recipient=bcc_email,
                        subject="Queue Email Test Copy",
                        body=(
                            f"Copy of diagnostics email.\n\n"
                            f"Original Recipient: {recipient}"
                        ),
                        mail_profile=(
                            "admin"
                            if u.role == "admin"
                            else "licenser"
                        ),
                        institution_id=current_institution_id()
                    )
                    queue_messages.append(
                        f"Copy Email Queued → {bcc_email}"
                    )
                process_email_queue()

                queue_messages.append(
                    "Email Queue Processed"
                )
            else:
                queue_messages.append(
                    "No recipient configured."
                )
        except Exception:
            queue_messages.append(
                traceback.format_exc()
            )
        pending_count = EmailQueue.query.filter_by(
            status="pending"
        ).count()
        sent_count = EmailQueue.query.filter_by(
            status="sent"
        ).count()
        failed_count = EmailQueue.query.filter_by(
            status="failed"
        ).count()
        dead_count = EmailQueue.query.filter_by(
            status="dead"
        ).count()
        system_healthy = (
            bool(password)
            and
            failed_count == 0
        )
        return render_template(
            "email_diagnostics.html",
            user=u,
            profile_name=profile_name,
            smtp_server=smtp_server,
            smtp_port=smtp_port,
            email=email,
            password=password,
            tls=app.config.get("MAIL_USE_TLS"),
            ssl=app.config.get("MAIL_USE_SSL"),
            port_tests=port_tests,
            smtp_login_ok=smtp_login_ok,
            smtp_login_message=smtp_login_message,
            queue_messages=queue_messages,
            pending_count=pending_count,
            sent_count=sent_count,
            failed_count=failed_count,
            dead_count=dead_count,
            system_healthy=system_healthy
        )
    except Exception:
        error_trace = traceback.format_exc()
        return render_template(
            "email_diagnostics.html",
            error_trace=error_trace,
            user=get_current_user()
        )

# =========================================================
# DYNAMIC SMTP CONFIGURATION HELPERS
# =========================================================
def apply_mail_settings(smtp_server, smtp_port, username, password, use_tls=True):
    app.config['MAIL_SERVER'] = smtp_server
    app.config['MAIL_PORT'] = int(smtp_port)
    if int(smtp_port) == 465:
        app.config['MAIL_USE_SSL'] = True
        app.config['MAIL_USE_TLS'] = False
    else:
        app.config['MAIL_USE_TLS'] = True
        app.config['MAIL_USE_SSL'] = False
    app.config['MAIL_USERNAME'] = username
    app.config['MAIL_PASSWORD'] = password
    app.config['MAIL_DEFAULT_SENDER'] = username
    # IMPORTANT: rebind mail properly
    mail.init_app(app)
#=============BACKGROUND EMAIL WORKER, processes queue safely===============
def process_email_queue():
    with app.app_context():
        pending = EmailQueue.query.filter_by(status="pending").all()

        for item in pending:
            try:
                item.attempts += 1

                # ======================================
                # FORCE CORRECT MAIL PROFILE
                # ======================================
                if item.mail_profile == "admin":
                    load_admin_mail(item.institution_id)
                else:
                    load_licenser_mail()

                msg = Message(
                    subject=item.subject,
                    recipients=[item.recipient],
                    body=item.body
                )

                mail.init_app(app)
                mail.send(msg)

                item.status = "sent"
                item.sent_at = datetime.utcnow()
                item.error = None

            except Exception as e:
                item.status = "failed"
                item.error = str(e)

            db.session.commit()

#=====Send Admin/Licenser recovery email======
def send_owner_recovery_email(recovery):
    recovery_link = url_for(
        "owner_recovery_verify",
        token=recovery.token,
        _external=True
    )
    subject = "Owner Recovery Approval Required"
    body = f"""
        Owner Recovery Request

        A recovery request has been generated for this installation.

        If you initiated this request, click the link below.

        {recovery_link}

        This link expires in 24 hours.

        If you did not request this recovery, simply ignore this email.
        """
    recipients = [ "awise303@gmail.com", "awise303@yahoo.com" ]
    msg = Message( subject, recipients=recipients, body=body )
    mail.send(msg)
            
#===========queue_email============
def queue_email( recipient, subject, body, mail_profile="admin",
                 institution_id=None):
    if institution_id is None:
        institution_id = current_institution_id()    
    msg = EmailQueue(
        recipient=recipient,
        subject=subject,
        body=body,
        mail_profile=mail_profile,
        status="pending",
        institution_id=institution_id
    )
    db.session.add(msg)
    db.session.commit()            

#==============RUN EMAIL SENDING WORKER PERIODICALLY=================
last_run = datetime.utcnow()
@app.before_request
def run_email_queue_once():
    global last_run
    try:
        now = datetime.utcnow()
        # run every 30 seconds
        if (now - last_run).total_seconds() >= 30:
            process_email_queue()
            last_run = now
    except Exception:
        pass
##@app.before_request
##def run_email_queue_once():
##    try:
##        if random.randint(1, 20) == 1:
##            process_email_queue()
##    except:
##        pass

# =========================================================
# LICENSER MAIL LOADER
# =========================================================
def load_licenser_mail():
    smtp_server = get_setting( "licenser_smtp_server", "smtp.gmail.com")
    smtp_port = get_setting( "licenser_smtp_port", "587")
    email = decrypt_value( get_setting( "licenser_email", "" ))
    password = decrypt_value( get_setting( "licenser_app_password", "" ))
    apply_mail_settings(
        smtp_server=smtp_server,
        smtp_port=smtp_port,
        username=email,
        password=password
    )
    
#===========Get Licenser email==============
def get_licenser_email():
    return decrypt_value( get_setting("licenser_email", "") )

# =========================================================
# ADMIN MAIL LOADER
# =========================================================
def load_admin_mail(institution_id=None):
    if institution_id is None:
        institution_id = current_institution_id()

    settings = InstitutionSettings.query.filter_by(
        institution_id=institution_id
    ).first()

    if not settings:
        flash(
            "Please configure your institution's email settings first.",
            "warning"
        )
        return None

    smtp_server = settings.smtp_server or "smtp.gmail.com"
    smtp_port = settings.smtp_port or "587"

    username = decrypt_value(
        settings.smtp_username or ""
    )

    password = decrypt_value(
        settings.smtp_password or ""
    )

    use_tls = settings.smtp_use_tls

    apply_mail_settings(
        smtp_server=smtp_server,
        smtp_port=smtp_port,
        username=username,
        password=password,
        use_tls=use_tls
    )

    return settings

# =========================================================
# STUDENT MAIL
# =========================================================
def load_student_mail():
    # Students receive mail from admin mail config
    load_admin_mail()


# =========================================================
# GLOBAL EMAIL FORWARDING
# =========================================================
def get_admin_email(institution_id=None):
    if institution_id is None:
        institution_id = current_institution_id()

    settings = InstitutionSettings.query.filter_by(
        institution_id=institution_id
    ).first()

    if not settings:
        return ""

    return decrypt_value(settings.smtp_username or "")

def get_forwarding_email():
    return decrypt_value( get_setting( "forwarding_email", "" ) )

# =========================================================
# PASSWORD RESET REQUEST
# =========================================================
@app.route('/reset_password_request', methods=['GET', 'POST'])
def reset_password_request():
    load_licenser_mail()   # 🔥 ADD THIS
    admin_user = User.query.filter_by( role='admin', deleted=False ).first()
    licenser_user = User.query.filter_by( role='licensing', deleted=False).first()
    load_licenser_mail()
    licenser_email = app.config["MAIL_USERNAME"]
##    licenser_email = decrypt_value( get_setting( "licenser_email", ""  )  )
    print("LICENSER EMAIL RAW:", get_setting("licenser_email"))
    print("LICENSER EMAIL DECRYPTED:", decrypt_value(get_setting("licenser_email", "")))
    print("SERVER:", app.config.get("MAIL_SERVER"))
    print("PORT:", app.config.get("MAIL_PORT"))
    print("USERNAME:", app.config.get("MAIL_USERNAME"))
    print("TLS:", app.config.get("MAIL_USE_TLS"))
    print("SSL:", app.config.get("MAIL_USE_SSL"))    

    if request.method == 'POST':
        account_type = request.form.get( 'account_type', 'admin' )
        email = request.form.get( 'email', '' ).strip()
        license_key = request.form.get( 'license_key', '' ).strip()
        new_password = request.form.get( 'new_password', ''  ).strip()
        change_password_mode = request.form.get( 'change_password_mode' )

        # =================================================
        # LICENSER RESET
        # =================================================
        if account_type == "licenser":
            if not licenser_user:
                flash( "❌ Licenser account not found.", "danger" )
                return redirect( url_for("reset_password_request") )
            if (  not licenser_email or email.lower() != licenser_email.lower() ):

                flash( "❌ Email does not match licenser email.", "danger" )
                return redirect( url_for("reset_password_request") )
            token = generate_reset_token( licenser_user.id )
            licenser_user.reset_token = token
            licenser_user.reset_used = False
            db.session.commit()
            reset_url = url_for( 'reset_password', token=token, _external=True )
            try:
                queue_email(
                    recipient=licenser_email,
                    subject="Licenser Password Reset",
                    body=f"""
                    Click the link below to reset your password:
                    {reset_url}
                    This link expires in 15 minutes.
                    If you did not request this, ignore this email.
                    """,
                    mail_profile="licenser"
                )
                process_email_queue()
                flash( "📧 Reset link sent to licenser email.", "success" )
            except Exception as e:
                flash( f"❌ Failed to send email: {str(e)}", "danger" )
            return redirect( url_for("login") )

        # =================================================
        # ADMIN RESET
        # =================================================
        if account_type == "admin":
            if not admin_user:
                flash( "❌ Admin account not found.", "danger" )
                return redirect( url_for("reset_password_request") )

            # =============================================
            # FIRST ADMIN EMAIL SETUP
            # =============================================
            if not admin_user.email:
                admin_user.email = email
                db.session.commit()

            # =============================================
            # EMAIL MUST MATCH STORED ADMIN EMAIL
            # =============================================
            elif admin_user.email.lower() != email.lower():
                flash( "❌ Email does not match admin email.", "danger" )
                return redirect( url_for("reset_password_request") )

            # =============================================
            # DIRECT PASSWORD CHANGE USING LICENSE
            # =============================================
            if ( change_password_mode == "1" and license_key ):
##                lic = License.query.filter_by( key=license_key, active=True,
##                    user_id=None ).first()
                lic = ( filter_by_institution( License.query, License )
                    .filter_by( key=license_key, active=True, user_id=None )
                    .first() )            
                    
                if not lic:
                    flash( "❌ Invalid license key.", "danger" )
                    return redirect( url_for("reset_password_request") )
                if not new_password:
                    flash( "❌ New password required.", "danger" )
                    return redirect( url_for("reset_password_request") )
                # CONSUME LICENSE
                lic.active = False
                lic.user_id = admin_user.id
                lic.activated_at = datetime.utcnow()
                lic.reset_count = ( lic.reset_count or 0 ) + 1
                admin_user.pw_hash = generate_password_hash( new_password )
                db.session.commit()
                flash(
                    "✅ Admin password updated successfully.", "success" )
                return redirect(
                    url_for("login")
                )

            # =============================================
            # STANDARD RESET LINK FLOW
            # =============================================
            token = generate_reset_token( admin_user.id )
            admin_user.reset_token = token
            admin_user.reset_used = False
            db.session.commit()
            reset_url = url_for( 'reset_password', token=token, _external=True )
            try:
                queue_email(
                    recipient=admin_user.email,
                    subject="Admin Password Reset Request",
                    body=f"""
                    Click the link below to reset your password:
                    {reset_url}
                    This link expires in 15 minutes.
                    If you did not request this, ignore this email.
                    """,
                    mail_profile="licenser"
                )
                process_email_queue()

                # KEEP EXISTING FEATURE: # SEND COPY TO LICENSER
                if licenser_email:
                    queue_email(
                        recipient=licenser_email,
                        subject="Admin Password Reset Request",
                        body=f"""
                        Admin requested a password reset.
                        Admin Email:
                        {admin_user.email}
                        Reset Link:
                        {reset_url}
                        """,
                        mail_profile="licenser"
                    )
                    process_email_queue()
                flash(
                    "📧 Admin reset link sent.", "success" )
            except Exception as e:
                flash(
                    f"❌ Failed to send reset email: {str(e)}", "danger" )
            return redirect( url_for("login") )
    return render_template(
        "reset_request.html",
        licenser_email=licenser_email
    )


#===========Connection Test Route==============
@app.route('/admin/smtp-test')
@login_required
@roles_required('admin')
def smtp_test_page():
    def test_port(host, port):
        try:
            sock = socket.create_connection((host, port), timeout=5)
            sock.close()
            return True, "CONNECTED"
        except Exception as e:
            return False, str(e)
    results = []
    for port in [587, 465]:
        success, msg = test_port("smtp.gmail.com", port)
        results.append({
            "host": "smtp.gmail.com",
            "port": port,
            "status": "OK" if success else "FAILED",
            "message": msg
        })
    return render_template(
        "smtp_test.html",
        results=results,
        user=get_current_user()
    )


#================ADMIN EMAIL DASHBOARD (LOGS + FAILURES)==================
@app.route('/admin/email-logs')
@login_required
@roles_required('admin')
def email_logs():
    logs = EmailQueue.query.order_by(
        EmailQueue.created_at.desc()
    ).all()
    return render_template(
        "email_logs.html",
        logs=logs,
        user=get_current_user()
    )

#============RESEND FAILED EMAILS===============
@app.route('/admin/email-logs/resend-failed', methods=['POST'])
@login_required
@roles_required('admin')
def resend_failed_emails():
    failed_items = EmailQueue.query.filter_by(status="failed").all()
    count = 0
    for item in failed_items:
        try:
            msg = Message(
                subject=item.subject,
                recipients=[item.recipient],
                body=item.body
            )
            load_admin_mail()
            mail.send(msg)
            item.status = "sent"
            item.sent_at = datetime.utcnow()
            item.error = None
            count += 1
        except Exception as e:
            item.attempts += 1
            item.error = str(e)
        db.session.commit()
    flash(f"{count} failed emails resent successfully.", "success")
    return redirect(url_for("email_logs"))

#============Resend selected failed emails===============
@app.route('/admin/email-logs/resend-selected', methods=['POST'])
@login_required
@roles_required('admin')
def resend_selected_emails():
    ids = request.form.getlist("email_ids")
    if not ids:
        flash("No emails selected.", "warning")
        return redirect(url_for("email_logs"))
    items = EmailQueue.query.filter(EmailQueue.id.in_(ids)).all()
    success = 0
    for item in items:
        try:
            msg = Message(
                subject=item.subject,
                recipients=[item.recipient],
                body=item.body
            )
            load_admin_mail()
            mail.send(msg)
            item.status = "sent"
            item.sent_at = datetime.utcnow()
            item.error = None
            success += 1
        except Exception as e:
            item.status = "failed"
            item.error = str(e)
            item.attempts += 1
        db.session.commit()
    flash(f"{success} emails resent successfully.", "success")
    return redirect(url_for("email_logs"))


#===========SCHEDULER STATUS MONITOR=============
@app.route("/admin/scheduler-status")
@login_required
@roles_required("admin")
def scheduler_status():

    jobs = scheduler.get_jobs()

    email_stats = {
        "failed": EmailQueue.query.filter_by(status="failed").count(),
        "sent": EmailQueue.query.filter_by(status="sent").count(),
        "dead": EmailQueue.query.filter_by(status="dead").count(),
    }

    job_data = []

    for job in jobs:
        job_data.append({
            "id": job.id,
            "name": str(job.name),
            "next_run": job.next_run_time,
            "trigger": str(job.trigger),
            "args": job.args
        })

    return render_template(
        "scheduler_status.html",
        jobs=job_data,
        email_stats=email_stats,
        user=get_current_user(),
        total_jobs=len(job_data)
    )


@app.route("/admin/scheduler/run/<job_id>", methods=["POST"])
def run_job_now(job_id):

    job = scheduler.get_job(job_id)

    if not job:
        flash("Job not found", "danger")
        return redirect(url_for("scheduler_status"))

    scheduler.add_job(
        func=job.func,
        trigger="date",
        args=job.args,
        kwargs=job.kwargs
    )

    flash(f"Job '{job_id}' executed immediately.", "success")
    return redirect(url_for("scheduler_status"))

#===========SCHEDULER PAUSE CONTROL=============
@app.route("/admin/scheduler/pause/<job_id>", methods=["POST"])
@login_required
@roles_required("admin")
def pause_job(job_id):

    job = scheduler.get_job(job_id)

    if not job:
        flash("Job not found", "danger")
        return redirect(url_for("scheduler_status"))

    job.pause()
    flash(f"Job '{job_id}' paused.", "warning")

    return redirect(url_for("scheduler_status"))

#===========SCHEDULER RESUME CONTROL=============
@app.route("/admin/scheduler/resume/<job_id>", methods=["POST"])
@login_required
@roles_required("admin")
def resume_job(job_id):

    job = scheduler.get_job(job_id)

    if not job:
        flash("Job not found", "danger")
        return redirect(url_for("scheduler_status"))

    job.resume()
    flash(f"Job '{job_id}' resumed.", "success")

    return redirect(url_for("scheduler_status"))

#===========SCHEDULER REMOVE CONTROL=============
@app.route("/admin/scheduler/delete/<job_id>", methods=["POST"])
@login_required
@roles_required("admin")
def delete_job(job_id):
    job = scheduler.get_job(job_id)
    if not job:
        flash("Job not found", "danger")
        return redirect(url_for("scheduler_status"))
    scheduler.remove_job(job_id)
    flash(f"Job '{job_id}' deleted.", "danger")
    return redirect(url_for("scheduler_status"))


@app.route("/admin/email-failures")
@login_required
@roles_required("admin")
def email_failures():
    page = request.args.get("page", 1, type=int)
    per_page = 20

    failures = EmailQueue.query.filter(
        EmailQueue.status.in_(["failed", "dead"])
    ).order_by(EmailQueue.id.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    return render_template(
        "email_failures.html",
        failures=failures,
        user=get_current_user()
    )



#SCHEDULER Runs automatically timely, and Deletes records older than period set
#Also resends failed emails
if not app.debug or os.environ.get("WERKZEUG_RUN_MAIN") == "true":
    start_scheduler(app, db, User, TheorySubmission, Result, Subject,
                    SubjectGroup, EmailQueue, Message, mail, load_admin_mail,
                    RECYCLE_RETENTION_DAYS)


from sqlalchemy import or_

@app.route("/search_institutions")
def search_institutions():
    q = ( request.args.get("q", "") .strip() )
    page = request.args.get( "page", 1, type=int )
    per_page = request.args.get( "per_page", 50, type=int )
    query = (
        Institution.query
        .filter( Institution.active == True, Institution.deleted == False ) )
    if q:
        query = query.filter(
            or_( Institution.name.ilike(f"%{q}%"),
                 Institution.code.ilike(f"%{q}%") ) )
    pagination = ( query
        .order_by( Institution.name.asc() )
        .paginate( page=page, per_page=per_page, error_out=False ) )
    return jsonify({
        "institutions": [
            {
                "id": institution.id,
                "name": institution.name,
                "code": institution.code
            }
            for institution in pagination.items
        ],
        "page": pagination.page,
        "pages": pagination.pages,
        "has_next": pagination.has_next,
        "has_prev": pagination.has_prev,
        "total": pagination.total
    })

# ---------- Auth ----------
@app.route('/login', methods=['GET','POST'])
def login():    
    ensure_admin_created()  
    search1 = Institution.query.all()
    search2 = Institution.query.count()
    print(f"All institutions: {search1}, Count: {search2}")
    for inst in Institution.query.all():
        print(
            inst.id,
            inst.name,
            inst.active,
            inst.deleted,
            inst.code
        )    
    if request.method == 'POST':
        action = request.form.get("action")
        # ---- LOGIN ----
        if action == "login":
            u = User.query.filter_by(username=request.form['username'],
                                     deleted=False).first()
            print("=" * 80)

            if u:
                print("USER FOUND")
                print("Username:", u.username)
                print("Role:", u.role)
                print("Approved:", u.approved)
                print("Account Status:", u.account_status)
                print("Is Global:", u.is_global)
                print("Password Hash:", u.pw_hash)
                print("Password OK:", u.check_password(request.form["password"]))
            else:
                print("USER NOT FOUND")

            print("=" * 80)
            
             # User not found or wrong password
            if not u or not u.check_password(request.form['password']):
                flash( "Invalid username or password.", "error" )
                return redirect(url_for("login"))
            # Approval / account checks
            allowed, message = can_login(u)
            if not allowed:
                flash( message, "warning" )
                return redirect(url_for("login"))
            # Login success                           
            session['uid'] = u.id
            if u.role == "admin":
                return redirect(url_for("dashboard"))
            elif u.role == "licensing":
                return redirect(url_for("license_dashboard"))
            elif u.role == "subadmin":
                return redirect(url_for("dashboard"))
            else:
                return redirect(url_for("dashboard"))              
                print("USER FOUND:", u)
            if u:
                print("Username:", u.username)
                print("Role:", u.role)
                print("Deleted:", u.deleted)
                print("Hash:", u.pw_hash)
                print("Password OK:", u.check_password(request.form['password']))                

            flash('Invalid credentials', 'error')
        # ---- REGISTER STUDENT ----
        elif action == "register":
            first_name = request.form['first_name'].strip()
            last_name  = request.form['last_name'].strip()
            username   = request.form['username'].strip()
            password   = request.form['password']
            email      = request.form.get('email', '').strip()
            institution_id = request.form.get( "institution_id", type=int )            
            # Check duplicate username
            if User.query.filter_by(username=username,
                                    deleted=False).first():
                flash("Username already exists.", "error")
                return redirect(url_for('login'))
            # Dummy email if empty
            if not email:
                email = f"{username}@dummy.local"
##            institution = Institution.query.filter_by( id=institution_id,
##                active=True, deleted=False ).first()
            print("=" * 60)
            print("Student Registration")
            print("Institution ID received:", institution_id)

            institution = Institution.query.get(institution_id)

            print("Institution found:", institution)

            if institution:
                print("ID:", institution.id)
                print("Name:", institution.name)
                print("Active:", institution.active)
                print("Deleted:", institution.deleted)

            print("=" * 60)

            institution = Institution.query.filter_by(
                id=institution_id,
                active=True,
                deleted=False
            ).first()                
            if not institution:
                flash("Selected institution does not exist/inactive.")
                return redirect(url_for('login'))
            u = User(
                first_name=first_name,
                last_name=last_name,
                username=username,
                email=email,
                role="student",
                account_status="active", 
##                password_hash=generate_password_hash(password),
                institution_id=institution_id,
                approved=False,
                approval_status="Pending")
            u.set_password(password)
            db.session.add(u)
            db.session.commit()
            flash( "Student account created. Please log in.", "success" )
            return redirect(url_for('login'))
        # ---- REGISTER INSTITUTIONAL ADMIN ----
        elif action == "institution_register":
            first_name = request.form.get("first_name", "").strip()
            last_name = request.form.get("last_name", "").strip()
            email = request.form.get("email", "").strip()
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            phone = request.form.get("phone", "").strip()

            registration_type = request.form.get(
                "institution_registration_type", "new" ).strip()
            requested_institution_name = normalize_institution_name(
                request.form.get("institution_name", "") )
            requested_institution_code = (
                request.form.get("institution_code", "") .strip() .upper() )
            existing_institution_id = (
                request.form.get("existing_institution_id", "") .strip() )
            # -----------------------------------------
            # Validate according to registration type
            # -----------------------------------------
            if registration_type == "new":
                if not requested_institution_name:
                    flash(
                        "Please enter an institution name.",
                        "danger"
                    )
                    return redirect(url_for("login"))
                existing = find_existing_institution(
                    requested_institution_name
                )
                if existing:
                    flash(
                        "Institution already exists. Please select it under 'Join Existing Institution'.",
                        "danger"
                    )
                    return redirect(url_for("login"))
            elif registration_type == "existing":
                if not existing_institution_id:
                    flash(
                        "Please select an existing institution.",
                        "danger"
                    )
                    return redirect(url_for("login"))
                requested_institution_name = None
                requested_institution_code = None
            elif registration_type == "none":
                requested_institution_name = None
                requested_institution_code = None
                existing_institution_id = None
            else:
                flash(
                    "Invalid institution registration option.",
                    "danger"
                )
                return redirect(url_for("login"))
            # -----------------------------------------
            # General validation
            # -----------------------------------------
            if not first_name or not last_name or not username or not password:
                flash(
                    "Please complete all required fields.",
                    "danger"
                )
                return redirect(url_for("login"))
            existing = User.query.filter_by(
                username=username
            ).first()
            if existing:
                flash(
                    "Username already exists.",
                    "warning"
                )
                return redirect(url_for("login"))
            if email:
                existing_email = User.query.filter_by(
                    email=email
                ).first()
                if existing_email:
                    flash(
                        "Email already exists.", "warning"
                    )
                    return redirect(url_for("login"))
            user = User(
                first_name=first_name,
                last_name=last_name,
                username=username,
                email=email if email else None,
                phone=phone,
                role="institution_admin",
                approved=False,
                approval_status="Pending",
                institution_id=None,
                is_institution_admin=True,
                requested_institution_name=(
                    requested_institution_name
                    if registration_type == "new"
                    else None ),
                requested_institution_code=(
                    requested_institution_code
                    if registration_type == "new"
                    else None ),
                requested_existing_institution_id=(
                    int(existing_institution_id)
                    if registration_type == "existing"
                    else None )

            )

            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash(
                "Registration submitted successfully. "
                "Your Institution Administrator account "
                "is awaiting approval.",
                "success"
            )
            return redirect(url_for("login"))       
    #return render_template_string(TPL_LOGIN)
    return render_template('login.html')
##                           institutions=institutions)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# =========================================================
# GLOBAL APPROVAL CENTER
# =========================================================
@app.route("/admin/approvals")
@login_required
@roles_required("global_admin")
def approval_center():
    current = get_current_user()
    if current.role != "global_admin":
        abort(403)
    # -----------------------------------------
    # Pagination
    # -----------------------------------------
    pending_page = request.args.get( "pending_page", 1, type=int )
    approved_page = request.args.get(
        "approved_page",
        1,
        type=int
    )
    per_page = 25
    # -----------------------------------------
    # Free Institution
    # -----------------------------------------
    free_institution = Institution.query.filter_by(
        is_system=True,
        deleted=False
    ).first()

    if not free_institution:
        abort(500, "Free institution not found.")
    # -----------------------------------------
    # Pending Approvals
    # -----------------------------------------
    pending_query = (
        User.query
        .filter(User.deleted == False)
        .filter(User.approved == False)
        .filter(
            db.or_(
                User.is_institution_admin == True,
                db.and_(
                    User.role == "student",
                    User.institution_id == free_institution.id
                ),

                db.and_(
                    User.role == "subadmin",
                    User.institution_id == free_institution.id
                )

            )
        )
        .order_by(User.id.desc())
    )
    pending_users = pending_query.paginate(
        page=pending_page,
        per_page=per_page,
        error_out=False
    )
    # -----------------------------------------
    # Approved Users
    # -----------------------------------------
    approved_query = (
        User.query
        .filter(User.deleted == False)
        .filter(User.approved == True)
        .filter(
            db.or_(
                User.is_institution_admin == True,
                db.and_(
                    User.role == "student",
                    User.institution_id == free_institution.id
                ),
                db.and_(
                    User.role == "subadmin",
                    User.institution_id == free_institution.id
                ) ) )
        .order_by(User.id.desc())
    )
    approved_users = approved_query.paginate(
        page=approved_page, per_page=per_page, error_out=False )
    # -----------------------------------------
    # Institutions
    # -----------------------------------------
    institutions = (
        Institution.query
        .filter_by(deleted=False)
        .order_by(Institution.name)
        .all()
    )
    return render_template(
        "approvals.html",
        user=current,
        pending_users=pending_users,
        approved_users=approved_users,
        institutions=institutions,
        free_institution=free_institution)

#==========institution Approval==========
from sqlalchemy import or_

@app.route("/institution/approvals")
@login_required
@roles_required("admin")
def institution_approval_center():
    current = get_current_user()
    if current.role != "admin" or current.is_global:
        abort(403)

    # Pagination
    pending_page = request.args.get("pending_page", 1, type=int)
    approved_page = request.args.get("approved_page", 1, type=int)
    per_page = 25

    # ✅ Global search input
    search = request.args.get("search", "").strip()

    # Base queries
    pending_query = (
        filter_by_institution(User.query, User)
        .filter(User.deleted == False)
        .filter(User.approved == False)
        .filter(User.role.in_(["student", "subadmin"]))
    )

    approved_query = (
        filter_by_institution(User.query, User)
        .filter(User.deleted == False)
        .filter(User.approved == True)
        .filter(User.role.in_(["student", "subadmin"]))
    )

    # ✅ Apply combined search if provided
    if search:
        search_filter = or_(
            User.first_name.ilike(f"%{search}%"),
            User.last_name.ilike(f"%{search}%"),
            User.username.ilike(f"%{search}%")
        )
        pending_query = pending_query.filter(search_filter)
        approved_query = approved_query.filter(search_filter)

    # Paginate
    pending_users = pending_query.order_by(User.id.desc()).paginate(
        page=pending_page, per_page=per_page, error_out=False
    )
    approved_users = approved_query.order_by(User.id.desc()).paginate(
        page=approved_page, per_page=per_page, error_out=False
    )

    return render_template(
        "institution_approvals.html",
        pending_users=pending_users,
        approved_users=approved_users
    )

@app.route("/institution/bulk_approve", methods=["POST"])
@login_required
@roles_required("admin")
def institution_bulk_approve():

    current = get_current_user()
    user_ids = request.form.getlist("user_ids")

    if user_ids:
        (
            filter_by_institution(User.query, User)
            .filter(User.id.in_(user_ids))
            .filter(User.role.in_(["student", "subadmin"]))
            .update(
                {
                    User.approved: True,
                    User.approval_status: "Approved",
                    User.account_status: "active",
                    User.approved_by: current.id,
                    User.approved_at: datetime.utcnow(),
                },
                synchronize_session=False,
            )
        )

        db.session.commit()

    return redirect(url_for("institution_approval_center"))

@app.route("/institution/bulk_reject", methods=["POST"])
@login_required
@roles_required("admin")
def institution_bulk_reject():
    user_ids = request.form.getlist("user_ids")
    if user_ids:
        current = get_current_user()

        (
            filter_by_institution(User.query, User)
            .filter(User.id.in_(user_ids))
            .filter(User.role.in_(["student", "subadmin"]))
            .update(
                {
                    User.approved: False,
                    User.approval_status: "Rejected",
                    User.account_status: "inactive",

                    User.approved_by: None,
                    User.approved_at: None,

                    User.deleted: True,
                    User.deleted_at: datetime.utcnow(),
                },
                synchronize_session=False,
            )
        )
        db.session.commit()
    return redirect(url_for("institution_approval_center"))

@app.route("/institution/bulk_disapprove", methods=["POST"])
@login_required
@roles_required("admin")
def institution_bulk_disapprove():
    user_ids = request.form.getlist("user_ids")
    if user_ids:
        current = get_current_user()

        (
            filter_by_institution(User.query, User)
            .filter(User.id.in_(user_ids))
            .filter(User.role.in_(["student", "subadmin"]))
            .update(
                {
                    User.approved: False,
                    User.approval_status: "Pending",
                    User.account_status: "inactive",

                    User.approved_by: None,
                    User.approved_at: None,
                },
                synchronize_session=False,
            )
        )
        db.session.commit()
    return redirect(url_for("institution_approval_center"))


@app.route( "/institution/approvals/<int:user_id>/approve", methods=["POST"] )
@login_required
@roles_required("admin")
def institution_approve_user(user_id):

    current = get_current_user()

    user = get_record_in_scope(
        User,
        user_id
    )

    if not user:
        abort(404)
    if user.role not in ("student", "subadmin"):
        abort(403)
    if user.deleted:
        flash("User no longer exists.", "warning")
        return redirect(url_for("institution_approval_center"))        

    user.approved = True
    user.approval_status = "Approved"
    user.account_status = "active"
    user.approved_by = current.id
    user.approved_at = datetime.utcnow()

    db.session.commit()

    flash(
        "User approved successfully.",
        "success"
    )

    return redirect(
        url_for("institution_approval_center")
    )

@app.route( "/institution/approvals/<int:user_id>/reject", methods=["POST"] )
@login_required
@roles_required("admin")
def institution_reject_user(user_id):
    current = get_current_user()
    user = get_record_in_scope( User, user_id )
    if not user:
        abort(404)
    if user.deleted:
        flash("User no longer exists.", "warning")
        return redirect(url_for("institution_approval_center"))        
    if user.role not in ("student", "subadmin"):
        abort(403)
    #==========Soft delete==========
    user.approved = False
    user.approval_status = "Rejected"
    user.account_status = "inactive"
    user.approved_by = None
    user.approved_at = None
    user.deleted = True
    user.deleted_at = datetime.utcnow()

    db.session.commit()

    flash(
        "Registration rejected.",
        "success"
    )

    return redirect( url_for("institution_approval_center") )


@app.route( "/institution/approvals/<int:user_id>/disapprove", methods=["POST"]
)
@login_required
@roles_required("admin")
def institution_disapprove_user(user_id):

    current = get_current_user()

    user = get_record_in_scope(
        User,
        user_id
    )

    if not user:
        abort(404)
    if user.deleted:
        flash("User no longer exists.", "warning")
        return redirect(url_for("institution_approval_center"))        
    if user.role not in ("student", "subadmin"):
        abort(403)        

    user.approved = False
    user.approval_status = "Pending"
    user.account_status = "inactive"
    user.approved_by = None
    user.approved_at = None

    db.session.commit()

    flash(
        "User moved back to pending.",
        "success"
    )

    return redirect(
        url_for("institution_approval_center")
    )



# =========================================================
# APPROVE USER
# =========================================================
@app.route("/admin/approvals/<int:user_id>/approve", methods=["POST"])
@login_required
@roles_required("global_admin")
def approve_user(user_id):
    current = get_current_user()
    if current.role != "global_admin":
        abort(403)
    user = User.query.filter_by( id=user_id, deleted=False ).first_or_404()
    free_institution = Institution.query.filter_by( is_system=True, deleted=False
    ).first()
    if not free_institution:
        abort(500, "Free institution not found.")
    if user.role == "admin":
        pass
    #Prevent Global Admin from approving students of another institution.
    elif user.role in ("student", "subadmin"):
        if user.institution_id != free_institution.id:
            abort(403)
    else:
        abort(403)
        
        return redirect(url_for("approval_center"))
    if user.approved:
        flash(
            "User has already been approved.",
            "warning"
        )
        return redirect(url_for("approval_center"))
    # =====================================================
    # Institution Administrator Approval
    # =====================================================
    if user.role == "admin":
        try:
            institution = resolve_requested_institution(
                user,
                manual_institution_id=request.form.get(
                    "institution_id",
                    type=int
                )
            )
        except ValueError as e:
            flash(
                str(e),
                "danger"
            )
            return redirect(
                url_for("approval_center")
            )
        activate_institution_admin(
            user=user,
            institution=institution,
            approved_by=current )
    # =====================================================
    # Student / Sub-admin Approval
    # =====================================================
    else:
        user.approved = True
        user.approval_status = "Approved"
        user.account_status = "active"
        user.approved_by = current.id
        user.approved_at = datetime.utcnow()
    # =====================================================
    # Save Changes
    # =====================================================
    db.session.commit()
    flash( "User approved successfully.", "success" )
    return redirect( url_for("approval_center") )

# =========================================================
# REJECT USER
# =========================================================
@app.route("/admin/approvals/<int:user_id>/reject", methods=["POST"])
@login_required
@roles_required("global_admin")
def reject_user(user_id):
    current = get_current_user()
    if current.role != "global_admin":
        abort(403)
    user = User.query.filter_by( id=user_id, deleted=False ).first_or_404()
    free_institution = Institution.query.filter_by( is_system=True,
        deleted=False ).first()

    if not free_institution:
        abort(500, "Free institution not found.")
    if user.role == "admin":
        # Institution Admins may belong to any institution.
        pass

    elif user.role in ("student", "subadmin"):

        if user.institution_id != free_institution.id:
            abort(403)

    else:
        abort(403)
    
    if user.approved:
        flash( "Approved users cannot be rejected.", "warning" )
        return redirect(url_for("approval_center"))
    #===========Soft delete=========
    user.approved = False
    user.approval_status = "Rejected"
    user.account_status = "inactive"
    user.deleted = True
    user.deleted_at = datetime.utcnow()
    user.approved_by = None
    user.approved_at = None
    db.session.commit()
    flash( "Registration rejected.", "success" )
    return redirect(url_for("approval_center"))


# =========================================================
# DISAPPROVE USER
# =========================================================
@app.route( "/admin/approvals/<int:user_id>/disapprove", methods=["POST"] )
@login_required
@roles_required("global_admin")
def disapprove_user(user_id):
    current = get_current_user()
    if not current.is_global:
        abort(403)
    user = User.query.filter_by( id=user_id, deleted=False ).first_or_404()
    free_institution = Institution.query.filter_by( is_system=True,
        deleted=False ).first()
    if not free_institution:
        abort(500, "Free institution not found.")
    if user.is_institution_admin:
        pass
    elif user.role in ("student", "subadmin"):
        if user.institution_id != free_institution.id:
            abort(403)
    else:
        abort(403)    
    # -----------------------------------------
    # Validation
    # -----------------------------------------    
    if not user.approved:
        flash( "User is already pending approval.", "warning" )
        return redirect(url_for("approval_center"))
    # -----------------------------------------
    # Protect Global Administrators
    # -----------------------------------------
    if user.is_global:
        flash( "Global Administrators cannot be disapproved.", "danger" )
        return redirect(url_for("approval_center"))
    # -----------------------------------------
    # Protect Global Licenser
    # -----------------------------------------
    if user.role == "licenser":
        flash( "Global Licenser cannot be disapproved.", "danger" )
        return redirect(url_for("approval_center"))
    # -----------------------------------------
    # Institution Administrator
    # Convert back to pending Institution Admin
    # -----------------------------------------
    if user.is_institution_admin:
        # Revert to pending Institution Administrator
        user.role = "admin"
        # Remove current assignment.
        # The Global Admin will assign it again
        # during the next approval.
        if user.institution_id:
            user.requested_existing_institution_id = user.institution_id
        user.institution_id = None
        # Leave these intact so they appear again
        # in the Pending Approvals table.
        #
        # requested_institution_name
        # requested_institution_code
        # requested_existing_institution_id
    # -----------------------------------------
    # Reset Approval Status
    # -----------------------------------------
    user.approved = False
    user.approval_status = "Pending"
    user.account_status = "inactive"
    user.approved_by = None
    user.approved_at = None
    # -----------------------------------------
    # Save
    # -----------------------------------------
    db.session.commit()
    flash(
        "User has been returned to the pending approval queue.",
        "success"
    )
    return redirect(url_for("approval_center"))    

# =====================================================
# Manage Institutions
# =====================================================
@app.route("/admin/institutions", methods=["GET", "POST"])
@login_required
@roles_required("global_admin")
def manage_institutions():
    user = get_current_user()
    if current.role != "global_admin":
        abort(403)    
    if request.method == "POST":
        name = request.form.get( "name", "" ).strip()
        code = request.form.get( "code", "" ).strip().upper()
        description = request.form.get( "description", "" ).strip()
        if not name:
            flash( "Institution name is required.", "danger" )
            return redirect( url_for("manage_institutions") )
        existing = Institution.query.filter(
            db.func.lower( Institution.name ) == name.lower() ).first()
        if existing:
            flash( "Institution already exists.", "warning")
            return redirect( url_for("manage_institutions") )
        institution = Institution(
            name=name,
            code=code if code else None,
            description=description,
            active=True
        )
        db.session.add( institution )
        db.session.commit()
        flash( "Institution created successfully.", "success"  )
        return redirect( url_for("manage_institutions") )
    institutions = (
        Institution.query.filter_by( deleted=False )
        .order_by( Institution.name )
        .all() )
    institution_stats = []
    for inst in institutions:
        institution_stats.append({
            "institution": inst,
            "students": User.query.filter_by(
                institution_id=inst.id,
                role="student",
                deleted=False
            ).count(),
            "admins": User.query.filter(
                User.institution_id == inst.id,
                User.role.in_([
                    "admin",
                    "subadmin"
                ]),
                User.deleted == False
            ).count(),
            "subjects": Subject.query.filter_by(
                institution_id=inst.id,
                deleted=False
            ).count(),
            "groups": SubjectGroup.query.filter_by(
                institution_id=inst.id,
                deleted=False
            ).count()
        })    
    return render_template(
        "institutions.html",
        user=user,
        institution_stats=institution_stats)

# =========================================================
# DELETE INSTITUTION ADMINISTRATOR
# =========================================================
@app.route( "/admin/institution-admins/delete/<int:user_id>", methods=["POST"] )
@login_required
@roles_required("global_admin")
def delete_institution_admin(user_id):
    current = get_current_user()
    if current.role != "global_admin" or not current.is_global:
        abort(403)
    #===========Load only non-deleted records============        
    admin = User.query.filter_by( id=user_id, deleted=False ).first_or_404()
    # -----------------------------------------
    # Validation
    # -----------------------------------------
    if admin.role != "admin" or not admin.is_institution_admin:
        flash(
            "Selected user is not an Institution Administrator.",
            "danger"
        )
        return redirect(url_for("approval_center"))
    if admin.is_temporary_admin:
        flash(
            "Temporary administrators cannot be deleted from here.",
            "warning"
        )
        return redirect(
            url_for("approval_center")
        )
    # -----------------------------------------
    # Move original admin to recycle bin
    # -----------------------------------------
    admin.deleted = True
    admin.deleted_at = datetime.utcnow()
    admin.account_status = "inactive"
    # -----------------------------------------
    # Create replacement administrator
    # -----------------------------------------
    create_temporary_institution_admin(admin)
    db.session.commit()
    flash(
        "Institution Administrator moved to recycle bin. "
        "A temporary administrator has been created.",
        "success"
    )
    return redirect( url_for("approval_center") )

# =========================================================
# RESTORE INSTITUTION ADMINISTRATOR
# =========================================================
@app.route( "/admin/institution-admins/restore/<int:user_id>", methods=["POST"])
@login_required
@roles_required("global_admin")
def restore_institution_admin(user_id):
    current = get_current_user()
    if current.role != "global_admin":
        abort(403)
    admin = User.query.filter_by( id=user_id, deleted=True ).first_or_404()
    # -----------------------------------------
    # Validation
    # -----------------------------------------
    if admin.role != "admin" or not admin.is_institution_admin:
        flash(
            "Selected user is not an Institution Administrator.",
            "danger"
        )
        return redirect(url_for("recycle_bin"))
    # -----------------------------------------
    # Restore original administrator
    # -----------------------------------------
    admin.deleted = False
    admin.deleted_at = None
    admin.account_status = "active"
    # -----------------------------------------
    # Remove temporary replacement
    # -----------------------------------------
    temp_admin = User.query.filter_by( replaced_admin_id=admin.id,
        is_temporary_admin=True,
        deleted=False
    ).first()
    if temp_admin:
        temp_admin.deleted = True
        temp_admin.deleted_at = datetime.utcnow()
        temp_admin.account_status = "inactive"
    db.session.commit()
    flash(
        "Institution Administrator restored successfully.", "success" )
    return redirect( url_for("recycle_bin") )

# =========================================================
# PERMANENTLY DELETE INSTITUTION ADMINISTRATOR
# =========================================================
@app.route( "/admin/institution-admins/permanent-delete/<int:user_id>",
    methods=["POST"] )
@login_required
@roles_required("global_admin")
def permanently_delete_institution_admin(user_id):
    current = get_current_user()
    if current.role != "global_admin" or not current.is_global:
        abort(403)
    admin = User.query.filter_by( id=user_id, deleted=True ).first_or_404()
    # -----------------------------------------
    # Must be an Institution Administrator
    # -----------------------------------------
    if admin.role != "admin" or not admin.is_institution_admin:
        flash(
            "Selected user is not an Institution Administrator.",
            "danger"
        )
        return redirect(url_for("recycle_bin"))
    # -----------------------------------------
    # Safety check
    # -----------------------------------------
    if not admin.is_temporary_admin:
        replacement = User.query.filter_by(
            replaced_admin_id=admin.id,
            is_temporary_admin=True
        ).first()
        if replacement and not replacement.deleted:
            flash(
                "Restore or remove the temporary administrator first.",
                "warning"
            )
            return redirect(url_for("recycle_bin"))
    # -----------------------------------------
    # Permanent delete
    # -----------------------------------------
    db.session.delete(admin)
    db.session.commit()
    flash(
        "Institution Administrator permanently deleted.", "success" )
    return redirect(url_for("recycle_bin"))

# =====================================================
# Edit Institution
# =====================================================
@app.route("/admin/institutions/edit/<int:institution_id>", methods=["POST"])
@login_required
@roles_required("global_admin")
def edit_institution(institution_id):
    current = get_current_user()
    if current.role != "global_admin" or not current.is_global:
        abort(403)    
    institution = Institution.query.filter_by( id=institution_id,
        deleted=False ).first_or_404()
    if institution.is_system:
        flash(
            "The Free Institution cannot be edited.",
            "warning"
        )
    return redirect(url_for("manage_institutions"))    
    name = request.form.get( "name", "" ).strip()
    code = request.form.get( "code", "" ).strip().upper()
    description = request.form.get( "description", "" ).strip()
    active = ( request.form.get("active") == "on" )
    if not name:
        flash( "Institution name is required.", "danger" )
        return redirect( url_for("manage_institutions") )
    duplicate = Institution.query.filter(
        Institution.id != institution.id,
        db.func.lower(Institution.name) == name.lower(),
        Institution.deleted == False
    ).first()
    if duplicate:
        flash( "Another institution already uses that name.", "warning" )
        return redirect( url_for("manage_institutions") )
    institution.name = name
    institution.code = code if code else None
    institution.description = description
    institution.active = active
    db.session.commit()
    flash( "Institution updated successfully.", "success" )
    return redirect( url_for("manage_institutions") )

# =========================================================
# CREATE INSTITUTION
# =========================================================
@app.route("/admin/institutions/create", methods=["POST"])
@login_required
@roles_required("global_admin")
def create_institution():
    current = get_current_user()
    if current.role != "global_admin" or not current.is_global:
        abort(403)    
    name = request.form.get( "name", "" ).strip()
    code = request.form.get( "code", "" ).strip().upper()
    description = request.form.get( "description", "" ).strip()
    if not name:
        flash( "Institution name is required.", "danger" )
        return redirect( url_for("manage_institutions") )
    exists = Institution.query.filter(
        db.func.lower(Institution.name) == name.lower(),
        Institution.deleted == False ).first()
    if exists:
        flash( "Institution already exists.", "warning" )
        return redirect( url_for("manage_institutions") )
    institution = Institution( name=name, code=code if code else None,
        description=description, active=True )
    db.session.add( institution )
    db.session.commit()
    # ============================================
    # CREATE DEFAULT SCHOOL SETTINGS
    # ============================================
    new_settings = SchoolSettings(
        institution_id=institution.id,
        academic_mode="nigerian_secondary",
        term="A",
        default_report_design="default.html",
        ca1_percentage=10,
        ca2_percentage=10,
        ca3_percentage=10,
        objective_percentage=50,
        theory_percentage=50
    )

    db.session.add(new_settings)
    
    create_default_academic_classes( institution.id )    
    flash( "Institution created successfully.", "success" )
    return redirect( url_for("manage_institutions") )

# =========================================================
# TOGGLE INSTITUTION STATUS
# =========================================================
@app.route("/admin/institutions/toggle/<int:institution_id>")
@login_required
@roles_required("global_admin")
def toggle_institution(institution_id):
##    institution = Institution.query.get_or_404( institution_id )
    institution = Institution.query.filter_by( id=institution_id,
        deleted=False ).first_or_404()
    current = get_current_user()
    if current.role != "global_admin" or not current.is_global:
        abort(403)    
    if institution.is_system:
        flash( "The Free institution cannot be disabled.", "warning" )
        return redirect( url_for("manage_institutions") )
    institution.active = not institution.active
    db.session.commit()
    flash( "Institution updated.", "success" )
    return redirect( url_for("manage_institutions") )

# =========================================================
# DELETE INSTITUTION
# =========================================================
@app.route("/admin/institutions/delete/<int:institution_id>")
@login_required
@roles_required("global_admin")
def delete_institution(institution_id):
##    institution = Institution.query.get_or_404( institution_id )
    institution = Institution.query.filter_by( id=institution_id,
        deleted=False ).first_or_404()
    current = get_current_user()
    if current.role != "global_admin" or not current.is_global:
        abort(403)    
    if institution.is_system:
        flash( "The Free institution cannot be deleted.", "warning" )
        return redirect( url_for("manage_institutions") )
    users = User.query.filter_by(
        institution_id=institution.id, deleted=False ).count()
    if users:
        flash( "Institution still has users.", "danger" )
        return redirect( url_for("manage_institutions") )
    institution.deleted = True
    institution.deleted_at = datetime.utcnow()
    db.session.commit()
    flash( "Institution moved to recycle bin.", "success" )
    return redirect( url_for("manage_institutions") )

@app.route("/admin/institutions/restore/<int:institution_id>")
@login_required
@roles_required("global_admin")
def restore_institution(institution_id):
##    institution = Institution.query.get_or_404(institution_id)
    institution = Institution.query.filter_by( id=institution_id,
        deleted=True ).first_or_404()
    current = get_current_user()
    if current.role != "global_admin" or not current.is_global:
        abort(403)    
    institution.deleted = False
    institution.deleted_at = None
    db.session.commit()
    flash("Institution restored.", "success")
    return redirect(url_for("recycle_bin"))



# (DASHBOARD/USERS/SUBJECTS/GROUPS 908) (Chat 2175) (License 2355)    
# (Questions 2719) (Exam 3466)(Report Card 4182) (DEBUG 5138)
# ================ DASHBOARD/USERS/SUBJECTS/GROUPS/ADMIN SETTINGS =====================
# ================ DASHBOARD/USERS/SUBJECTS/GROUPS/ADMIN SETTINGS =====================
# ================ DASHBOARD/USERS/SUBJECTS/GROUPS/ADMIN SETTINGS =====================
# ================ DASHBOARD/USERS/SUBJECTS/GROUPS/ADMIN SETTINGS =====================
# ================ DASHBOARD/USERS/SUBJECTS/GROUPS/ADMIN SETTINGS =====================
# ================ DASHBOARD/USERS/SUBJECTS/GROUPS/ADMIN SETTINGS =====================
# global pagination config (fallback = 20)
# Global pagination config
PAGINATION_CONFIG = {"per_page": 20}

from sqlalchemy import func, or_

@app.route('/', methods=["GET", "POST"])
@login_required
def dashboard():

    if request.args.get("theory_timeout") == "1":
        flash(
            "Exam ended due to time-out. Typed answers were saved automatically. Uploaded files not previously saved may be lost.",
            "warning"
        )

    u = get_current_user()
    unread_count = 0

    if not u:
        return redirect(url_for("login"))

    search_query = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = PAGINATION_CONFIG.get("per_page", 20)

    # =====================================================
    # ================= ADMIN PAGINATION ==================
    # =====================================================
    if (
        (u.role == "global_admin" and u.is_global)
        or (u.role == "admin" and u.is_institution_admin)
    ) and request.method == "POST":

        new_per_page = request.form.get("per_page", type=int)

        if new_per_page and new_per_page > 0:
            PAGINATION_CONFIG["per_page"] = new_per_page

            flash(
                f"Pagination updated to {new_per_page} per page.",
                "success"
            )

        return redirect(url_for("dashboard"))

    # =====================================================
    # ================= STUDENT DASHBOARD =================
    # =====================================================
    if u.role == "student":

        # =========================================
        # HIDE SUBJECTS FROM STUDENTS
        # =========================================
        subjects = (
            filter_by_institution(
                Subject.query,
                Subject
            )
            .filter(
                Subject.hidden_from_students == False,
                Subject.deleted == False,
                or_(
                    Subject.class_level == u.student_class,
                    Subject.result_type == "practice"
                )
            )
            .order_by(Subject.name)
            .all()
        )

        # ensure label always exists safely
        for s in subjects:
            if not hasattr(s, "label"):
                s.label = None

        db.session.expire_all()

        from sqlalchemy.orm import selectinload

        groups_query = (
            filter_by_institution(
                SubjectGroup.query,
                SubjectGroup
            )
            .options(
                selectinload(SubjectGroup.subjects),
                selectinload(
                    SubjectGroup.children
                ).selectinload(
                    SubjectGroup.children
                )
            )
            .order_by(SubjectGroup.name)
        )

        if search_query:
            groups_query = groups_query.filter(
                SubjectGroup.name.ilike(f"%{search_query}%")
            )

##        groups_paginated = groups_query.paginate(
##            page=page,
##            per_page=per_page,
##            error_out=False
##        )
##
##        groups = groups_paginated

        groups = (
            filter_by_institution(
                SubjectGroup.query,
                SubjectGroup
            )
            .filter(
                SubjectGroup.deleted == False
            )
            .options(
                selectinload(SubjectGroup.subjects),
                selectinload(SubjectGroup.children),
                selectinload(SubjectGroup.parent)
            )
            .order_by(SubjectGroup.name)
            .all()
        )

        # =========================================
        # FILTER HIDDEN SUBJECTS INSIDE GROUPS
        # =========================================
        for group in groups:

            group.subjects = [
                s for s in group.subjects
                if (
                    not s.hidden_from_students
                    and not s.deleted
                    and (
                        s.class_level == u.student_class
                        or s.result_type == "practice"
                    )
                )
            ]

            for s in group.subjects:
                if not hasattr(s, "label"):
                    s.label = None

        # Hide deleted children
        for group in groups:
            group.children = [
                child
                for child in group.children
                if not child.deleted
            ]

        # Root groups
        roots = [ g for g in groups if g.parent is None or g.parent.deleted ]
        #==students always message their institution's administrator==
        admin = ( filter_by_institution(User.query, User)
            .filter(
                User.role == "admin",
                User.deleted == False,
                User.is_institution_admin == True,
            ) .first() )
        unread_map = get_unread_count(receiver_id=u.id)
        unread_count = ( unread_map.get(admin.id, 0) if admin else 0 )
        my_sessions = ( filter_by_institution( ExamSession.query, ExamSession )
            .filter_by( user_id=u.id, completed=False )
            .order_by( ExamSession.started_at.desc() )
            .all() )
        my_results = ( filter_by_institution( Result.query, Result )
            .filter_by( user_id=u.id )
            .order_by( Result.taken_at.desc() )
            .paginate( page=page, per_page=per_page, error_out=False ) )
        # =========================================
        # DEDUPLICATE RESULTS BY SESSION
        # =========================================
        seen = set()
        unique = []
        for r in my_results.items:
            if r.session_id not in seen:
                unique.append(r)
                seen.add(r.session_id)
        my_results.items = unique
        return render_template(
            "dashboard.html", user=u, subjects=subjects, groups=groups,
            sessions=my_sessions, results=my_results, students=None,
            per_page=per_page, unread_count=unread_count, admin=admin,
            search_query=search_query, roots=roots, now=datetime.now() )

    # =====================================================
    # ============== ADMIN / SUBADMIN DASHBOARD ===========
    # =====================================================
    print("=" * 80)
    print("DASHBOARD USER")
    print("Username:", u.username)
    print("Role:", u.role)
    print("Is Global:", u.is_global)
    print("Institution:", u.institution_id)
    print("=" * 80)    
    if (
        (u.role == "global_admin" and u.is_global)
        or (u.role == "admin" and u.is_institution_admin)
        or u.role == "subadmin"
    ):
        subject_query = ( filter_by_institution( Subject.query, Subject )
            .filter(Subject.deleted == False) )
        if u.role == "subadmin":
            subject_query = subadmin_subject_filter( subject_query, u )
        subjects = ( subject_query
            .order_by(Subject.name)
            .all() )
        groups = (
            filter_by_institution(
                SubjectGroup.query.filter_by( deleted=False ), SubjectGroup )
            .order_by(SubjectGroup.name)
            .all() )
        results_q = (
            filter_by_institution(Result.query, Result)
            .filter(Result.deleted == False)
            .join(Result.subject)
            .join(Result.user) )
        if u.role == "subadmin":
            allowed_group_ids = json_load( u.allowed_group_ids_json, [] )
            results_q = results_q.filter(
                Subject.group_id.in_( allowed_group_ids ) )
        if search_query:
            full_name_concat = func.concat(
                func.coalesce(User.first_name, ""), " ",
                func.coalesce(User.last_name, "") )
            results_q = results_q.filter(
                or_(
                    User.username.ilike( f"%{search_query}%" ),
                    full_name_concat.ilike( f"%{search_query}%" ) ) )
        results_paginated = ( results_q
            .order_by(Result.taken_at.desc())
            .paginate( page=page, per_page=per_page, error_out=False ) )
        # =========================================
        # MAP STUDENTS → RESULTS
        # =========================================
        students_dict = {}
        for r in results_paginated.items:
            students_dict.setdefault( r.user, [] ).append(r)
        unread_count = (
            filter_by_institution(ChatMessage.query, ChatMessage)
            .filter(
                ChatMessage.receiver_id == u.id,
                ChatMessage.read == False,
                ChatMessage.deleted == False or None,
            ) .count() )
            # ============================
        # DASHBOARD STATISTICS
        # ============================
        student_count = ( filter_by_institution( User.query, User )
            .filter_by( role="student", deleted=False )
            .count() )
        subject_count = (
            filter_by_institution(Subject.query, Subject)
            .filter(Subject.deleted == False)
            .count() )
        group_count = (
            filter_by_institution(SubjectGroup.query, SubjectGroup)
            .filter(SubjectGroup.deleted == False)
            .count()
        )
        result_count = results_q.count()
        
        return render_template(
            "dashboard.html", user=u, subjects=subjects, groups=groups,
            sessions=[], results=results_paginated, students=students_dict,
            per_page=per_page, unread_count=unread_count,
            search_query=search_query, student_count=student_count,
            subject_count=subject_count, group_count=group_count,
            result_count=result_count, now=datetime.now(), )
    # =====================================================
    # ================= FORBIDDEN ==========================
    # =====================================================
    abort(403)


# =========================================================
# ADMIN SETTINGS
# =========================================================
@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin')
def admin_settings():
    current = get_current_user()
    settings = get_institution_settings()
    
    if current.role == "global_admin":
        pass
    elif current.role == "admin" and current.is_institution_admin:
        pass
    else:
        abort(403)    

    # =========================================
    # MAIL SETTINGS FORM
    # =========================================
    if (
        request.method == 'POST'
        and 'admin_smtp_server' in request.form
    ):

        provider = request.form.get('admin_smtp_provider', 'custom')
        admin_email = request.form.get( "admin_email", "" ).strip()        

        smtp_server = request.form.get('admin_smtp_server', '').strip()
        smtp_port = request.form.get('admin_smtp_port', '587').strip()
        username = request.form.get('admin_smtp_username', '').strip()
        password = request.form.get('admin_smtp_password', '').strip()
        use_tls = request.form.get('admin_smtp_use_tls') == "on"

        settings.admin_email = encrypt_value(admin_email)
        settings.smtp_provider = provider
        settings.smtp_server = smtp_server
        settings.smtp_port = smtp_port
        settings.smtp_username = encrypt_value(username)
        settings.auto_forward_scores = bool(
            request.form.get("auto_forward_scores")
        )

        if password:
            settings.smtp_password = encrypt_value(password)

        settings.smtp_use_tls = use_tls

        if current.is_global and "app_id" in request.form:
            app_id = request.form.get("app_id", "").strip().upper()

            if app_id:
                set_setting("app_id", app_id)
                flash("App ID updated successfully.", "success")

            return redirect(url_for("admin_settings"))        
        db.session.commit()

        flash("Admin SMTP settings updated.", "success")
        return redirect(url_for("admin_settings"))

    # =========================================
    # PAGE DISPLAY
    # =========================================
    print("=" * 80)
    print("ADMIN DASHBOARD")
    print("=" * 80)    
    return render_template(
        'admin_settings.html',
        verify_report=session.get("verify_report"),
        admin_email=decrypt_value(settings.admin_email or ""),
        smtp_provider=settings.smtp_provider,
        smtp_server=settings.smtp_server,
        smtp_port=settings.smtp_port,
        smtp_username=decrypt_value(settings.smtp_username or ""),
        smtp_password=decrypt_value(settings.smtp_password or ""),
        smtp_use_tls=settings.smtp_use_tls,
        user=get_current_user(),
        app_id=get_setting("app_id", ""),
        can_edit_app_id=current.is_global,
        auto_forward=settings.auto_forward_scores
    )


# ---------- Admin: Groups ----------
@app.route('/admin/groups', methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin')
def manage_groups():
    current = get_current_user()
    if current.role == "global_admin":
        pass
    elif current.role == "admin" and current.is_institution_admin:
        pass
    else:
        abort(403)    

    def generate_unique_group_name(base_name, exclude_id=None):
        """
        Generates a globally unique group name.

        Example:
            Maths
            Maths (1)
            Maths (2)
            Maths (3)
        """

        name = base_name
        counter = 1

        while True:
                                                                                                                                                                                                              
            q = (
                filter_by_institution( SubjectGroup.query, SubjectGroup
                )
                .filter(
                    SubjectGroup.deleted == False,
                    SubjectGroup.name == name ) )

            if exclude_id is not None:
                q = q.filter(SubjectGroup.id != exclude_id)

            if not q.first():
                return name

            name = f"{base_name} ({counter})"
            counter += 1

    if request.method == 'POST':

        action = request.form.get('action')

        # =====================================================
        # CREATE
        # =====================================================
        if action == 'create':

            name = request.form.get('name', '').strip()

            parent_id = request.form.get('parent_id')
##            parent_id = int(parent_id) if parent_id else None
            if parent_id:
                parent = get_record_in_scope( SubjectGroup, int(parent_id) )
                if not parent:
                    abort(404)
                parent_id = parent.id
            else:
                parent_id = None
            

            if name:

                # Always generate a unique name
                name = generate_unique_group_name(name)

                g = SubjectGroup(
                    name=name,
                    institution_id=current_institution_id(),
                    hide_scores_for_subadmins=bool(
                        request.form.get('hide_scores') ),
                    parent_id=parent_id
                )

                db.session.add(g)

                try:
                    db.session.commit()

                except IntegrityError as e:

                    db.session.rollback()

                    print("=" * 80)
                    print(e)
                    print("=" * 80)

                    flash(
                        "⚠ Unable to create group.",
                        "error"
                    )

        # =====================================================
        # UPDATE
        # =====================================================
        elif action == 'update':

            gid = int(request.form['id'])

            g = get_record_in_scope(SubjectGroup,
                gid
            )            

            new_name = request.form.get(
                'name',
                g.name
            ).strip()

            pid = request.form.get("parent_id")
            if pid:
                parent = get_record_in_scope(SubjectGroup, int(pid))
                if not parent:
                    abort(404)
                new_parent = parent.id
            else:
                new_parent = None

            # Always generate a unique name
            new_name = generate_unique_group_name(
                new_name,
                exclude_id=g.id
            )

            g.name = new_name
            g.parent_id = new_parent
            g.hide_scores_for_subadmins = bool(
                request.form.get('hide_scores')
            )

            try:
                db.session.commit()

            except IntegrityError as e:

                db.session.rollback()

                print("=" * 80)
                print(e)
                print("=" * 80)

                flash(
                    "⚠ Unable to update group.",
                    "error"
                )

        # =====================================================
        # DELETE
        # =====================================================
        elif action == 'delete':

            gid = int(request.form['id'])

            g = get_record_in_scope( SubjectGroup, gid )        

            g.deleted = True
            g.deleted_at = datetime.utcnow()

            db.session.commit()

            print("=" * 80)
            print(g.id)
            print(g.name)
            print(g.deleted)
            print("=" * 80)

            flash(
                f"Group '{g.name}' Moved to recycle bin (auto-delete in 3 months).",
                "warning"
            )

        return redirect(url_for('manage_groups'))

    # =====================================================
    # DISPLAY GROUPS
    # =====================================================
    groups = (
        filter_by_institution(
            SubjectGroup.query,
            SubjectGroup
        )
        .filter(SubjectGroup.deleted == False)
        .options(
            joinedload(SubjectGroup.children),
            joinedload(SubjectGroup.parent),
            joinedload(SubjectGroup.subjects)
        )
        .order_by(SubjectGroup.name)
        .all()
    )

    for group in groups:
        group.children = [ c for c in group.children if not c.deleted ]

        group.subjects = [ s for s in group.subjects if not s.deleted ]
    roots = [
        g for g in groups
        if g.parent is None or g.parent.deleted
    ]


    print("\n================ GROUPS ================\n")

    for g in groups:
        print(
            f"id={g.id:<3} "
            f"name={g.name:<35} "
            f"parent_id={g.parent_id}"
        )

    print("\n============== ROOT GROUPS ==============\n")

    for g in groups:
        if g.parent_id is None:
            print(g.id, g.name)

    print("\n=========================================\n")

    return render_template(
        "groups.html",
        groups=groups,
        roots=roots,
        user=get_current_user()
    )
    

# ---------- Admin/Subadmin: Subjects ----------
@app.route('/admin/subjects', methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin', 'subadmin')
def manage_subjects():

    u = get_current_user()

    # =====================================================
    # POST
    # =====================================================
    if request.method == 'POST':

        action = request.form.get('action')

        # =================================================
        # CREATE
        # =================================================
        if action == 'create' and u.role == 'admin':

            name = request.form.get('name', '').strip()
            session_val = request.form.get('session', '').strip()

            result_type = request.form.get(
                'result_type',
                'exam'
            ).strip().lower()

            order = request.form.get('order', '1ST')
            term = request.form.get('term', 'A')

            label = f"{session_val} {result_type.upper()} {term} {order}"

            try:
                duration = int(
                    request.form.get('duration', 60)
                )
            except ValueError:
                duration = 60

            try:
                theory_duration = int(
                    request.form.get('theory_duration', 30)
                )
            except ValueError:
                theory_duration = 30

            group_id = request.form.get("group_id")
            if group_id:
                group = get_record_in_scope( SubjectGroup, int(group_id) )
                group_id = group.id
            else:
                group_id = None
                
            class_level = request.form.get( 'class_level', type=int )            

            hide_scores = 'hide_scores' in request.form
            hidden_students = 'hidden_from_students' in request.form

            # =========================================
            # PRACTICE OVERRIDE
            # =========================================
            if result_type == 'practice':
                show_answers = True
            else:
                show_answers = (
                    request.form.get(
                        'show_answers_to_students'
                    ) == '1'
                )

            # =========================================
            # SERIES FIX
            # =========================================
            is_series = request.form.get('is_series') == '1'            

            subj = Subject(
                name=name,
                label=label,
                session=session_val,
                term=term,
                result_type=result_type,
                duration_minutes=duration,
                theory_duration_minutes=theory_duration,
                group_id=group_id,
                hide_scores_for_subadmins=hide_scores,
                hidden_from_students=hidden_students,
                show_answers_to_students=show_answers,
                series_group=name if is_series else None,
                institution_id=current_institution_id(),
                class_level=class_level
            )


            db.session.add(subj)
            db.session.commit()
            print("=" * 80)
            print("SUBJECT CREATED")
            print("ID:", subj.id)
            print("Name:", subj.name)
            print("Institution:", subj.institution_id)
            print("=" * 80)            

            flash(
                f"Subject '{name}' created successfully.",
                "success"
            )

        # =================================================
        # UPDATE
        # =================================================
        elif action == 'update':

            sid = int(request.form['id'])

            subj = get_record_in_scope( Subject, sid )

            new_name = request.form.get(
                'name',
                ''
            ).strip()

            session_val = request.form.get(
                'session',
                ''
            ).strip()

            result_type = request.form.get(
                'result_type',
                'exam'
            ).strip().lower()

            term = request.form.get('term', 'A')
            order = request.form.get('order', '1ST')

            new_label = (
                f"{session_val} "
                f"{result_type.upper()} "
                f"{term} "
                f"{order}"
            )

            subj.name = new_name
            subj.label = new_label
            subj.result_type = result_type
            subj.session = session_val
            subj.term = term            

            try:
                subj.duration_minutes = int(
                    request.form.get(
                        'duration',
                        subj.duration_minutes
                    )
                )
            except ValueError:
                pass                
            try:
                subj.theory_duration_minutes = int(
                    request.form.get(
                        'theory_duration',
                        subj.theory_duration_minutes or 30
                    )
                )               
            except ValueError:
                pass
            #==Prevent someone moving subject into another institution==
            gid = request.form.get("group_id")
            if gid:
                group = get_record_in_scope( SubjectGroup, int(gid)
                )
                subj.group_id = group.id
            else:
                subj.group_id = None
    
            class_level = request.form.get( 'class_level', type=int )            
            subj.class_level = class_level
            subj.hide_scores_for_subadmins = (
                'hide_scores' in request.form
            )

            subj.hidden_from_students = (
                'hidden_from_students' in request.form
            )

            # =========================================
            # SERIES UPDATE FIX
            # =========================================
            is_series = (
                request.form.get('is_series') == '1'
            )

            if is_series:
                subj.series_group = new_name
            else:
                subj.series_group = None

            # =========================================
            # PRACTICE FIX
            # =========================================
            if result_type == 'practice':

                subj.show_answers_to_students = True

            else:

                subj.show_answers_to_students = (
                    request.form.get(
                        'show_answers_to_students'
                    ) == 'on'
                )

            db.session.commit()

            flash( "Updated successfully", "success" )

        # =================================================
        # DELETE
        # =================================================
        elif action == 'delete' and u.role == 'admin':

            sid = int(request.form['id'])

            subj = get_record_in_scope( Subject, sid)

            subj.deleted = True
            subj.deleted_at = datetime.utcnow()

            db.session.commit()

            flash(f"'{subj.name}' Moved to recycle bin (auto-delete in 3 months).",
                  "warning")

        return redirect(url_for('manage_subjects'))

    # =====================================================
    # GET
    # =====================================================
    page = request.args.get( 'page', 1, type=int )

    search = request.args.get( 'search', '' ).strip()

    print("=" * 80)
    print("CURRENT USER:", u.username)
    print("CURRENT USER INSTITUTION:", u.institution_id)
    print("IS GLOBAL:", u.is_global)
    print("=" * 80)    

##    base_query = filter_by_institution( Subject.query.filter(
##            Subject.deleted == False ), Subject )
    base_query = ( filter_by_institution( Subject.query, Subject )
        .filter(Subject.deleted == False) )   

    query = ( base_query if u.role == 'admin' else subadmin_subject_filter(
            base_query, u ) )

    if search:
        query = query.filter(
            Subject.name.ilike(f"%{search}%")
        )

    print("=" * 80)
    print("ALL SUBJECTS IN DATABASE")

    for s in Subject.query.all():
        print(
            s.id,
            s.name,
            s.institution_id,
            s.deleted
        )

    print("=" * 80)

    all_subjects = query.order_by(
        Subject.name.asc(),
        Subject.id.desc()
    ).all()

    from collections import defaultdict
    from math import ceil

    grouped = defaultdict(list)

    # =========================================
    # GROUP SUBJECTS
    # =========================================
    for s in all_subjects:

        key = (
            s.series_group
            if s.series_group
            else s.name
        )

        grouped[key].append(s)

    # =========================================
    # HELPERS
    # =========================================
    def extract_term(label):

        if not label:
            return '-'

        parts = label.split()

        if len(parts) < 3:
            return '-'

        # =====================================
        # SERIES FORMAT
        # 2025/2026 PRACTICE Series 2 1ST
        # =====================================
        if parts[2].lower() == 'series':

            if len(parts) >= 4:
                return f"{parts[2]} {parts[3]}"

        # =====================================
        # NORMAL FORMAT
        # 2025/2026 EXAM A 1ST
        # =====================================
        return parts[2]


    import re

    def sort_key(subj):

        label = (subj.label or "").upper()

        session = 0

        m = re.match(r"(\d{4})/(\d{4})", label)

        if m:
            session = int(m.group(1))

        return (
            -session,   # Newest session first
            -subj.id    # Newest subject first within that session
        )

    # =========================================
    # BUILD DISPLAY SUBJECTS
    # =========================================
    subjects = []

    SUB_PER_PAGE = 10

    for name, items in grouped.items():

        items = sorted(
            items,
            key=sort_key
        )

        parent = items[0]

        parent.term = extract_term(
            parent.label
        )

        # =====================================
        # CHILD PAGINATION
        # =====================================
        sub_page = request.args.get(
            f"subpage_{parent.id}",
            1,
            type=int
        )

        total_children = max(
            len(items) - 1,
            0
        )

        child_start = (
            (sub_page - 1)
            * SUB_PER_PAGE
        )

        child_end = (
            child_start
            + SUB_PER_PAGE
        )

        parent.sub_subjects = (
            items[1:][child_start:child_end]
        )

        # =====================================
        # FIX CHILD TERMS
        # =====================================
        for child in parent.sub_subjects:
            child.term = extract_term(
                child.label
            )

        parent.sub_page = sub_page

        parent.sub_total = total_children

        parent.sub_pages = (
            ceil(total_children / SUB_PER_PAGE)
            if total_children > 0
            else 1
        )

        subjects.append(parent)

    # =========================================
    # PARENT PAGINATION
    # =========================================
    start = (page - 1) * 10
    end = start + 10

    display_subjects = subjects[start:end]

    # =========================================
    # SIMPLE PAGINATION
    # =========================================
    class SimplePagination:

        def __init__(
            self,
            page,
            total,
            per_page
        ):
            self.page = page
            self.per_page = per_page
            self.total = total

            self.pages = ceil(
                total / per_page
            )

        @property
        def has_prev(self):
            return self.page > 1

        @property
        def has_next(self):
            return self.page < self.pages

        @property
        def prev_num(self):
            return self.page - 1

        @property
        def next_num(self):
            return self.page + 1

        def iter_pages(self):
            return range(
                1,
                self.pages + 1
            )

    pagination = SimplePagination(
        page,
        len(subjects),
        10
    )
    groups = (
        filter_by_institution( SubjectGroup.query, SubjectGroup )
        .filter(
            SubjectGroup.deleted == False
        )
        .order_by(SubjectGroup.name)
        .all() )
    print("=" * 80)
    print("GROUPS SENT TO TEMPLATE")
    for g in groups:
        print(g.id, g.name, g.deleted)
    print("=" * 80)    

    return render_template(
        'subjects.html',
        subjects=display_subjects,
        groups=groups,
        user=u,
        pagination=pagination,
        search=search,
        theory_default=30
    )

####
# ---------- Admin: Users ----------
from sqlalchemy import or_, func
import json

@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin')
def manage_users():
    u = get_current_user()
    # ================= POST =================
    if request.method == 'POST':
        action = request.form.get('action')
        # =====================================================
        # ================= CREATE USER =======================
        # =====================================================
        if action == 'create':
            username = request.form['username'].strip()
            role = request.form['role']
            pw = request.form.get('password', '123456')
            email = request.form.get('email', '').strip()
            first_name = request.form.get('first_name', '').strip()
            last_name = request.form.get('last_name', '').strip()

            registration_number = (
                request.form.get(
                    'registration_number'
                ) or ""
            ).strip()

            age = request.form.get(
                'age',
                type=int
            )

            if username and role in ('admin', 'subadmin', 'student'):
                # =====================================
                # CHECK DUPLICATE USERNAME
                # =====================================
                if (
                    filter_by_institution( User.query, User )
                    .filter_by( username=username, deleted=False )
                    .first() ):
                    flash('Username already exists.', 'error')
                    return redirect(url_for('manage_users'))

                # =====================================
                # CHECK DUPLICATE EMAIL
                # =====================================
                if role == 'student' and email:
                    if (
                        filter_by_institution( User.query, User )
                        .filter_by( email=email, deleted=False )
                        .first() ):
                        flash('Email already exists.', 'error')
                        return redirect(url_for('manage_users'))

                # =====================================
                # LOAD ACADEMIC CLASS
                # =====================================
                academic_class_id = request.form.get(
                    'academic_class_id',
                    type=int
                )

                academic_class = None

                if academic_class_id:
                    academic_class = get_record_in_scope(
                        AcademicClass,
                        academic_class_id
                    )

                # =====================================
                # CREATE USER
                # =====================================
                new_u = User(
                    username=username,
                    role=role,
                    email=email or None,
                    first_name=first_name or None,
                    last_name=last_name or None,
                    institution_id=u.institution_id
                )

                if academic_class:
                    new_u.student_class = academic_class.class_code
                    new_u.academic_class_id = academic_class.id
                else:
                    new_u.student_class = None
                    new_u.academic_class_id = None

                new_u.registration_number = registration_number
                new_u.age = age

                new_u.set_password(pw)

                db.session.add(new_u)
                db.session.commit()
                if role == "student" and new_u.academic_class:
                    settings = (
                        filter_by_institution( SchoolSettings.query,
                            SchoolSettings )
                        .first() )
                    record_student_class_history(
                        student=new_u,
                        academic_class=new_u.academic_class,
                        session=settings.current_session if settings else "",
                        term=settings.current_term if settings else "",
                        change_type="Initial Admission",
                        remarks="Student admitted",
                        changed_by=u.id
                    )
                    db.session.commit()                

                # =====================================
                # AUTO LICENSE ASSIGNMENT
                # =====================================
                flash(
                    f'User {username} created successfully.',
                    'success'
                )
                return redirect(url_for('manage_users'))
        # =====================================================
        # ================= RESET PASSWORD ====================
        # =====================================================
        elif action == 'reset':
            reset_key = request.form.get( 'reset_key', '' )
            main_admin = (
                filter_by_institution( User.query, User )
                .filter_by(
                    username="admin",
                    role="admin",
                    deleted=False
                ) .first() )
            if ( not main_admin or
                not main_admin.check_password(reset_key) ):
                flash( 'Invalid admin password.', 'error' )
                return redirect( url_for('manage_users') )
            u_reset = get_record_in_scope( User, int(request.form['id']) )
            u_reset.set_password( request.form.get( 'new_password', '123456'))
            db.session.commit()
            flash( "Password reset successful", "success" )
        # =====================================================
        # ================= DELETE USER =======================
        # =====================================================
        elif action == 'delete':
            uid = int(request.form['id'])
            if uid == u.id:
                flash( "You cannot delete yourself.", "error" )
            else:
                del_u = get_record_in_scope( User, uid)
                del_u.deleted = True
                del_u.deleted_at = datetime.utcnow()
                db.session.commit()
                flash( "User moved to the recycle bin.", "success" )
        # =====================================================
        # ================= UPDATE USER =======================
        # =====================================================
        elif action == 'update_user':
            uid = int(request.form['id'])
            edit_u = get_record_in_scope(
                User,
                uid
            )
            edit_u.first_name = (
                request.form.get(
                    'first_name',
                    ''
                ).strip() or None
            )
            edit_u.last_name = (
                request.form.get(
                    'last_name',
                    ''
                ).strip() or None
            )
            edit_u.username = request.form.get(
                'username',
                ''
            ).strip()
            edit_u.email = (
                request.form.get(
                    'email',
                    ''
                ).strip() or None
            )
            edit_u.role = request.form.get(
                'role',
                'student'
            )
            registration_number = (
                request.form.get(
                    'registration_number'
                ) or ""
            ).strip()
            age = request.form.get(
                'age',
                type=int
            )
            # =====================================
            # LOAD ACADEMIC CLASS
            # =====================================
            academic_class_id = request.form.get(
                'academic_class_id',
                type=int
            )
            academic_class = None
            if academic_class_id:
                academic_class = get_record_in_scope(
                    AcademicClass,
                    academic_class_id
                )
            if academic_class:
                edit_u.student_class = academic_class.class_code
                edit_u.academic_class_id = academic_class.id
            else:
                edit_u.student_class = None
                edit_u.academic_class_id = None
            edit_u.registration_number = registration_number
            edit_u.age = age
            # =====================================
            # PASSWORD (OPTIONAL)
            # =====================================
            new_password = request.form.get(
                'password',
                ''
            ).strip()
            if new_password:
                edit_u.set_password(new_password)
            db.session.commit()
            flash(
                'User updated successfully',
                'success' )
        # =====================================================
        # ================= UPDATE GROUPS =====================
        # =====================================================
        elif action == 'update_groups':
            uid = int(request.form['id'])
            edit_u = get_record_in_scope( User, uid )
            
            allowed = []
            for gid in request.form.getlist("allowed_groups"):
                group = get_record_in_scope( SubjectGroup, int(gid) )
                allowed.append(group.id)
            edit_u.allowed_group_ids_json = json.dumps(allowed)
            
            db.session.commit()
            flash( "Groups updated successfully", "success" )
        # =====================================================
        # ================= UPDATE SUBJECTS ===================
        # =====================================================
        elif action == 'update_subjects':
            uid = int(request.form['id'])
            edit_u = get_record_in_scope( User, uid )
            
            allowed = []
            for sid in request.form.getlist("allowed_subjects"):
                subject = get_record_in_scope( Subject, int(sid) )
                allowed.append(subject.id)
            edit_u.allowed_subject_ids_json = json.dumps(allowed)
            
            db.session.commit()
            flash( "Subjects updated successfully", "success" )
        return redirect( url_for('manage_users') )
    # ================= GET =================
    search = request.args.get( 'search', '' ).strip()
    filter_students = request.args.get( 'filter_students' )
    filter_subadmins = request.args.get( 'filter_subadmins' )
    if filter_students is None and filter_subadmins is None:
        filter_students = "1"
        filter_subadmins = "0"
    page = request.args.get( 'page', 1, type=int )
    per_page = 50
    query = filter_by_institution( User.query.filter_by(deleted=False), User )
    if search:
        full_name = func.concat(
            func.coalesce( User.first_name, '' ), ' ',
            func.coalesce( User.last_name, '' ) )
        query = query.filter(
            or_(
                User.username.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%"),
                User.registration_number.ilike(f"%{search}%"),
                full_name.ilike(f"%{search}%") ) )
    roles = ['admin']
    if filter_students == "1":
        roles.append("student")
    elif filter_subadmins == "1":
        roles.append("subadmin")
    query = query.filter( User.role.in_(roles) )
    users = query.order_by( User.id.desc() ).paginate( page=page,
        per_page=per_page, error_out=False )
    subjects = ( filter_by_institution( Subject.query.filter_by(
        deleted=False ), Subject )
        .order_by(Subject.name)
        .all() )
    groups = ( filter_by_institution( SubjectGroup.query.filter_by(
                deleted=False ), SubjectGroup )
        .order_by(SubjectGroup.name)
        .all() )
    # =========================================
    # SAFE JSON LOADER
    # =========================================
    def safe_json(value):
        try:
            return json.loads(value) if value else []
        except:
            return []
    academic_classes = get_academic_classes()        
    return render_template(
        'users.html',
        users=users,
        subjects=subjects,
        groups=groups,
        search=search,
        filter_students=filter_students,
        filter_subadmins=filter_subadmins,
        json_load=safe_json,
        user=u,
        academic_classes=academic_classes,
        recent_user_ids=[
            x.id
            for x in (
                filter_by_institution(
                    User.query.filter_by(
                        deleted=False
                    ),
                    User
                )
                .order_by(User.id.desc())
                .limit(5)
                .all()
            )
        ]
    )

#==================Academic class====================
@app.route("/admin/academic_classes", methods=["GET", "POST"])
@roles_required("global_admin", "admin")
def admin_academic_classes():
    if request.method == "POST":
        class_name = ( request.form.get("class_name") or ""
        ).strip()
        class_code = request.form.get( "class_code", type=int )
        display_order = request.form.get( "display_order", type=int )
        education_level = ( request.form.get("education_level") or ""
        ).strip()
        active = bool( request.form.get("active") )
        if not class_name:
            flash( "Class name is required.", "danger" )
            return redirect( url_for("admin_academic_classes") )
        existing = (
            filter_by_institution( AcademicClass.query, AcademicClass )
            .filter_by( class_name=class_name )
            .first() )
        if existing:
            flash( "A class with that name already exists.", "warning" )
            return redirect( url_for("admin_academic_classes") )
        existing_code = (
            filter_by_institution(AcademicClass.query, AcademicClass)
            .filter_by(class_code=class_code)
            .first() )
        if existing_code:
            flash("A class with that code already exists.", "warning")
            return redirect(url_for("admin_academic_classes"))
        
        new_class = AcademicClass(
            institution_id=current_institution_id(),
            class_name=class_name,
            class_code=class_code,
            display_order=display_order
            if display_order is not None
            else class_code,
            education_level=education_level,
            active=active
        )
        db.session.add(new_class)
        db.session.commit()
        flash( "Academic class created successfully.", "success" )
        return redirect( url_for("admin_academic_classes") )
    classes = (
        filter_by_institution( AcademicClass.query, AcademicClass )
        .order_by( AcademicClass.display_order, AcademicClass.class_name )
        .all() )
    return render_template( "academic_classes.html", classes=classes )

#=================Edit Academic Class===============
@app.route( "/admin/academic_classes/<int:class_id>/edit", methods=["POST"] )
@roles_required("global_admin", "admin")
def edit_academic_class(class_id):
    academic_class = get_record_in_scope( AcademicClass, class_id )
    academic_class.class_name = ( request.form.get("class_name") or ""
    ).strip()
    academic_class.class_code = request.form.get( "class_code", type=int )
    
    duplicate = (
        filter_by_institution(AcademicClass.query, AcademicClass)
        .filter(
            AcademicClass.id != academic_class.id,
            AcademicClass.class_code == academic_class.class_code )
        .first() )
    if duplicate:
        flash( "Another class already uses that class code.", "warning" )
        return redirect(url_for("admin_academic_classes"))

    academic_class.display_order = request.form.get( "display_order",
        type=int )
    academic_class.education_level = (
        request.form.get("education_level") or ""
    ).strip()
    academic_class.active = bool( request.form.get("active") )
    db.session.commit()
    flash( "Academic class updated successfully.", "success" )
    return redirect( url_for("admin_academic_classes") )

#=============Delete Academic Class==============
@app.route( "/admin/academic_classes/<int:class_id>/delete", methods=["POST"] )
@roles_required("global_admin", "admin")
def delete_academic_class(class_id):
    academic_class = get_record_in_scope( AcademicClass, class_id )
    # -----------------------------------
    # Prevent deleting if students exist
    # -----------------------------------
    student_count = (
        filter_by_institution( User.query, User )
        .filter( User.academic_class_id == academic_class.id )
        .count() )
    if student_count:
        flash(
            "This class cannot be deleted because students are assigned to it.",
            "danger" )
        return redirect( url_for("admin_academic_classes") )
    # -----------------------------------
    # Prevent deleting if history exists
    # -----------------------------------
    history_count = (
        filter_by_institution(
            StudentClassHistory.query, StudentClassHistory )
        .filter(
            StudentClassHistory.academic_class_id  == academic_class.id )
        .count() )
    if history_count:
        flash(
            "This class cannot be deleted because it exists in student academic history.",
            "danger" )
        return redirect( url_for("admin_academic_classes") )
    db.session.delete( academic_class )
    db.session.commit()
    flash( "Academic class deleted successfully.", "success" )
    return redirect( url_for("admin_academic_classes") )

#=============Create the Classes Page==============
@app.route("/classes")
@roles_required("global_admin", "admin", "subadmin")
def classes():
    academic_classes = (
        filter_by_institution( AcademicClass.query, AcademicClass )
        .filter_by(active=True)
        .order_by( AcademicClass.display_order, AcademicClass.class_name )
        .all() )
    class_data = []
    for academic_class in academic_classes:
        students = get_students_in_academic_class( academic_class.id )
        summary = get_academic_class_summary( academic_class.id )
        class_data.append({
            "academic_class": academic_class,
            "students": students,
            "summary": summary
        })
    return render_template( "classes.html", class_data=class_data )

#===============Promotion=================
@app.route( "/classes/<int:class_id>/promote", methods=["GET", "POST"] )
@roles_required("global_admin", "admin")
def promote_academic_class(class_id):
    academic_class = get_record_in_scope( AcademicClass, class_id )
    students = get_students_in_academic_class( academic_class.id )
    settings = (
        filter_by_institution( SchoolSettings.query, SchoolSettings
        ).first() )
    if request.method == "POST":
        promoted = 0
        skipped = 0
        for student in students:
            ok, _ = promote_student(
                student,
                session=(settings.current_session if settings else ""),
                term=(settings.current_term if settings else ""),
                changed_by=get_current_user().id,
                remarks="Bulk promotion" )
            if ok:
                promoted += 1
            else:
                skipped += 1
        db.session.commit()
        flash( f"{promoted} students promoted. {skipped} skipped.", "success"
        )
        return redirect( url_for("classes") )
    next_class = get_next_academic_class( academic_class )
    return render_template(
        "promote_class.html",
        academic_class=academic_class,
        next_class=next_class,
        students=students
    )

#==============Academic History===============
@app.route( "/classes/<int:class_id>/history" )
@roles_required("global_admin", "admin", "subadmin")
def academic_class_history(class_id):
    academic_class = get_record_in_scope( AcademicClass, class_id )
    history = (
        filter_by_institution( StudentClassHistory.query, StudentClassHistory)
        .filter_by( academic_class_id=academic_class.id )
        .order_by( StudentClassHistory.changed_at.desc() )
        .all() )
    return render_template( "academic_class_history.html",
        academic_class=academic_class, history=history )

#============Student Transfer============
@app.route( "/classes/<int:class_id>/transfer", methods=["GET", "POST"] )
@roles_required("global_admin", "admin")
def transfer_students(class_id):
    academic_class = get_record_in_scope( AcademicClass, class_id )
    students = get_students_in_academic_class( academic_class.id )
    destination_classes = (
        filter_by_institution( AcademicClass.query, AcademicClass )
        .filter( AcademicClass.active == True,
            AcademicClass.id != academic_class.id )
        .order_by( AcademicClass.display_order )
        .all() )
    settings = (
        filter_by_institution( SchoolSettings.query, SchoolSettings
        ).first() )
    if request.method == "POST":
        destination_class_id = request.form.get(
            "destination_class_id", type=int )
        destination_class = get_record_in_scope(
            AcademicClass, destination_class_id )
        selected_students = request.form.getlist( "student_ids" )
        transferred = 0
        skipped = 0
        if not destination_class:
            flash( "Please select a valid destination class.", "danger" )
            return redirect( request.url )
        if not selected_students:
            flash( "Please select at least one student.", "warning" )
            return redirect( request.url )        
        for sid in selected_students:
            student = get_record_in_scope( User, int(sid) )
            ok, _ = transfer_student( student=student,
                destination_class=destination_class,
                session=( settings.current_session if settings else "" ),
                term=( settings.current_term if settings else "" ),
                changed_by=get_current_user().id,
                remarks="Manual transfer" )
            if ok:
                transferred += 1
            else:
                skipped += 1
        db.session.commit()
        flash(
            f"{transferred} student(s) transferred successfully. "
            f"{skipped} skipped.",
            "success"
        )
        return redirect( url_for( "classes" ) )
    return render_template(
        "transfer_students.html",
        academic_class=academic_class,
        students=students,
        destination_classes=destination_classes
    )


# (DASHBOARD/USERS/SUBJECTS/GROUPS 908) (Chat 2175) (License 2355)    
# (Questions 2719) (Exam 3466)(Report Card 4182) (DEBUG 5138)
#======================CHAT===========================
#======================CHAT===========================
#======================CHAT===========================
#======================CHAT===========================
#======================CHAT===========================
#======================CHAT===========================
# ------------------ CHAT VIEW (SHARED) ------------------
def render_chatbox(user1, user2, messages, pagination=None):
    """Helper to render the same chatbox for student/admin"""
    return render_template('chat_view.html',
                           me=user1,
                           peer=user2,
                           messages=messages,
                           pagination=pagination)

# ------------------ STUDENT CHAT (WITH ADMIN) -----endpoint--#to make students dashboard url_for('chat', ... work without changing links
@app.route("/chat", methods=["GET", "POST"], endpoint="chat")
@login_required
@roles_required("student")
def student_chat():
    u = get_current_user()
    admin = ( filter_by_institution( User.query, User )
        .filter_by( role="admin", deleted=False )
        .first() )
    if not admin:
        abort(404, "No admin found")
    if request.method == "POST":
        text = request.form.get("message", "").strip()
        if text:
            msg = ChatMessage(
                sender_id=u.id,
                receiver_id=admin.id,
                message=text,
                institution_id=current_institution_id()
            )
            db.session.add(msg)
            db.session.commit()

            # Notify admin in real-time
            socketio.emit(
                "new_message",
                {
                    "sender_id": u.id,
                    "receiver_id": admin.id,
                    "message": text,
                    "timestamp": msg.timestamp.strftime("%Y-%m-%d %H:%M"),
                },
                room=f"user_{admin.id}"
            )
            return redirect(url_for("chat"))
    # ✅ Pagination for messages
    page = request.args.get("page", 1, type=int)
    per_page = 50
##    pagination = ChatMessage.query.filter(
##        ((ChatMessage.sender_id == u.id) & (ChatMessage.receiver_id == admin.id)) |
##        ((ChatMessage.sender_id == admin.id) & (ChatMessage.receiver_id == u.id))
##    ).order_by(ChatMessage.timestamp).paginate(page=page, per_page=per_page,
##                                               error_out=False)
    pagination = ( filter_by_institution( ChatMessage.query, ChatMessage )
        .filter( (
                (ChatMessage.sender_id == u.id) &
                (ChatMessage.receiver_id == admin.id)
            ) | (
                (ChatMessage.sender_id == admin.id) &
                (ChatMessage.receiver_id == u.id)
            ) )
        .order_by( ChatMessage.timestamp )
        .paginate( page=page, per_page=per_page, error_out=False ) )    
    messages = pagination.items
    # Mark admin → student messages as read
##    ChatMessage.query.filter_by( receiver_id=u.id, sender_id=admin.id,
##        read=False ).update({"read": True})
    filter_by_institution( ChatMessage.query, ChatMessage
    ).filter_by( receiver_id=u.id, sender_id=admin.id, read=False
    ).update( {"read": True} )    
    db.session.commit()
    # Update unread badge for admin
    new_unread = get_unread_count(receiver_id=admin.id, peer_id=u.id)
    socketio.emit(
        "unread_count",
        {"from": u.id, "to": admin.id, "count": new_unread},
        room=f"user_{admin.id}"
    )
    # ✅ FIX: Use render_template, not render_template_string
    return render_template(
        "chat.html",       # <- template file
        messages=messages,
        pagination=pagination,
        user=u,
        peer=admin
    )

# ---------------- Admin Chat ----------------
@app.route("/support_chat/<int:student_id>", methods=["GET", "POST"])
@login_required
@roles_required("global_admin", "admin")
def support_chat(student_id):
    u = get_current_user()
##    student = User.query.get_or_404(student_id)
    student = get_record_in_scope( User, student_id )
    if request.method == "POST":
        text = request.form.get("message", "").strip()
        if text:
            msg = ChatMessage(
                sender_id=u.id,
                receiver_id=student.id,
                message=text,
                institution_id=current_institution_id()
            )
            db.session.add(msg)
            db.session.commit()

            # Notify student in real-time
            socketio.emit(
                "new_message",
                {
                    "sender_id": u.id,
                    "receiver_id": student.id,
                    "message": text,
                    "timestamp": msg.timestamp.strftime("%Y-%m-%d %H:%M"),
                },
                room=f"user_{student.id}"
            )
            return redirect(url_for("support_chat", student_id=student.id))

    # ✅ Pagination for messages
    page = request.args.get("page", 1, type=int)
    per_page = 50
##    pagination = ChatMessage.query.filter(
##        ((ChatMessage.sender_id == u.id) & (ChatMessage.receiver_id == student.id)) |
##        ((ChatMessage.sender_id == student.id) & (ChatMessage.receiver_id == u.id))
##    ).order_by(ChatMessage.timestamp).paginate(page=page, per_page=per_page,
##                                               error_out=False)
    pagination = ( filter_by_institution( ChatMessage.query, ChatMessage )
        .filter( (
                (ChatMessage.sender_id == u.id) &
                (ChatMessage.receiver_id == student.id)
            ) | (
                (ChatMessage.sender_id == student.id) &
                (ChatMessage.receiver_id == u.id)
            ) )
        .order_by( ChatMessage.timestamp )
        .paginate( page=page, per_page=per_page, error_out=False ) )    
    messages = pagination.items

    # Mark as read
##    ChatMessage.query.filter_by(
##        receiver_id=u.id, sender_id=student.id, read=False
##    ).update({"read": True})
    filter_by_institution( ChatMessage.query, ChatMessage
    ).filter_by( receiver_id=u.id, sender_id=student.id, read=False
    ).update( {"read": True} )    
    db.session.commit()

    # Update unread badge for student
    new_unread = get_unread_count(receiver_id=student.id, peer_id=u.id)
    socketio.emit(
        "unread_count",
        {"from": u.id, "to": student.id, "count": new_unread},
        room=f"user_{student.id}"
    )

    return render_template('chat.html',
                           messages=messages,
                           pagination=pagination,
                           user=u,
                           peer=student)

@app.route("/admin/chat_dashboard")
@login_required
@roles_required("global_admin", "admin")
def admin_chat_dashboard():
##    students = User.query.filter_by(role="student", deleted=False).all()
    students = ( filter_by_institution( User.query, User )
        .filter_by( role="student", deleted=False )
        .all() )    
    chats = []
    admin = get_current_user()

    for s in students:
        # Get last message between student and admin
##        last_msg = ChatMessage.query.filter(
##            ((ChatMessage.sender_id == s.id) & (ChatMessage.receiver_id == admin.id)) |
##            ((ChatMessage.sender_id == admin.id) & (ChatMessage.receiver_id == s.id))
##        ).order_by(ChatMessage.timestamp.desc()).first()
        last_msg = ( filter_by_institution( ChatMessage.query, ChatMessage )
            .filter( (
                    (ChatMessage.sender_id == s.id) &
                    (ChatMessage.receiver_id == admin.id)
                ) | (
                    (ChatMessage.sender_id == admin.id) &
                    (ChatMessage.receiver_id == s.id)
                ) )
            .order_by( ChatMessage.timestamp.desc() )
            .first() )        

        # Count unread messages student → admin
##        unread_count = ChatMessage.query.filter_by(
##            sender_id=s.id, receiver_id=admin.id, read=False ).count()
        unread_count = (
            filter_by_institution( ChatMessage.query, ChatMessage )
            .filter_by( sender_id=s.id, receiver_id=admin.id, read=False )
            .count() )        

        chats.append({
            "id": s.id,
            "username": s.username,
            "last_message": last_msg.message if last_msg else None,
            "last_time": last_msg.timestamp if last_msg else None,
            "unread_count": unread_count,
        })

    # Sort AFTER all chats have been collected
    chats.sort(
        key=lambda c: (
            c["unread_count"] == 0,
            -(c["last_time"].timestamp() if c["last_time"] else 0)
        )
    )

    # Paginate AFTER sorting
    page = request.args.get("page", 1, type=int)
    per_page = 100
    start = (page - 1) * per_page
    end = start + per_page
    total = len(chats)
    pages = (total + per_page - 1) // per_page
    chats_page = chats[start:end]
    return render_template(
        "admin_chat_dashboard.html",
        chats=chats_page,
        page=page,
        pages=pages,
        total=total,
        per_page=per_page,
        user=admin,
    )



# (DASHBOARD/USERS/SUBJECTS/GROUPS 908) (Chat 2175) (License 2355)    
# (Questions 2719) (Exam 3466)(Report Card 4182) (DEBUG 5138)
# ============= AUTO LICENSE CONFIG ==============
# ============= AUTO LICENSE CONFIG ==============
# ============= AUTO LICENSE CONFIG ==============
# ============= AUTO LICENSE CONFIG ==============
# ============= AUTO LICENSE CONFIG ==============
# ============= AUTO LICENSE CONFIG ==============
@app.route('/licensing/auto_activate', methods=['GET', 'POST'])
@login_required
@roles_required('licensing')
def auto_activate_students():

    print("ACCESS GRANTED TO:", get_current_user().username)

    config = AutoLicenseConfig.query.first()

    # ❌ DO NOT auto-create empty config (caused NOT NULL email crash)

    if request.method == 'POST':

        count = int(request.form.get('count', 0))
        period = int(request.form.get('period', 365))
        email = (request.form.get('email') or '').strip()

        # 🔥 IMPORTANT SAFETY CHECK
        if not email:
            flash("Email is required for auto activation.", "error")
            return redirect(url_for('auto_activate_students'))

        # create config ONLY when valid data exists
        if not config:
            config = AutoLicenseConfig(
                remaining_count=count,
                valid_days=period,
                email=email,
                enabled=(count > 0)
            )
            db.session.add(config)

        else:
            config.remaining_count = count
            config.valid_days = period
            config.email = email
            config.enabled = True if count > 0 else False

        db.session.commit()

        flash(f"Auto activation enabled for next {count} students.", "success")
        return redirect(url_for('license_dashboard'))

    return render_template(
        'auto_activate.html',
        user=get_current_user(),
        config=config
    )


#---------------Auto license request---------------------
@app.route('/admin/activate_auto_license', methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin')
def activate_auto_license():
    if request.method == 'POST':
        key = (request.form.get('key') or '').strip()
        req = AutoLicenseRequest.query.filter_by(
            request_key=key,
            activated=False
        ).first()
        if not req:
            flash(
                "Invalid or already used key.",
                "error"
            )
            return redirect( url_for('activate_auto_license') )
        if not req.is_auto_license:
            flash(
                "This activation key is not an auto-license request.",
                "error"
            )
            return redirect( url_for('activate_auto_license') )        
        config = AutoLicenseConfig.query.first()
        # 🔥 only create if missing AND req has valid email
        if not config:
            config = AutoLicenseConfig(
                remaining_count=req.count,
                valid_days=req.valid_days,
                email=req.email,
                enabled=True
            )
            db.session.add(config)
        else:
            config.remaining_count = req.count
            config.valid_days = req.valid_days
            config.email = req.email
            config.enabled = True
        req.activated = True
        db.session.commit()
        flash(f"Auto licensing activated for next {req.count} students.", "success")
        return redirect(url_for('license_dashboard'))
    return render_template( 'activate_auto_license.html',
                            user=get_current_user() )



# =========================================================
# LICENSER SETTINGS PAGE
# =========================================================

@app.route('/licensing/settings', methods=['GET', 'POST'])
@login_required
@roles_required('licensing')
def licenser_settings():

    if request.method == 'POST':

        # =================================================
        # LICENSER EMAIL
        # =================================================
        lic_email = request.form.get(
            'licenser_email',
            ''
        ).strip()

        lic_password = request.form.get(
            'licenser_password',
            ''
        ).strip()

        # =================================================
        # SMTP SETTINGS
        # =================================================
        smtp_server = request.form.get(
            'smtp_server',
            'smtp.gmail.com'
        ).strip()

        smtp_port = request.form.get(
            'smtp_port',
            '587'
        ).strip()

        # =================================================
        # GLOBAL FORWARDING EMAIL
        # =================================================
        forwarding_email = request.form.get(
            'forwarding_email',
            ''
        ).strip()

        # =================================================
        # APP ID
        # =================================================
        app_id = request.form.get(
            'app_id',
            ''
        ).strip()

        # =================================================
        # SAVE SETTINGS
        # =================================================

        set_setting(
            "licenser_email",
            encrypt_value(lic_email)
        )

        set_setting(
            "licenser_app_password",
            encrypt_value(lic_password)
        )

        set_setting(
            "licenser_smtp_server",
            smtp_server
        )

        set_setting(
            "licenser_smtp_port",
            smtp_port
        )

        # ⚠️ FIXED KEY CONSISTENCY (IMPORTANT)
        set_setting(
            "forwarding_email",
            encrypt_value(forwarding_email)
        )

        set_setting(
            "app_id",
            app_id
        )

        flash(
            "Licensing settings updated successfully.",
            "success"
        )

        return redirect(url_for('licenser_settings'))

    # =====================================================
    # PAGE LOAD
    # =====================================================

    return render_template(
        'licenser_settings.html',

        licenser_email=decrypt_value(
            get_setting("licenser_email", "")
        ),

        licenser_password=decrypt_value(
            get_setting("licenser_app_password", "")
        ),

        smtp_server=get_setting(
            "licenser_smtp_server",
            "smtp.gmail.com"
        ),

        smtp_port=get_setting(
            "licenser_smtp_port",
            "587"
        ),

        # ✅ FIXED: same key as POST
        forwarding_email=decrypt_value(
            get_setting("forwarding_email", "")
        ),

        app_id=get_setting(
            "app_id",
            "EXAM-APP-001"
        ),

        user=get_current_user()
    )


#----------------Licensing -----------------------
@app.route('/license_dashboard')
@login_required
def license_dashboard():
    u = get_current_user()
    # only admin or licensing
    if not u or u.role not in ['global_admin', 'licensing', 'admin']:
        abort(403)
    search_query = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 100  # 👈 fixed pagination size
##    licenses_query = License.query.join(User, isouter=True)
    licenses_query = ( filter_by_institution( License.query, License )
        .join(User, isouter=True) )    
    if search_query:
        full_name_concat = func.concat(
            func.coalesce(User.first_name, ''),
            ' ',
            func.coalesce(User.last_name, '')
        )
        licenses_query = licenses_query.filter(
            or_(
                User.username.ilike(f"%{search_query}%"),
                User.first_name.ilike(f"%{search_query}%"),
                User.last_name.ilike(f"%{search_query}%"),
                full_name_concat.ilike(f"%{search_query}%")
            ))
    licenses_paginated = licenses_query.order_by(
        case(
            (License.user_id.is_(None), 0),   # unassigned first
            else_=1
        ),License.created_at.desc()
    ).paginate(page=page,per_page=per_page,error_out=False)
##    auto_configs = ( AutoLicenseConfig.query
##        .order_by(AutoLicenseConfig.created_at.desc())
##        .all() )
    auto_configs = (
        filter_by_institution( AutoLicenseConfig.query, AutoLicenseConfig )
        .order_by( AutoLicenseConfig.created_at.desc() )
        .all() )    
    return render_template(
        'license_dashboard.html',
        user=u,
        licenses=licenses_paginated,
        search_query=search_query,
        auto_configs=auto_configs,
        timedelta=timedelta
    )


#Only admin or licenser should be able to reset
@app.route('/reset_device/<int:license_id>', methods=['POST'])
@login_required
def reset_device(license_id):
    u = get_current_user()
    if not u or u.role != "licensing":  #not in ["admin", "licenser"]:
        abort(403)   # Only licensing role can reset devices

##    lic = License.query.get_or_404(license_id)
    lic = get_record_in_scope( License, license_id )    
    lic.device_id = None   # 🔓 unbind device
    lic.reset_count = (lic.reset_count or 0) + 1  #+= 1
    db.session.commit()

    flash(f"Device binding reset for license {lic.key}.", "success")
    return redirect(url_for("license_dashboard"))
##    return redirect(url_for("dashboard"))


@app.route('/reset_all_devices', methods=['POST'])
@login_required
def reset_all_devices():
    u = get_current_user()

    if not u or u.role != "licensing":
        abort(403)

    # Reset ALL active licenses
    licenses = License.query.filter(
        License.device_id.isnot(None)
    ).all()

    count = 0

    for lic in licenses:
        lic.device_id = None
        lic.reset_count = (lic.reset_count or 0) + 1
        count += 1

    db.session.commit()

    flash(f"All device bindings reset ({count} licenses affected).", "success")
    return redirect(url_for("license_dashboard"))
##    return redirect(url_for("dashboard"))



# ---------- Licensing: Generate Licenses ----------
@app.route('/licensing/generate', methods=['GET', 'POST'])
@login_required
@roles_required('licensing')
def generate_license():
    if request.method == 'POST':
        count = int(request.form.get("count", 1))
        period = int(request.form.get("period", 365))
        email = request.form.get("email")
        keys = []

        for _ in range(count):
            k = secrets.token_hex(16)
##            lic = License(key=k, valid_days=period)
            lic = License(key=k, email=email, valid_days=period,
                          institution_id=current_institution_id(),
                          generated_by="licenser")
            db.session.add(lic)
            keys.append(k)
        db.session.commit()

        flash(f"{count} license(s) generated. Each lasts {period} days from activation.", "ok")
        return render_template_string("""
            <h2>Generated Licenses</h2>
            <ul>{% for k in keys %}<li><b>{{k}}</b></li>{% endfor %}</ul>
            <p>Save them securely. Expiry starts on activation.</p>
            <a href="{{url_for('license_dashboard')}}">Back</a>
        """, keys=keys, user=get_current_user())
##                <a href="{{url_for('dashboard')}}">Back</a>

    return render_template('gen_license.html',
                           user=get_current_user())




# ----------- Admin License Request Page -----------
@app.route('/admin/request_license', methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin')
def admin_request_license():

    u = get_current_user()
    u_email = u.email

    if request.method == 'POST':

        num = int(request.form['num'])
        duration = int(request.form['duration'])
        e_mail = request.form['email']
        phone = request.form['phone']

        auto_license = request.form.get("auto_license")
        print("AUTO LICENSE =", auto_license)

        # ======================================
        # GENERATE REQUEST KEY (ONLY IF NEEDED)
        # ======================================
##        request_key = secrets.token_hex(16) if auto_license else None
        # ALWAYS generate request key (required by DB constraint)
        request_key = secrets.token_hex(16)

        req = AutoLicenseRequest(
            request_key=request_key,
            count=num,
            valid_days=duration,
            email=e_mail,
            admin_email=u.email,
            is_auto_license=bool(auto_license)
        )

        db.session.add(req)
        db.session.commit()

        # ======================================
        # LICENSER EMAIL (ONLY SOURCE OF TRUTH)
        # ======================================
        licenser_email = decrypt_value(
            get_setting("licenser_email", "")
        )

        if not licenser_email:
            flash("Licenser email has not been configured.", "danger")
            return redirect(url_for("admin_request_license"))

        # ======================================
        # EMAIL BODY BUILDING
        # ======================================
        base_info = f"""
            Admin License Request

            Count: {num}
            Duration: {duration} days

            Admin Email:
            {e_mail}
            {u_email}

            Phone:
            {phone}
            """

        if auto_license:
            email_body = base_info + f"""

            🚨 AUTO LICENSE MODE ENABLED 🚨

            Activation Key:
            {request_key}
            """
        else:
            email_body = base_info

        # ======================================
        # MESSAGE (STRICT: LICENSER PROFILE)
        # ======================================
        msg = Message(
            subject="Admin License Request",
            recipients=[licenser_email],
            body=email_body
        )

        # forward copy (optional)
        forward_email = get_forwarding_email()
        if forward_email:
            msg.bcc = [forward_email]

        # ======================================
        # QUEUE EMAIL USING LICENSER PROFILE
        # ======================================
        try:
            queue_email(
                recipient=licenser_email,
                subject=msg.subject,
                body=msg.body,
                mail_profile="licenser"   # 🔥 IMPORTANT FIX
            )

            flash("✅ License request sent to Licenser email.", "success")

        except Exception as e:
            flash(f"❌ Error sending email: {str(e)}", "danger")

        # ======================================
        # PRICING LOGIC (UNCHANGED)
        # ======================================
        amount = calculate_amount(num, duration)

        if amount is None:
            flash("❌ Invalid duration entered.", "danger")
            return redirect(url_for("admin_request_license"))

        return redirect(
            url_for(
                "pay_license",
                amount=amount,
                num=num,
                duration=duration,
                email=e_mail,
                phone=phone
            )
        )

    return render_template('admin_license_request.html', user=u)

# Extra years are priced automatically using bisect
def calculate_amount(num, duration):
    thresholds = [31, 90, 133, 365] #cutoff days.
    multipliers = [1, 2, 3, 6] #mapped multiplier for each cutoff
    #find where duration fits in threshold, returns index position i (binary search)
    i = bisect.bisect_right(thresholds, duration)
    if i < len(multipliers):
        return num * multipliers[i]

    # unlimited case: scale by years
    extra_years = (duration - thresholds[-1]) // 365 # [-1] last in list, floor division (returns integer)
    return num * (multipliers[-1] + extra_years * 6)

# ----------- Payment Page Placeholder -----------
@app.route('/admin/pay_license')
@login_required
@roles_required('global_admin', 'admin')
def pay_license():
    u = get_current_user()
    # grab query parameters
    #data comes from query string, not database or direct variables
    #→ must request.args.get() it → then pass into template.
    amount = request.args.get('amount')
    num = request.args.get('num')
    duration = request.args.get('duration')
    e_mail = request.args.get('email')
    phone = request.args.get('phone')

    return render_template('pay.html', amount=amount, num=num,
                           duration=duration, e_mail=e_mail,
                           phone=phone, user=u)



# ----- Delete license completely ------
@app.route('/delete_license/<int:license_id>', methods=['POST'])
@login_required
def delete_license(license_id):
    u = get_current_user()
    if not u or u.role not in ['licensing', 'admin']:
        abort(403)  # only licensing/admin can delete

##    lic = License.query.get_or_404(license_id)
    lic = get_record_in_scope( License, license_id )    
    db.session.delete(lic)
    db.session.commit()

    flash(f"🗑️ License {lic.key} deleted.", "success")
    return redirect(url_for('license_dashboard'))
##    return redirect(url_for('dashboard'))


# -----Bulk Delete license completely ------
@app.route('/bulk_delete_licenses', methods=['POST'])
@login_required
def bulk_delete_licenses():

    u = get_current_user()

    if not u or u.role not in ['admin', 'licensing']:
        abort(403)

    ids = request.form.getlist('license_ids')

    if not ids:
        flash("No licenses selected.", "error")
        return redirect(url_for('license_dashboard'))

    licenses = License.query.filter(License.id.in_(ids)).all()

    count = len(licenses)

    for lic in licenses:
        db.session.delete(lic)

    db.session.commit()

    flash(f"{count} licenses deleted successfully.", "success")
    return redirect(url_for('license_dashboard'))



# ---------Route for Students to Activate ---------
@app.route('/activate_license', methods=['GET','POST'])
@login_required
@roles_required('student')
def activate_license():
    if request.method == 'POST':
        key = request.form['key'].strip()
##        lic = License.query.filter_by(key=key, active=True).first()
        lic = ( filter_by_institution( License.query, License )
            .filter_by( key=key, active=True )
            .first() )       
        if not lic:
            flash("Invalid license key.", "error")
            return redirect(url_for('activate_license'))
        # If already linked to another student
        if lic.user_id and lic.user_id != get_current_user().id:
            flash("License already used by another account.", "danger")
            return redirect(url_for('activate_license'))
        # If already locked to a different device
        device_id = get_device_id()
        if lic.device_id and lic.device_id != device_id:
            flash("License already activated on another device.", "danger")
            return redirect(url_for('activate_license'))
        if lic.user_id:
            flash("This key is already used.", "error")
            return redirect(url_for('activate_license'))

        # First-time activation
        lic.activated_at = datetime.utcnow()
        lic.expires_at = lic.activated_at + timedelta(days=lic.valid_days)
        lic.user_id = get_current_user().id
        lic.device_id = device_id   # 🔒 lock to this device
        db.session.commit()

        print("Activating license:", lic.key, "for user:", get_current_user().username)
        flash(f"License activated! Valid until {lic.expires_at.date()}.", "ok")
        return redirect(url_for('dashboard'))

    return render_template('act_license.html',
                           user=get_current_user())



# (DASHBOARD/USERS/SUBJECTS/GROUPS 908) (Chat 2175) (License 2355)    
# (Questions 2719) (Exam 3466)(Report Card 4182) (DEBUG 5138)
# ================= UPLOAD / MANAGE QUESTIONS =================
# ================= UPLOAD / MANAGE QUESTIONS =================
# ================= UPLOAD / MANAGE QUESTIONS =================
# ================= UPLOAD / MANAGE QUESTIONS =================
# ================= UPLOAD / MANAGE QUESTIONS =================

# ==========================================================
# ALLOWED FILE TYPES (GLOBAL)
# ==========================================================
ALLOWED_EXTENSIONS = {

    # ---------- Images ----------
    "png", "jpg", "jpeg", "gif", "bmp",
    "webp", "tif", "tiff", "svg", "ico",

    # ---------- Microsoft Office ----------
    "doc", "docx",
    "xls", "xlsx",
    "ppt", "pptx",

    # ---------- OpenDocument ----------
    "odt", "ods", "odp",

    # ---------- PDF ----------
    "pdf",

    # ---------- Text ----------
    "txt", "csv", "rtf",

    # ---------- Archives ----------
    "zip", "rar", "7z",

    # ---------- Programming ----------
    "py", "js", "html", "css",
    "json", "xml", "sql",
    "cpp", "c", "h",
    "java",
    "cs",
    "php",

    # ---------- Adobe ----------
    "psd",
    "ai",
    "indd",

    # ---------- Corel ----------
    "cdr",

    # ---------- Scratch ----------
    "sb",
    "sb2",
    "sb3",

    # ---------- Android ----------
    "apk",

    # ---------- Audio ----------
    "mp3", "wav", "ogg", "m4a",

    # ---------- Video ----------
    "mp4", "avi", "mov", "mkv", "wmv",

    # ---------- Misc ----------
    "epub",
    "md"
}

def allowed_file(filename):
    if not filename:
        return False
    if "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_EXTENSIONS
##def allowed_file(filename):
##    return (
##        "." in filename
##        and
##        filename.rsplit(".", 1)[1].lower()
##        in ALLOWED_EXTENSIONS
##    )

import uuid

def save_uploaded_image(image_file):
    """
    Save an uploaded image and return its unique filename.
    """
    filename = (
        f"{uuid.uuid4()}_"
        f"{secure_filename(image_file.filename)}"
    )

    image_file.save(
        os.path.join(
            app.config["UPLOAD_FOLDER"],
            filename
        )
    )

    return filename

def clean_text(text):
    """
    Strips leading/trailing spaces,
    reduces multiple spaces to one,
    converts non-strings safely,
    capitalizes the first letter.
    """
    if text is None:
        return ""

    text = str(text).strip()

    if not text:
        return ""

    # Remove repeated whitespace
    text = re.sub(r"\s+", " ", text)

    # Capitalize first letter
    return text[0].upper() + text[1:]


UPLOAD_FOLDER = os.path.join(app.root_path, "static", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER


@app.template_filter('fromjson')
def fromjson_filter(value):
    try:
        return json.loads(value)
    except:
        return []


@app.route('/admin/upload', methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin', 'subadmin')
def upload_questions():

    user = get_current_user()
    errors = {}
    # ================= FILTERS =================
    session_filter = request.args.get("session_filter", "").strip()
    term_filter = request.args.get("term_filter", "").strip()
    type_filter = request.args.get("type_filter", "").strip()

    # ================= SUBJECT QUERY =================
##    subjects_query = Subject.query
    subjects_query = filter_by_institution( Subject.query, Subject )    

    if user.role == "subadmin":
        subjects_query = subadmin_subject_filter(subjects_query, user)

    if session_filter:
        subjects_query = subjects_query.filter(
            or_(
                Subject.label.ilike(f"%{session_filter}%"),
                Subject.name.ilike(f"%{session_filter}%")
            )
        )

    if term_filter:
        subjects_query = subjects_query.filter(
            Subject.label.ilike(f"%{term_filter}%")
        )

    if type_filter:
        subjects_query = subjects_query.filter(
            Subject.result_type == type_filter
        )

    subject_page = request.args.get(
        "subject_page",
        1,
        type=int
    )

    subjects_pagination = (
        subjects_query
        .order_by(
            Subject.result_type.asc(),
            Subject.name.asc(),
            Subject.label.asc()
        )
        .paginate(
            page=subject_page,
            per_page=20
        )
    )

    subjects = subjects_pagination.items

    # ================= POST =================

    if request.method == "POST":

        try:

            subject_id = int(
                request.form["subject_id"]
            )

##            subj_q = Subject.query
            subj_q = filter_by_institution( Subject.query, Subject )            

            if user.role == "subadmin":
                subj_q = subadmin_subject_filter(
                    subj_q,
                    user
                )

            subject = (
                subj_q
                .filter_by(id=subject_id)
                .first_or_404()
            )

            # =====================================
            # READ ALL FORM VALUES
            # =====================================

            bulk_file = request.files.get("file")

            image_file = request.files.get("image")

            prompt = request.form.get(
                "prompt",
                ""
            ).strip()

            choices = request.form.getlist(
                "choices"
            )

            answer = request.form.get(
                "answer",
                ""
            ).strip()

            theory_text = request.form.get(
                "theory_text",
                ""
            ).strip()

            theory_image = request.files.get(
                "theory_image"
            )

            theory_question_id = request.form.get(
                "theory_question_id"
            )

            created = 0

            # =====================================
            # HELPER
            # =====================================

            def add_question( prompt, choices, answer, image=None ):
                nonlocal created

                choices_clean = []

                for c in (choices or []):

                    c = clean_text(c)

                    if c:
                        choices_clean.append(c)

                answer_clean = clean_text(answer)

                q = Question(
                    subject_id=subject.id,
                    prompt=clean_text(prompt),
                    choices_json=json.dumps(
                        choices_clean
                    ),
                    answer=answer_clean,
                    institution_id=current_institution_id(),
                    image=image
                )

                db.session.add(q)

                created += 1

            # =====================================
            # THEORY TEXT UPLOAD
            # =====================================

            if "save_theory_text" in request.form:

                text = theory_text.strip()

                matches = list(
                    re.finditer(
                        r'^\s*\d+(?:\([a-zA-Z]\))?\s*[:.)]?',
                        text,
                        flags=re.MULTILINE
                    )
                )

                uploaded = 0

                for i, match in enumerate(matches):

                    start = match.start()

                    end = (
                        matches[i + 1].start()
                        if i + 1 < len(matches)
                        else len(text)
                    )

                    block = text[start:end].strip()

                    if not block:
                        continue

                    parts = re.split(
                        r'^\s*answer\s*:',
                        block,
                        maxsplit=1,
                        flags=re.IGNORECASE | re.MULTILINE
                    )

                    question_text = parts[0].strip()

##                    answer_text = (
##                        parts[1].strip()
##                        if len(parts) > 1
##                        else ""
##                    )
                    answer_text = clean_text(
                        parts[1] if len(parts) > 1 else ""
                    )                    

                    question_text = re.sub(
                        r'^\s*\d+(?:\([a-zA-Z]\))?\s*[:.)]?\s*',
                        '',
                        question_text
                    )

                    question_text = clean_text(
                        question_text
                    )

                    existing = (
                        filter_by_institution(
                            TheoryQuestion.query,
                            TheoryQuestion
                        )
                        .filter_by(
                            subject_id=subject.id,
                            prompt=question_text
                        )
                        .first()
                    )                    

                    if existing:

                        existing.answer = answer_text

                    else:

                        db.session.add(
                            TheoryQuestion(
                                subject_id=subject.id,
                                prompt=question_text,
                                answer=answer_text,
                                institution_id=current_institution_id()
                            )
                        )

                    uploaded += 1

                db.session.commit()

                flash(
                    f"{uploaded} theory question(s) uploaded successfully.",
                    "success"
                )

                return redirect(
                    url_for(
                        "upload_questions",
                        **request.args
                    )
                )

            # =====================================
            # THEORY IMAGE UPLOAD
            # =====================================

            if "upload_theory_image" in request.form:

                if (
                    not theory_question_id
                ):
                    flash(
                        "Please select a theory question.",
                        "error"
                    )

                    return redirect(
                        url_for(
                            "upload_questions",
                            **request.args
                        )
                    )

                if ( not theory_image or not theory_image.filename ):
                    flash( "Please choose an image.", "error" )
                    return redirect(
                        url_for( "upload_questions", **request.args ) )
                
                if not allowed_file(theory_image.filename):
                    flash( "Unsupported file type.", "error" )
                    return redirect(
                        url_for( "upload_questions", **request.args ) )                

                question = (
                    filter_by_institution(
                        TheoryQuestion.query,
                        TheoryQuestion
                    )
                    .filter_by(
                        id=int(theory_question_id),
                        subject_id=subject.id
                    )
                    .first()
                )

                if not question:

                    flash(
                        "Invalid theory question.",
                        "error"
                    )

                    return redirect(
                        url_for(
                            "upload_questions",
                            **request.args
                        )
                    )

                import uuid

                filename = (
                    f"{uuid.uuid4()}_"
                    f"{secure_filename(theory_image.filename)}"
                )

                theory_image.save(
                    os.path.join(
                        app.config["UPLOAD_FOLDER"],
                        filename
                    )
                )

##                question.image = filename
##
##                db.session.commit()
                # --------------------------------------
                # Delete previous image (if any)
                # --------------------------------------
                if question.image:
                    old_path = os.path.join(
                        app.config["UPLOAD_FOLDER"],
                        question.image
                    )

                    if os.path.isfile(old_path):
                        try:
                            os.remove(old_path)
                        except OSError:
                            pass

                # Save new filename
                question.image = filename

                db.session.commit()
                flash(
                    "Theory image uploaded successfully.",
                    "success"
                )

                return redirect(
                    url_for(
                        "upload_questions",
                        **request.args
                    )
                )

            # =====================================
            # OBJECTIVE
            # =====================================

            prompt = clean_text(request.form.get("prompt", ""))

            choices = [
                clean_text(c)
                for c in request.form.getlist("choices")
            ]

            answer = clean_text(
                request.form.get("answer", "")
            )

            image_filename = None

            if image_file and image_file.filename:

                if not allowed_file(image_file.filename):

                    flash(
                        "Unsupported image type.",
                        "error"
                    )

                    return redirect(
                        url_for(
                            "upload_questions",
                            **request.args
                        )
                    )

                image_filename = save_uploaded_image(image_file)

            # ==================================================
            # SINGLE OBJECTIVE QUESTION
            # ==================================================

            if "submit_single_objective" in request.form:

                if not prompt:

                    flash(
                        "Please type a question before submitting.",
                        "error"
                    )

                    return redirect(
                        url_for(
                            "upload_questions",
                            **request.args
                        )
                    )

                add_question(
                    prompt,
                    choices,
                    answer,
                    image_filename
                )

                db.session.commit()

                flash(
                    "Question uploaded successfully.",
                    "success"
                )

                return redirect(
                    url_for(
                        "upload_questions",
                        **request.args
                    )
                )


            # ==================================================
            # BULK OBJECTIVE UPLOAD
            # ==================================================

            if "submit_bulk_objective" in request.form:

                if not bulk_file or not bulk_file.filename:

                    flash(
                        "Please choose a TXT, CSV or XLSX file.",
                        "error"
                    )

                    return redirect(
                        url_for(
                            "upload_questions",
                            **request.args
                        )
                    )

                filename = bulk_file.filename.lower()
                data = bulk_file.read()

                # TXT / CSV / XLSX parsing starts here

            # ==================================================
            # TXT
            # ==================================================
            print(len(data))
            print(repr(data[:100]))
            if filename.endswith(".txt"):

                text = data.decode(
                    "utf-8",
                    errors="ignore"
                )

                lines = text.splitlines()

                i = 0

                while i < len(lines):

                    line = lines[i].strip()

                    if not line:
                        i += 1
                        continue

                    # ---------- Pipe format ----------

                    if "|" in line:

                        parts = [
                            p.strip()
                            for p in line.split("|")
                        ]

                        if len(parts) == 2:

                            add_question(
                                parts[0],
                                [],
                                parts[1]
                            )

                        else:

                            add_question(
                                parts[0],
                                parts[1:-1],
                                parts[-1]
                            )

                        i += 1
                        continue

                    # ---------- 6-line format ----------

                    if i + 5 < len(lines):

                        block = lines[i:i + 6]

                        if all(block):

                            add_question(
                                block[0],
                                block[1:5],
                                block[5]
                            )

                            i += 6
                            continue

                    i += 1

            # ==================================================
            # CSV
            # ==================================================

            elif filename.endswith(".csv"):

                import csv
                import io

                text = data.decode(
                    "utf-8",
                    errors="ignore"
                )

                reader = csv.DictReader(
                    io.StringIO(text)
                )

                for row in reader:

                    add_question(

                        row.get("question", ""),

                        [
                            row.get("choice1", ""),
                            row.get("choice2", ""),
                            row.get("choice3", ""),
                            row.get("choice4", "")
                        ],

                        row.get("answer", "")
                    )

            # ==================================================
            # EXCEL
            # ==================================================

            elif filename.endswith(".xlsx"):

                from openpyxl import load_workbook
                from io import BytesIO

                wb = load_workbook(
                    BytesIO(data)
                )

                ws = wb.active

                headers = {}

                for col in range(
                    1,
                    ws.max_column + 1
                ):

                    value = ws.cell(
                        1,
                        col
                    ).value

                    if value:

                        headers[
                            str(value).strip().lower()
                        ] = col

                for row in range(
                    2,
                    ws.max_row + 1
                ):

                    question = ws.cell(
                        row,
                        headers.get("question")
                    ).value

                    if not question:
                        continue

                    add_question(

                        str(question),

                        [

                            str(
                                ws.cell(
                                    row,
                                    headers.get("choice1")
                                ).value or ""
                            ),

                            str(
                                ws.cell(
                                    row,
                                    headers.get("choice2")
                                ).value or ""
                            ),

                            str(
                                ws.cell(
                                    row,
                                    headers.get("choice3")
                                ).value or ""
                            ),

                            str(
                                ws.cell(
                                    row,
                                    headers.get("choice4")
                                ).value or ""
                            )

                        ],

                        str(
                            ws.cell(
                                row,
                                headers.get("answer")
                            ).value or ""
                        )

                    )

            else:

                flash(
                    "Unsupported file format.",
                    "error"
                )

                return redirect(
                    url_for(
                        "upload_questions",
                        **request.args
                    )
                )

            db.session.commit()

            flash(
                f"{created} question(s) uploaded.",
                "success"
            )

            return redirect(
                url_for(
                    "upload_questions",
                    **request.args
                )
            )


        except Exception as e:
            db.session.rollback()
            flash(str(e), "error")
            return redirect(url_for('upload_questions', **request.args))

    # ================= PAGINATED QUESTIONS =================
    subject_questions = {}

    for s in subjects:
##        subject_questions[s.id] = Question.query.filter_by(
##            subject_id=s.id ).order_by( Question.id.desc() ).all()
        subject_questions[s.id] = (
            filter_by_institution( Question.query, Question )
            .filter_by( subject_id=s.id )
            .order_by( Question.id.desc() )
            .all() )

    subject_theory_questions = {}

    for s in subjects:
##        subject_theory_questions[s.id] = (
##            TheoryQuestion.query
##            .filter_by(subject_id=s.id) .order_by(TheoryQuestion.id) .all() )        
        subject_theory_questions[s.id] = (
            filter_by_institution( TheoryQuestion.query, TheoryQuestion )
            .filter_by( subject_id=s.id )
            .order_by( TheoryQuestion.id )
            .all() )

    # ================= SUBADMINS =================
##    subadmins = User.query.filter_by(role='subadmin', deleted=False).all()
    subadmins = ( filter_by_institution( User.query, User )
        .filter_by( role="subadmin", deleted=False )
        .all() )

    subject_subadmins = {}
    for sa in subadmins:
        for subj in getattr(sa, "subjects", []):
            subject_subadmins.setdefault(subj.id, []).append(sa)

    recent_subjects = subjects[:10]
    
    theory_questions_json = {}

    for s in subjects:

        theory_questions_json[s.id] = [
            {
                "id": q.id,
                "prompt": q.prompt
            }
            for q in subject_theory_questions[s.id]
        ]


##    for sub in submissions:
##        if sub.answers_json:
##            try:
##                sub.answers = json.loads(sub.answers_json)
##            except Exception:
##                sub.answers = []
##        else:
##            sub.answers = []        

    return render_template(
        "upload.html",
        subjects=subjects,
        user=user,
        subjects_pagination=subjects_pagination,
        subject_questions=subject_questions,
        subject_subadmins=subject_subadmins,
        recent_subjects=recent_subjects,
        session_filter=session_filter,
        term_filter=term_filter,
        type_filter=type_filter,
        subject_theory_questions=subject_theory_questions,
        theory_questions_json=theory_questions_json,
        allowed_extensions=sorted(ALLOWED_EXTENSIONS),

        # Validation errors
        errors=locals().get("errors", {})
    )



@app.route('/admin/subject/<int:subject_id>/edit/questions')
@login_required
@roles_required('global_admin', 'admin','subadmin')
def edit_questions_main(subject_id):
    # Redirect to question 1
    first_q = Question.query.filter_by(subject_id=subject_id).order_by(Question.id).first()
    if not first_q:
        flash("No questions available for this subject.", "error")
        return redirect(url_for('manage_subjects'))
    return redirect(url_for('edit_single_question', subject_id=subject_id, qid=first_q.id))


@app.route('/admin/subject/<int:subject_id>/edit/<int:qid>', methods=['GET','POST'])
@login_required
@roles_required('global_admin', 'admin','subadmin')
def edit_single_question(subject_id, qid):
##    subject = Subject.query.get_or_404(subject_id)
    subject = get_record_in_scope( Subject, subject_id )    
##    q = Question.query.get_or_404(qid)
    q = get_record_in_scope( Question, qid )
    # Load all questions for prev/next navigation
    all_q = Question.query.filter_by(subject_id=subject_id).order_by(Question.id).all()
    ids = [x.id for x in all_q]
    idx = ids.index(qid)

    prev_q = all_q[idx-1] if idx > 0 else None
    next_q = all_q[idx+1] if idx < len(all_q)-1 else None

    if request.method == 'POST':
        action = request.form.get("action")

        if action == "delete":
            db.session.delete(q)
            db.session.commit()
            flash("Question deleted.", "ok")
            return redirect(url_for('edit_questions_main', subject_id=subject_id))

        # Save edits
        q.prompt = request.form.get("prompt").strip()

        choices = [
            request.form.get("optA").strip(),
            request.form.get("optB").strip(),
            request.form.get("optC").strip(),
            request.form.get("optD").strip(),
        ]
        choices = [c for c in choices if c]

        q.choices_json = json.dumps(choices)
        q.answer = request.form.get("answer").strip()

        img = request.files.get("image")
        if img and img.filename:
            filename = secure_filename(img.filename)
            img.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            q.image = filename

        db.session.commit()
        flash("Saved.", "success")

    return render_template('single_question_edit.html',
                           q=q,
                           subject=subject,
                           index=idx+1,
                           total=len(all_q),
                           prev_q=prev_q,
                           next_q=next_q,
                           json_load=json_load,
                           user=get_current_user())

@app.route('/admin/subject/<int:subject_id>/review', methods=['GET','POST'])
@login_required
@roles_required('global_admin', 'admin','subadmin')
def review_questions(subject_id):
##    subject = Subject.query.get_or_404(subject_id)
    subject = get_record_in_scope( Subject, subject_id )    
    questions = ( filter_by_institution( Question.query, Question )
        .filter_by(subject_id=subject_id)
        .order_by(Question.id)
        .all() )

    if request.method == 'POST':
        ids = request.form.getlist("del_ids")
##        for id_ in ids:
##            q = Question.query.get(id_)
##            if q:
##                db.session.delete(q)
        for id_ in ids:
            q = get_record_in_scope( Question, int(id_) )
            if q:
                db.session.delete(q)        
        db.session.commit()
        flash(f"{len(ids)} question(s) deleted.", "ok")
        return redirect(url_for('review_questions', subject_id=subject_id))

    return render_template('review_questions.html',
                           subject=subject,
                           questions=questions,
                           user=get_current_user())

# ---------- Export (Admin/Subadmin with restrictions) ----------
@app.route('/export', methods=['GET'])
@login_required
def export_scores():
    print("STEP 1")
    u = get_current_user()
    print("STEP 2", u.username if u else None, u.role if u else None)
    cleanup_deleted_results()
    print("STEP 3")
    if u.role not in ("global_admin", "admin", "institution_admin", "subadmin"):
        print("STEP 4 - aborting")
        abort(403)
    print("STEP 5")        

    from datetime import datetime, timedelta

    page = request.args.get("page", 1, type=int)
    per_page = 30

    today = datetime.today().date()
    default_start = today - timedelta(days=30)
    default_end = today

    start_date = request.args.get('start_date', default_start.isoformat())
    end_date = request.args.get('end_date', default_end.isoformat())
    quick = request.args.get("quick")
    export_csv = request.args.get("export") == "1"

    # ================= FILTERS (NEW) =================
    username = request.args.get("username", "").strip()
    first_name = request.args.get("first_name", "").strip()
    last_name = request.args.get("last_name", "").strip()
    subject = request.args.get("subject", "").strip()

    query = ( filter_by_institution( Result.query, Result )
        .filter(Result.deleted == False) )
    fname_part = "all_time"

    # ================= QUICK EXPORT FILTERS =================
    if quick == "today":
        start = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)
        query = query.filter(Result.taken_at >= start, Result.taken_at < end)
        fname_part = f"{start.date()}_today"

    elif quick == "30":
        end = datetime.today()
        start = end - timedelta(days=30)
        query = query.filter(Result.taken_at >= start, Result.taken_at < end)
        fname_part = f"{start.date()}_to_{end.date()}"

    elif quick == "all":
        fname_part = "all_time"
        start = None
        end = None

    else:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Result.taken_at >= start)
        except:
            start = default_start

        try:
            end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(Result.taken_at < end)
        except:
            end = default_end

        fname_part = f"{start_date}_to_{end_date}"

    # ================= FILTER LOGIC (NEW) =================
    if username:
        query = query.join(Result.user).filter(User.username.ilike(f"%{username}%"))

    if first_name:
        query = query.join(Result.user).filter(User.first_name.ilike(f"%{first_name}%"))

    if last_name:
        query = query.join(Result.user).filter(User.last_name.ilike(f"%{last_name}%"))

    if subject:
        query = query.join(Subject).filter(Subject.name.ilike(f"%{subject}%"))

    # ================= SUBADMIN RESTRICTIONS =================
    if u.role == 'subadmin':

        allowed_subjects = json_load(u.allowed_subject_ids_json, [])
        allowed_groups = json_load(u.allowed_group_ids_json, [])

        query = query.join(Subject)
        query = query.filter(Subject.hide_scores_for_subadmins == False)

        query = query.join(SubjectGroup, isouter=True).filter(
            (SubjectGroup.hide_scores_for_subadmins == False) |
            (SubjectGroup.id == None)
        )

        if allowed_subjects:
            query = query.filter(Subject.id.in_(allowed_subjects))

        elif allowed_groups:
            query = query.filter(Subject.group_id.in_(allowed_groups))

        else:
            flash("You have no allowed subjects or groups to export.", "error")
            return redirect(url_for("dashboard"))

    # ================= CSV EXPORT =================
    if export_csv:

        results = query.order_by(Result.taken_at.desc()).all()

        import io, csv
        out = io.StringIO()
        writer = csv.writer(out)

        writer.writerow(['Username','First Name','Last Name','Subject',
                         'Label','Score','Total','Date'])

        for r in results:
            writer.writerow([
                r.user.username,
                r.user.first_name or "",
                r.user.last_name or "",
                r.subject.name if r.subject else "N/A",
                r.label or "",
                r.score,
                r.total,
                r.taken_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        out.seek(0)

        username_export = "_".join(
            p.title() for p in [u.first_name, u.last_name] if p
        ) if (u.first_name or u.last_name) else u.username.title()

        filename = f"{username_export}_{fname_part}.csv"

        return Response(
            out,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    # ================= LATEST 10 (NEW SECTION) =================
##    latest_10 = query.order_by(Result.taken_at.desc()).limit(10).all()
    #Ignore filters (global latest submissions only)
##    latest_10 = Result.query.filter(Result.deleted == False)\
##    .order_by(Result.taken_at.desc()).limit(10).all()
    latest_10 = ( filter_by_institution( Result.query, Result )
        .filter(Result.deleted == False)
        .order_by(Result.taken_at.desc())
        .limit(10)
        .all() )    

    # ================= PAGINATION =================
    pagination = query.order_by(Result.taken_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    results = pagination.items

    return render_template(
        'export_scores.html',
        user=u,
        results=results,
        pagination=pagination,
        latest_10=latest_10,   # ✅ ADD THIS
        start=start_date,
        end=end_date
    )

def timefmt(seconds):
    if seconds is None:
        return "0s"
    minutes = seconds // 60
    secs = seconds % 60
    return f"{minutes}m {secs}s"

app.jinja_env.filters['timefmt'] = timefmt



#================= Theory Question Management =================
@app.route('/admin/subject/<int:subject_id>/theory_questions', methods=['GET', 'POST'])
@login_required
def manage_theory_questions(subject_id):

    user = get_current_user()

    if user.role not in ['admin', 'subadmin']:
        abort(403)

##    subject = Subject.query.get_or_404(subject_id)
    subject = get_record_in_scope( Subject, subject_id )
    if user.role == 'subadmin':
        allowed_subjects = json.loads(user.allowed_subject_ids_json)

        if subject.id not in allowed_subjects:
            abort(403)

##    questions = (
##        TheoryQuestion.query
##        .filter_by(subject_id=subject_id)
##        .order_by(TheoryQuestion.id)
##        .all()
##    )
    questions = ( filter_by_institution( TheoryQuestion.query, TheoryQuestion )
        .filter_by( subject_id=subject_id )
        .order_by( TheoryQuestion.id )
        .all() )    

    # ==========================================
    # CREATE NEW THEORY QUESTION
    # ==========================================

    if request.method == 'POST' and 'create_question' in request.form:

##        prompt = request.form.get('prompt', '').strip()
##        answer = request.form.get('answer', '').strip()
        prompt = clean_text(request.form.get('prompt', ''))
        answer = clean_text(request.form.get('answer', ''))        

        image_file = request.files.get("image")
        image_filename = None

        if prompt:

            # Optional image upload
            if image_file and image_file.filename:

                import uuid

                image_filename = (
                    f"{uuid.uuid4()}_"
                    f"{secure_filename(image_file.filename)}"
                )

                image_file.save(
                    os.path.join(
                        app.config["UPLOAD_FOLDER"],
                        image_filename
                    )
                )

            tq = TheoryQuestion(
                subject_id=subject.id,
                prompt=prompt,
                answer=answer,
                image=image_filename,
                institution_id=current_institution_id()
            )            

            db.session.add(tq)
            db.session.commit()

            flash(
                f"Question {len(questions)+1} created successfully!",
                "success"
            )

            return redirect(
                url_for(
                    'manage_theory_questions',
                    subject_id=subject.id
                )
            )

    return render_template(
        'manage_theory_questions.html',
        subject=subject,
        questions=questions,
        user=user
    )

#================= DELETE Single Theory Question =================
@app.route('/admin/theory/delete/<int:q_id>', methods=['POST'])
@login_required
@roles_required('global_admin', 'admin','subadmin')
def delete_theory(q_id):
##    q = TheoryQuestion.query.get_or_404(q_id)
    q = get_record_in_scope( TheoryQuestion, int(q_id) )    
    db.session.delete(q)
    db.session.commit()
    flash("Theory question deleted!", "success")
    return redirect(request.referrer)

#================= BULK DELETE =================
@app.route('/admin/theory/delete_bulk', methods=['POST'])
@login_required
@roles_required('global_admin', 'admin','subadmin')
def delete_theory_bulk():
    ids = request.form.getlist('selected_ids')
    if not ids:
        flash("No questions selected", "error")
        return redirect(request.referrer)
    try:
        ids = [int(q_id) for q_id in ids]
    except ValueError:
        flash("Invalid question IDs", "error")
        return redirect(request.referrer)
##    questions = TheoryQuestion.query.filter(TheoryQuestion.id.in_(ids)).all()
    questions = ( filter_by_institution( TheoryQuestion.query, TheoryQuestion )
        .filter( TheoryQuestion.id.in_(ids) )
        .all() )    
    if not questions:
        flash("No matching questions found", "error")
        return redirect(request.referrer)
    for q in questions:
        db.session.delete(q)
    db.session.commit()
    flash(f"{len(questions)} questions deleted successfully!", "success")
    return redirect(request.referrer)

#================= EDIT =================
@app.route('/admin/theory/edit/<int:q_id>', methods=['GET','POST'])
@login_required
@roles_required('global_admin', 'admin','subadmin')
def edit_theory(q_id):
##    q = TheoryQuestion.query.get_or_404(q_id)
    q = get_record_in_scope( TheoryQuestion, int(q_id) )    
    if request.method == 'POST':
        q.prompt = request.form.get('prompt')
        q.answer = request.form.get('answer')
        db.session.commit()
        flash("Theory question updated!", "success")
        return redirect(url_for('manage_theory_questions', subject_id=q.subject_id))
    return render_template('edit_theory.html', q=q,
                           user = get_current_user())

@app.template_filter('fromjson')
def fromjson_filter(value):
    try:
        return json.loads(value)
    except:
        return []



#================= RECYCLE BIN (FINAL MERGED VERSION) =================
@app.route('/recycle_bin')
@login_required
def recycle_bin():
    u = get_current_user()
    if u.role not in ('global_admin', 'admin', 'subadmin'):
        abort(403)
    per_page = 20
    now = datetime.utcnow()
    # ================= RESULTS =================
    results_page = request.args.get( 'results_page', 1, type=int )
    results_q = ( filter_by_institution( Result.query, Result )
        .filter_by( deleted=True )
        .order_by( Result.deleted_at.desc() ) )    
    total_results = results_q.count()
    results = results_q.offset(
        (results_page - 1) * per_page
    ).limit(
        per_page
    ).all()
    results_total_pages = max( 1, math.ceil(total_results / per_page) )
    for r in results:
        if r.deleted_at:
            delete_due = ( r.deleted_at + timedelta(days=730) )
            r.delete_timestamp = int( delete_due.timestamp() )
        else:
            r.delete_timestamp = None
    # ================= THEORY =================
    theory_page = request.args.get( 'theory_page', 1, type=int )
    theory_q = (
        filter_by_institution( TheorySubmission.query, TheorySubmission )
        .filter( TheorySubmission.deleted == True )
        .order_by( TheorySubmission.deleted_at.desc() ) )    
    total_theory = theory_q.count()
    theory_submissions = theory_q.offset( (theory_page - 1) * per_page
    ).limit( per_page ).all()
    theory_total_pages = max( 1, math.ceil(total_theory / per_page) )
    for sub in theory_submissions:
        if sub.deleted_at:
            delete_due = ( sub.deleted_at + timedelta(days=730) )
            sub.delete_timestamp = int( delete_due.timestamp() )
        else:
            sub.delete_timestamp = None
    # ================= SUBJECTS =================
    subjects_page = request.args.get( 'subjects_page', 1, type=int )
    subjects_q = (
        filter_by_institution( Subject.query, Subject )
        .filter( Subject.deleted == True )
        .order_by( Subject.deleted_at.desc() ) )    
    total_subjects = subjects_q.count()
    subjects = subjects_q.offset( (subjects_page - 1) * per_page ).limit(
        per_page ).all()
    subjects_total_pages = max( 1, math.ceil(total_subjects / per_page) )
    # AUTO DELETE TIMER
    for s in subjects:
        s.delete_timestamp = None
        if s.deleted_at is not None:
            try:
                # 90 DAYS AFTER DELETE
                delete_due = ( s.deleted_at + timedelta(days=90) )
                s.delete_timestamp = int( delete_due.timestamp() )
            except Exception as e:
                print( f"Timestamp error for subject {s.id}:", e )
                s.delete_timestamp = None
    # ================= GROUPS =================
    groups = (
        filter_by_institution( SubjectGroup.query, SubjectGroup )
        .filter( SubjectGroup.deleted == True )
        .order_by( SubjectGroup.name )
        .all() )    
    for g in groups:
        g.delete_timestamp = None
        if g.deleted_at:
            delete_due = ( g.deleted_at + timedelta(days=90) )
            g.delete_timestamp = int( delete_due.timestamp() )
        print( g.id, g.deleted_at, g.delete_timestamp )

    # ================= INSTITUTION ADMINISTRATORS =================
    institution_admins = []
    institution_admins_page = 1
    institution_admins_total_pages = 1

    if u.is_global:        
        institution_admins_page = request.args.get(
            "institution_admins_page", 1, type=int )
        institution_admins_q = (
            User.query
            .filter( User.deleted == True, User.is_institution_admin == True )
            .order_by( User.deleted_at.desc() ) )
        total_institution_admins = institution_admins_q.count()
        institution_admins = ( institution_admins_q
            .offset( (institution_admins_page - 1) * per_page )
            .limit(per_page)
            .all() )
        institution_admins_total_pages = max( 1,
            math.ceil( total_institution_admins / per_page ) )
        for admin in institution_admins:
            admin.delete_timestamp = None
            if admin.deleted_at:
                delete_due = (
                    admin.deleted_at +
                    timedelta(days=90) )
                admin.delete_timestamp = int(
                    delete_due.timestamp() )
    # ================= INSTITUTIONS =================
    institutions = []

    if u.is_global:

        institutions = (
            Institution.query
            .filter_by(deleted=True)
            .order_by(Institution.deleted_at.desc())
            .all()
        )

        for inst in institutions:
            inst.delete_timestamp = None

            if inst.deleted_at:
                delete_due = inst.deleted_at + timedelta(days=90)
                inst.delete_timestamp = int(delete_due.timestamp())                
        
    # ================= USERS =================
    users_page = request.args.get( 'users_page', 1, type=int )
    users_q = ( filter_by_institution( User.query, User )
        .filter( User.deleted == True, User.is_institution_admin == False )
        .order_by( User.deleted_at.desc() ) )   
    total_users = users_q.count()
    users = users_q.offset( (users_page - 1) * per_page
    ).limit( per_page
    ).all()
    users_total_pages = max( 1, math.ceil(total_users / per_page) )
    # AUTO DELETE TIMER
    for usr in users:
        usr.delete_timestamp = None
        if usr.deleted_at:
            delete_due = ( usr.deleted_at + timedelta(days=90) )
            usr.delete_timestamp = int( delete_due.timestamp() )
    return render_template(
        'recycle_bin.html',
        results=results,
        theory_submissions=theory_submissions,
        results_page=results_page,
        results_total_pages=results_total_pages,
        theory_page=theory_page,
        theory_total_pages=theory_total_pages,
        user=u,
        subjects=subjects,
        subjects_page=subjects_page,
        subjects_total_pages=subjects_total_pages,
        groups=groups,
        # ================= USERS =================
        users=users,
        users_page=users_page,
        users_total_pages=users_total_pages,
        # ================= INSTITUTION ADMINS =================
        institution_admins=institution_admins,
        institution_admins_page=institution_admins_page,
        institution_admins_total_pages=institution_admins_total_pages,
        # ================= INSTITUTIONS =================
        institutions=institutions)


# =========================================================
# PERMANENTLY DELETE INSTITUTION
# =========================================================
@app.route(
    "/admin/institutions/permanent-delete/<int:institution_id>",
    methods=["POST"]
)
@login_required
@roles_required("global_admin")
def permanently_delete_institution(institution_id):

    current = get_current_user()

    if current.role != "global_admin" or not current.is_global:
        abort(403)

    institution = Institution.query.filter_by(
        id=institution_id,
        deleted=True
    ).first_or_404()

    # -----------------------------------------
    # Protect Free Institution
    # -----------------------------------------
    if institution.is_system:
        flash(
            "The Free Institution cannot be permanently deleted.",
            "warning"
        )
        return redirect(url_for("recycle_bin"))

    # -----------------------------------------
    # Safety Check
    # -----------------------------------------
    active_users = User.query.filter_by(
        institution_id=institution.id,
        deleted=False
    ).count()

    if active_users:
        flash(
            "Institution still contains active users.",
            "danger"
        )
        return redirect(url_for("recycle_bin"))

    # -----------------------------------------
    # Remove any remaining deleted records
    # -----------------------------------------
    User.query.filter_by(
        institution_id=institution.id,
        deleted=True
    ).delete(synchronize_session=False)

    Subject.query.filter_by(
        institution_id=institution.id,
        deleted=True
    ).delete(synchronize_session=False)

    SubjectGroup.query.filter_by(
        institution_id=institution.id,
        deleted=True
    ).delete(synchronize_session=False)

    Question.query.filter_by(
        institution_id=institution.id,
        deleted=True
    ).delete(synchronize_session=False)

    TheoryQuestion.query.filter_by(
        institution_id=institution.id,
        deleted=True
    ).delete(synchronize_session=False)

    Result.query.filter_by(
        institution_id=institution.id,
        deleted=True
    ).delete(synchronize_session=False)

    TheorySubmission.query.filter_by(
        institution_id=institution.id,
        deleted=True
    ).delete(synchronize_session=False)

    ChatMessage.query.filter_by(
        institution_id=institution.id,
        deleted=True
    ).delete(synchronize_session=False)

    # -----------------------------------------
    # Finally remove the Institution
    # -----------------------------------------
    db.session.delete(institution)
    db.session.commit()

    flash(
        "Institution permanently deleted.",
        "success"
    )
    return redirect(url_for("recycle_bin"))


@app.route('/restore_theory_submissions', methods=['POST'])
@login_required
@roles_required('global_admin', 'admin','subadmin')
def restore_theory_submissions():
    ids = request.form.getlist('submission_ids')
    for sid in ids:
##        sub = TheorySubmission.query.get(sid)
        sub = ( filter_by_institution( TheorySubmission.query,TheorySubmission)
            .filter_by( id=sid )
            .first() )        
        if sub:
            sub.deleted = False
            sub.deleted_at = None
    db.session.commit()
    flash("Theory submissions restored", "success")
    return redirect(url_for('recycle_bin'))


#----Restore ROUTE (WITH CONFLICT LOGIC)for result from recycle bin----
@app.route('/restore_results', methods=['POST'])
@login_required
def restore_results():
    ids = request.form.getlist('result_ids')
    for rid in ids:
##        r = Result.query.get(rid)
        r = ( filter_by_institution( Result.query, Result )
            .filter_by( id=rid )
            .first() )        
        if not r:
            continue
        # check conflict (same user + subject + session)
##        existing = Result.query.filter_by( user_id=r.user_id,
##            subject_id=r.subject_id, session_id=r.session_id, deleted=False )
##        .first()
        existing = ( filter_by_institution( Result.query, Result )
            .filter_by( user_id=r.user_id, subject_id=r.subject_id,
                session_id=r.session_id, deleted=False )
            .first() )        
        if existing:
            # move existing to recycle bin
            existing.deleted = True
            existing.deleted_at = datetime.utcnow()
        # resore old one
        r.deleted = False
        r.deleted_at = None
    db.session.commit()
    flash("Results restored successfully", "success")
    return redirect(url_for('recycle_bin'))

#========RESTORE SUBJECT ROUTE========
@app.route( '/restore_subjects', methods=['POST'] )
@login_required
@roles_required('global_admin', 'admin')
def restore_subjects():
    ids = request.form.getlist( 'subject_ids' )
    if not ids:
        flash("No subjects selected","warning" )
        return redirect( url_for('recycle_bin') )
##    subjects = Subject.query.filter( Subject.id.in_(ids) ).all()
    subjects = ( filter_by_institution( Subject.query, Subject )
        .filter( Subject.id.in_(ids) )
        .all() )    
    for s in subjects:
        s.deleted = False
        s.deleted_at = None
    db.session.commit()
    flash( "Subjects restored successfully", "success" )
    return redirect( url_for('recycle_bin') )


#======== RESTORE USERS ROUTE ========
@app.route('/restore_users', methods=['POST'])
@login_required
@roles_required('global_admin', 'admin')
def restore_users():
    ids = request.form.getlist( 'user_ids' )
    if not ids:
        flash( "No users selected.", "warning" )
        return redirect( url_for('recycle_bin') )
##    users = User.query.filter( User.id.in_(ids) ).all()
    users = ( filter_by_institution( User.query, User )
        .filter( User.id.in_(ids) )
        .all() )    
    restored = 0
    for u in users:
        u.deleted = False
        u.deleted_at = None
        restored += 1
    db.session.commit()
    flash( f"{restored} user(s) restored successfully.", "success" )
    return redirect( url_for('recycle_bin') )

#==========PERMANENT DELETE SUBJECT ROUTE=========
@app.route( '/permanently_delete_subject/<int:subject_id>' )
@login_required
@roles_required('global_admin', 'admin')
def permanently_delete_subject(subject_id):
##    subj = Subject.query.get_or_404( subject_id )
    subj = get_record_in_scope( Subject, subject_id )    
    db.session.delete(subj)
    db.session.commit()
    flash( "Subject permanently deleted", "success" )
    return redirect( url_for('recycle_bin') )

#========RESTORE GROUP ROUTE========
@app.route('/restore_groups', methods=['POST'])
@login_required
@roles_required('global_admin', 'admin')
def restore_groups():
    group_ids = request.form.getlist('group_ids')
    if group_ids:
##        groups = SubjectGroup.query.filter( SubjectGroup.id.in_(group_ids)
##        ).all()
        groups = (
            filter_by_institution( SubjectGroup.query, SubjectGroup )
            .filter( SubjectGroup.id.in_(group_ids) )
            .all() )        
        for g in groups:
            g.deleted = False
            g.deleted_at = None
        db.session.commit()
        flash("Groups restored successfully", "success")
    return redirect(url_for('recycle_bin'))

#==========PERMANENT DELETE GROUP ROUTE=========
@app.route('/permanently_delete_group/<int:group_id>')
@login_required
@roles_required('global_admin', 'admin')
def permanently_delete_group(group_id):
##    g = SubjectGroup.query.get_or_404(group_id)
    g = get_record_in_scope( SubjectGroup, group_id )
    # optional safety check
    if g.subjects:
        flash("Cannot delete group with subjects inside", "error")
        return redirect(url_for('recycle_bin'))
    db.session.delete(g)
    db.session.commit()
    flash("Group permanently deleted", "success")
    return redirect(url_for('recycle_bin'))



# (DASHBOARD/USERS/SUBJECTS/GROUPS 908) (Chat 2175) (License 2355)    
# (Questions 2719) (Exam 3466)(Report Card 4182) (DEBUG 5138)
#=================EXAMS===================
#=================EXAMS===================
#=================EXAMS===================
#=================EXAMS===================
#=================EXAMS===================
#-----Start Exam options ---------
@app.route('/subject/<int:subject_id>/options')
@login_required
def subject_options(subject_id):
    user = get_current_user()
    if user.role != 'student':
        abort(403)
##    subject = Subject.query.get_or_404(subject_id)
    subject = get_record_in_scope( Subject, subject_id )    
    return render_template('subject_options.html', subject=subject, user=user)




# ---------- Start / Resume Exam / Enforce License Check Before Exam----------
@app.route('/start/<int:subject_id>')
@login_required
def start(subject_id):
    user = get_current_user()  
##    subj = Subject.query.get_or_404(subject_id)
    subj = get_record_in_scope( Subject, subject_id )
    # =====================================================
    # RETAKE PERMISSION
    # =====================================================

    existing_result = (
        Result.query
        .filter_by(
            user_id=user.id,
            subject_id=subj.id
        )
        .order_by(Result.taken_at.desc())
        .first()
    )

    if existing_result:

        if existing_result.can_retake:

            # Remove any abandoned session
            ExamSession.query.filter_by(
                user_id=user.id,
                subject_id=subj.id,
                completed=False
            ).delete()

            db.session.commit()

        else:

            flash(
                "You have already taken this exam.",
                "warning"
            )

            return redirect(url_for("dashboard"))
    
    # License check
    ok, msg = check_license(user)
    if not ok:
        flash(msg, "error")
        return redirect(url_for("activate_license"))

    # Check if unfinished session already exists for this subject
    es = (ExamSession.query
          .filter_by(user_id=user.id, subject_id=subj.id, completed=False)
          .first())
    if not es:
        # Build shuffled order of question IDs
        q_ids = [q.id for q in Question.query.filter_by(subject_id=subj.id).order_by(Question.id).all()]
        if not q_ids:
            flash('No questions in this subject yet.', 'error')
            return redirect(url_for('dashboard'))
        random.shuffle(q_ids)

        es = ExamSession(
            user_id=user.id,
            subject_id=subj.id,
            institution_id=current_institution_id(),
            current_index=0,
            progress_count=0,
            question_order_json=json.dumps(q_ids),  # ✅ shuffled order
            answers_json='{}',
            navigation_history_json='[]',
            paused=False,
            completed=False,
            started_at=datetime.utcnow(),
            remaining_seconds=(subj.duration_minutes or 60) * 60  # ✅ fallback to 60 min
        )
        db.session.add(es)
        db.session.commit()

    # Jump to first question
    return redirect(url_for('exam', session_id=es.id))



# ----This route handles pause, resume, or any other session actions -----

# ---------- Unified Session Action ----------
@app.route("/session_action/<int:session_id>", methods=["POST"])
@login_required
def session_action(session_id):
##    es = ExamSession.query.get_or_404(session_id)
    es = get_record_in_scope( ExamSession, session_id )   
    if es.user_id != get_current_user().id:
        abort(403)

    action = request.form.get("action")
    now = datetime.utcnow()

    if action == "pause":
        elapsed = int((now - es.started_at).total_seconds())
        es.remaining_seconds = max(0, es.remaining_seconds - elapsed)
        es.paused = True
        db.session.commit()
        flash("Exam paused.", "ok")

    elif action == "resume":
        es.started_at = now
        es.paused = False
        db.session.commit()
        flash("Exam resumed.", "ok")

    elif action == "submit":
        if not es.completed:
            finalize_exam(es, now)
        return redirect(url_for("result", session_id=es.id))

    return redirect(request.referrer or url_for("dashboard"))

# ---------- Exam Page ----------
@app.route("/exam/<int:session_id>", methods=["GET","POST"])
@login_required
def exam(session_id):
##    es = ExamSession.query.get_or_404(session_id)
    es = get_record_in_scope( ExamSession, session_id )    
    user = get_current_user()

    if es.user_id != user.id:
        abort(403)
    if es.completed:
        return redirect(url_for("dashboard"))

    now = datetime.utcnow()

    if es.paused:
        remaining = es.remaining_seconds or (es.subject.duration_minutes or 60)*60
    else:
        elapsed = int((now - es.started_at).total_seconds())
        remaining = (es.remaining_seconds or (es.subject.duration_minutes or 60)*60) - elapsed

        if remaining <= 0:
            finalize_exam(es, now)
            return redirect(url_for("result", session_id=es.id))

        es.remaining_seconds = remaining
        es.started_at = now
        db.session.commit()

    q_order = json_load(es.question_order_json, [])
    answers = json_load(es.answers_json, {})
    history = json_load(es.navigation_history_json, [])
    total = len(q_order)

    if total == 0:
        flash("No questions available for this exam.", "error")
        return redirect(url_for("dashboard"))

    # ================= POST =================
    if request.method == "POST":
        idx_safe = clamp(es.current_index, 0, total-1)
        q_id = q_order[idx_safe]

        # -------- ALWAYS reload fresh flagged from DB --------
        flagged = json_load(es.flagged_json, [])

        # -------- Save Answer --------
        choice = request.form.get("choice", "")
        answers[str(q_id)] = choice
        es.answers_json = json.dumps(answers)

        # -------- Toggle Flag --------
        if request.form.get("flag"):
            qid_str = str(q_id)

            if qid_str in flagged:
                flagged.remove(qid_str)
            else:
                flagged.append(qid_str)

        # -------- SAVE FLAGGED (ALWAYS) --------
        es.flagged_json = json.dumps(flagged)

        # -------- Actions --------
        action = request.form.get("action")

        if action == "prev":
            if history:
                es.current_index = history.pop()
            es.navigation_history_json = json.dumps(history)
            es.progress_count = max(0, (es.progress_count or 0) - 1)

        elif action == "next":
            history.append(es.current_index)
            es.navigation_history_json = json.dumps(history)

            if es.current_index < total-1:
                es.current_index += 1

            es.progress_count = min(total, (es.progress_count or 0) + 1)

        elif action in ["submit", "pause"]:
            db.session.commit()
            return redirect(url_for("session_action", session_id=es.id), code=307)

        # -------- Jump --------
        goto = request.form.get("goto")
        if goto is not None:
            try:
                goto_index = int(goto)
                if 0 <= goto_index < total:
                    history.append(es.current_index)
                    es.navigation_history_json = json.dumps(history)
                    es.current_index = goto_index
            except:
                pass

        # 🔥 SINGLE COMMIT POINT
        db.session.commit()

        return redirect(url_for("exam", session_id=es.id))

    # ================= GET =================
    flagged = json_load(es.flagged_json, [])

    idx = clamp(es.current_index, 0, total-1)
    q_id = q_order[idx]
##    q = Question.query.get_or_404(q_id)
    q = get_record_in_scope( Question, q_id )
    try:
        choices = json.loads(q.choices_json or "[]")
        if not isinstance(choices, list):
            choices = []
    except:
        choices = []

    saved_answer = answers.get(str(q.id), "")
    is_first = idx == 0
    is_last = idx == total-1

    # =====================================================
    # DOM PRACTICE EXAM ANSWER VISIBILITY
    # =====================================================

    subject_label = (es.subject.label or "").lower()

    allow_show_answers = (
        "dom practice exam" in subject_label
    )

    # Optional:
    # Admin can still force enable manually from subject settings
    if es.subject.show_answers_to_students:
        allow_show_answers = True

    if not es.remaining_seconds:
        es.remaining_seconds = (es.subject.duration_minutes or 60)*60
        db.session.commit()

    return render_template(
        'exam.html',
        es=es,
        q=q,
        idx=idx,
        total=total,
        choices=choices,
        saved_answer=saved_answer,
        is_first=is_first,
        is_last=is_last,
        remaining=remaining,
        user=user,
        q_order=q_order,
        answers=answers,
        flagged=flagged,
        allow_show_answers=allow_show_answers
    )


def finalize_exam(es, now=None):
    """Mark exam as completed, save or update result, and queue result email."""
    if es.completed:
        return None

    now = now or datetime.utcnow()
    score, total = compute_score(es)

    es.completed = True
    es.submitted_at = now

    # Check for existing result
    existing_result = (
        filter_by_institution(Result.query, Result)
        .filter_by(
            user_id=es.user_id,
            subject_id=es.subject_id
        )
        .first()
    )

    if existing_result:

        # If this exam was a granted retake,
        # increase the completed retake count.
        if existing_result.can_retake:
            if existing_result.retake_count is None:
                existing_result.retake_count = 0
            existing_result.retake_count += 1

        existing_result.score = score
        existing_result.total = total
        existing_result.taken_at = now

        # Retake permission has now been used.
        existing_result.can_retake = False

        existing_result.session_id = es.id
        existing_result.result_type = es.subject.result_type
        existing_result.label = es.subject.label

        r = existing_result

    else:

        r = Result(
            user_id=es.user_id,
            subject_id=es.subject_id,
            institution_id=current_institution_id(),
            score=score,
            total=total,
            taken_at=now,
            session_id=es.id,
            label=es.subject.label,
            result_type=es.subject.result_type,
            can_retake=False
        )

        db.session.add(r)

    db.session.commit()

    # =====================================================
    # SEND RESULT EMAIL (IF ENABLED)
    # =====================================================
    u = User.query.get(es.user_id)

    settings = InstitutionSettings.query.filter_by(
        institution_id=current_institution_id()
    ).first()

    auto_forward = (
        settings.auto_forward_scores
        if settings else False
    )

    app_id = get_setting("app_id", "NOT SET")

    if auto_forward and u and u.email:

        try:
            subj = (
                filter_by_institution(Subject.query, Subject)
                .filter_by(id=es.subject_id)
                .first()
            )

            msg = Message(
                subject=f"Exam Result - {subj.name if subj else 'Exam'}",
                recipients=[u.email],
                body=f"""
App ID: {app_id}

Hello {u.username},

You scored {score} / {total}.

Regards,
Exam System
"""
            )

            # Load this institution's SMTP settings.
            if load_admin_mail():

                queue_email(
                    recipient=msg.recipients[0],
                    subject=msg.subject,
                    body=msg.body,
                    institution_id=current_institution_id()
                )

        except Exception as e:
            app.logger.error(f"Auto forward failed: {e}")

    return r


def capitalize_choices_and_answer(choices, answer):
    """
    Normalize spaces and capitalize first letters of all choices and the answer.
    Handles cases where answer is a letter (A/B/C/D) or full text.
    """
    capitalized_choices = [clean_text(c) for c in choices]

    capitalized_answer = answer.strip() if answer else ''
    # If answer is a single letter and choices exist, make it uppercase
    if capitalized_answer and len(capitalized_answer) == 1 and capitalized_choices:
        capitalized_answer = capitalized_answer.upper()
    else:
        capitalized_answer = clean_text(capitalized_answer)

    return capitalized_choices, capitalized_answer


# ---------- Results ----------
def compute_score(es: ExamSession):
    """
    Compute exam score.
    Supports:
    - MCQs with stored full-text answers
    - MCQs where the answer is a letter
    - Open-ended answers
    """
    q_order = json_load(es.question_order_json, [])
    answers = json_load(es.answers_json, {})
    correct = 0
    total = len(q_order)

    for qid in q_order:
        q = Question.query.get(qid)
        if not q:
            total -= 1
            continue

        user_answer = str(answers.get(str(qid), '')).strip()
        if not user_answer:
            continue

        stored_answer = (q.answer or '').strip()
        choices = json_load(q.choices_json, [])

        # --- Case 1: Multiple-choice question ---
        if choices:
            ok = False
            # If stored answer matches one of the choices (full text)
            if stored_answer in choices and user_answer in choices:
                ok = stored_answer.lower() == user_answer.lower()
            else:
                # If stored answer is a letter (A/B/C/D) and choices exist
                letter_map = {chr(65+i): c for i, c in enumerate(choices)}  # A→choices[0], B→choices[1], etc.
                # Compare letter selection
                if user_answer.upper() in letter_map:
                    selected_text = letter_map[user_answer.upper()]
                    ok = selected_text.lower() == stored_answer.lower()
                # Compare stored letter to user's text selection
                elif stored_answer.upper() in letter_map:
                    correct_text = letter_map[stored_answer.upper()]
                    ok = correct_text.lower() == user_answer.lower()
                else:
                    # fallback: exact text match
                    ok = user_answer.lower() == stored_answer.lower()
            if ok:
                correct += 1
        else:
            # --- Case 2: Open-ended ---
            if user_answer.lower() == stored_answer.lower():
                correct += 1

    return correct, total

#===========Retake Exam Permission===========
@app.route('/admin/retake_permissions', methods=['GET'])
@login_required
@roles_required('global_admin', 'admin', 'subadmin')
def retake_permissions():
    user = get_current_user()
    # ==========================================
    # PARAMETERS
    # ==========================================
    search = request.args.get( "search", "" ).strip()
    student_id = request.args.get( "student_id", type=int )
    selected_subject_id = request.args.get( "subject_id", type=int )
    student = None
    selected_subject = None
    subjects = []
    result = None
    theory_submission = None
    # ==========================================
    # LOAD STUDENT LIST
    # ==========================================
##    students_query = User.query.filter_by( role="student", deleted=False )
    students_query = (
        filter_by_institution( User.query, User )
        .filter_by( role="student", deleted=False ) )    
    if search:
        students_query = students_query.filter(
            db.or_(
                User.first_name.ilike( f"%{search}%" ),
                User.last_name.ilike( f"%{search}%" ),
                User.username.ilike( f"%{search}%" ) ) )
    students = (
        students_query
        .order_by( User.first_name, User.last_name )
        .all() )
    # ==========================================
    # LOAD SELECTED STUDENT
    # ==========================================
    if student_id:
##        student = User.query.filter_by(
##            id=student_id, role="student", deleted=False ).first()
        student = (
            filter_by_institution( User.query, User )
            .filter_by( id=student_id, role="student", deleted=False )
            .first() )        
    # ==========================================
    # LOAD SUBJECTS TAKEN
    # ==========================================
    if student:
        subject_map = {}
        # ---------- THEORY ----------
##        theory_submissions = ( TheorySubmission.query
##            .filter_by( user_id=student.id )
##            .all() )
        theory_submissions = (
            filter_by_institution( TheorySubmission.query, TheorySubmission )
            .filter_by( user_id=student.id )
            .all() )        
        for sub in theory_submissions:
            if sub.subject:
                subject_map[ sub.subject.id ] = sub.subject
        # ---------- RESULTS ----------
##        results = ( Result.query
##            .filter_by( user_id=student.id )
##            .all() )
        results = ( filter_by_institution( Result.query, Result )
            .filter_by( user_id=student.id )
            .all() )        
        for row in results:
            if row.subject:
                subject_map[ row.subject.id ] = row.subject
        subjects = sorted(
            subject_map.values(),
            key=lambda s: ( s.name or "" ).lower() )
    # ==========================================
    # LOAD SELECTED SUBJECT
    # ==========================================
    if student and selected_subject_id:
##        selected_subject = Subject.query.get( selected_subject_id )
        selected_subject = ( filter_by_institution( Subject.query, Subject )
            .filter_by( id=selected_subject_id )
            .first() )        
        if selected_subject:
            # ---------- THEORY ----------
##            theory_submission = ( TheorySubmission.query
##                .filter_by( user_id=student.id,
##                subject_id=selected_subject.id ) .first() )
            theory_submission = (
                filter_by_institution( TheorySubmission.query,
                    TheorySubmission )
                .filter_by( user_id=student.id,
                    subject_id=selected_subject.id )
                .first()
            )            
            # ---------- RESULT ----------
            result = (
                Result.query
                .filter_by( user_id=student.id,
                    subject_id=selected_subject.id )
                .order_by( Result.taken_at.desc() )
                .first() )
            result = (
                filter_by_institution( Result.query, Result )
                .filter_by( user_id=student.id,
                    subject_id=selected_subject.id )
                .order_by( Result.taken_at.desc() )
                .first() )            
    # ==========================================
    # RENDER
    # ==========================================
    return render_template(
        "retake_permissions.html",
        user=user,
        search=search,
        students=students,
        student=student,
        subjects=subjects,
        selected_subject=selected_subject,
        result=result,
        theory_submission=theory_submission
    )


#======Toggle Retake Permission=========
@app.route("/admin/toggle_retake_permission", methods=["POST"])
@login_required
@roles_required('global_admin', "admin", "subadmin")
def toggle_retake_permission():
    student_id = request.form.get( "student_id", type=int )
    subject_id = request.form.get( "subject_id", type=int )
    assessment = (
        request.form.get( "assessment" ) or "" ).strip().lower() 
##    student = User.query.get_or_404(student_id)
    student = get_record_in_scope( User, student_id )
##    subject = Subject.query.get_or_404(subject_id)
    subject = get_record_in_scope( Subject, subject_id )    
    admin = get_current_user()
    # ==========================================
    # NON-THEORY ASSESSMENTS ARE DETERMINED
    # FROM THE SUBJECT ITSELF
    # ==========================================
    if assessment != "theory":
        assessment = subject.result_type or ""    
##        admin = get_current_user()
    # ===========================================
    # THEORY
    # ===========================================
    if assessment == "theory":
##        submission = TheorySubmission.query.filter_by(
##            user_id=student.id, subject_id=subject.id ).first()
        submission = (
            filter_by_institution( TheorySubmission.query, TheorySubmission )
            .filter_by( user_id=student.id, subject_id=subject.id )
            .first() )        
        if not submission:
            flash( "Theory submission not found.", "warning" )
        else:

            if submission.can_retake:
                submission.can_retake = False
            else:
                submission.can_retake = True
                submission.retake_granted_by = admin.id
                submission.retake_granted_at = datetime.utcnow()            
            # ===================================
            # AUDIT INFORMATION
            # ===================================
            # NOTE:
            # Do NOT clear the audit fields when disabling.
            # We want to preserve the last grant information.
            db.session.commit()
            flash(
                "Theory retake {}.".format(
                    "enabled"
                    if submission.can_retake
                    else "disabled"
                ), "success" )
    # ===========================================
    # OBJECTIVE / CA1 / CA2 / CA3
    # ===========================================
    else:
##        result = ( Result.query
##            .filter_by( user_id=student.id, subject_id=subject.id )
##            .order_by( Result.taken_at.desc() ) .first() )
        result = ( filter_by_institution( Result.query, Result )
            .filter_by( user_id=student.id, subject_id=subject.id )
            .order_by( Result.taken_at.desc() )
            .first() )        
        if not result:
            flash( f"{assessment.upper()} result not found.", "warning" )
        else:
            if result.can_retake:
                result.can_retake = False
            else:
                result.can_retake = True
                result.retake_granted_by = admin.id
                result.retake_granted_at = datetime.utcnow()            
            # ===================================
            # AUDIT INFORMATION
            # ===================================
            # NOTE:
            # Do NOT clear the audit fields when disabling.
            # Preserve the last granted information.
            db.session.commit()
            # ===================================
            # FRIENDLY RESULT TYPE NAME
            # ===================================
            if subject.result_type == "exam":
                result_name = "Objective"
            elif subject.result_type == "ca":
                label_upper = (subject.label or "").upper()
                if "1ST" in label_upper:
                    result_name = "CA 1"
                elif "2ND" in label_upper:
                    result_name = "CA 2"
                elif "3RD" in label_upper:
                    result_name = "CA 3"
                else:
                    result_name = "Continuous Assessment"
            elif subject.result_type == "practice":
                result_name = "Practice"
            else:
                result_name = (
                    subject.result_type or "Assessment"
                ).title()
            flash(
                f"{result_name} retake "
                + (
                    "enabled."
                    if result.can_retake
                    else "disabled."
                ),
                "success"
            )
    search = request.args.get(
        "search",
        request.form.get(
            "search",
            ""
        )
    )
    return redirect(
        url_for(
            "retake_permissions",
            student_id=student.id,
            subject_id=subject.id,
            search=search
        )
    )   

@app.route('/result/<int:session_id>')
@login_required
def result(session_id):
##    es = ExamSession.query.get_or_404(session_id)
    es = get_record_in_scope( ExamSession, session_id )    
    u = get_current_user()
    if es.user_id != u.id and u.role != 'admin':
        # subadmin can view depending on hide flag
        if u.role == 'subadmin':
            subj = es.subject
            group_hide = subj.group.hide_scores_for_subadmins if subj.group else False
            if subj.hide_scores_for_subadmins or group_hide:
                abort(403)
        else:
            abort(403)
# --- fetch result directly by session_id ---
##    res = Result.query.filter_by(session_id=es.id).first()
    res = ( filter_by_institution( Result.query, Result )
        .filter_by( session_id=es.id )
        .first() )    

    if not res:
        # fallback (e.g., if something failed to save before)
        flash("No stored result found, recomputing...")
        score, total = compute_score(es)
        res = Result(user_id=es.user_id, subject_id=es.subject_id,
                     institution_id=current_institution_id(), score=score,
                     total=total, taken_at=es.submitted_at, session_id=es.id)
        db.session.add(res)
        db.session.commit()

    return render_template('result.html',
                           score=res.score,
                           total=res.total, taken_at=res.taken_at,
                           es=es, user=u)

# ----------Hide/Unhide results on students dashboards by students ----------
@app.route("/toggle_result/<int:result_id>")
@login_required
def toggle_result(result_id):
##    r = Result.query.get_or_404(result_id)
    r = get_record_in_scope( Result, result_id )    
    if r.user_id != get_current_user().id and get_current_user().role != "admin":
        abort(403)
    r.hidden = not r.hidden
    db.session.commit()
    return redirect(url_for("dashboard"))

# --------Delete result ---------
@app.route("/delete_result/<int:result_id>", methods=["POST"])
@login_required
@roles_required('global_admin', "admin")
def delete_result(result_id):
##    r = Result.query.get_or_404(result_id)
    r = get_record_in_scope( Result, result_id )    
    db.session.delete(r)
    db.session.commit()
    flash("Result deleted.", "ok")
    return redirect(url_for("dashboard"))

#----Delete Results Route----
@app.route('/delete_results', methods=['POST'])
@login_required
def delete_results():
    u = get_current_user()
    if u.role not in ('admin', 'subadmin'):
        abort(403)

    ids = request.form.getlist('result_ids')

##    for rid in ids:
##        r = Result.query.get(rid)
##        if r:
##            r.deleted = True
##            r.deleted_at = datetime.utcnow()
    for rid in ids:
        r = get_record_in_scope( Result, int(rid) )
        if r:
            r.deleted = True
            r.deleted_at = datetime.utcnow()    

    db.session.commit()
    flash("Selected results moved to recycle bin", "success")
    return redirect(url_for('export_scores'))



#-----Theory exam submit route-----
@app.route('/subject/<int:subject_id>/theory/submit', methods=['POST'])
def submit_theory(subject_id):
    answers = {}
    for key, value in request.form.items():
        answers[key] = value
    flash("Your theory answers have been submitted successfully!", "success")
    return redirect(url_for('dashboard'))

#----Student Theory Exam Page----
@app.route('/subject/<int:subject_id>/theory', methods=['GET', 'POST'])
@login_required
def start_theory(subject_id):
    user = get_current_user()
##    subject = Subject.query.get_or_404(subject_id)
    subject = get_record_in_scope( Subject, subject_id )    
# ==========================================
# LICENSE CHECK
# ==========================================
    ok, msg = check_license(user)
    if not ok:
        flash(msg, "error")
        return redirect(url_for("activate_license"))    
##    questions = (
##        TheoryQuestion.query .filter_by(subject_id=subject_id)
##        .order_by(TheoryQuestion.id) .all() )
    questions = (
        filter_by_institution( TheoryQuestion.query, TheoryQuestion )
        .filter_by( subject_id=subject_id )
        .order_by( TheoryQuestion.id )
        .all() )    
    # =====================================================
    # GET EXISTING SUBMISSION
    # =====================================================
##    submission = TheorySubmission.query.filter_by(
##        user_id=user.id, subject_id=subject_id ).first()
    submission = (
        filter_by_institution( TheorySubmission.query, TheorySubmission )
        .filter_by( user_id=user.id, subject_id=subject_id )
        .first() )    
    # =====================================================
    # CREATE FIRST SUBMISSION
    # =====================================================

    if not submission:

        submission = TheorySubmission(
            user_id=user.id,
            subject_id=subject_id,
            institution_id=current_institution_id(),
            answers_json=json.dumps({}),
            submitted_at=None,
            started_at=datetime.utcnow(),
            remaining_seconds=(
                (subject.theory_duration_minutes or 30) * 60
            )
        )

        db.session.add(submission)
        db.session.commit()

    # =====================================================
    # RETAKE RESET
    # =====================================================
    elif submission.submitted_at and submission.can_retake:
        # Count this granted retake as being used.
        submission.retake_count += 1

        submission.answers_json = json.dumps({})

        submission.submitted_at = None

        submission.started_at = datetime.utcnow()

        submission.remaining_seconds = (
            (subject.theory_duration_minutes or 30) * 60
        )

        submission.completed = False

        submission.score = None

        submission.marked_by = None

        # Consume the retake permission.
        submission.can_retake = False

        db.session.commit()

    # =====================================================
    # ALREADY SUBMITTED
    # =====================================================

    elif submission.submitted_at:

        flash(
            "You have already submitted this theory exam.",
            "warning"
        )

        return redirect(url_for("dashboard"))

    print(
        "submitted_at =",
        submission.submitted_at if submission else None
    )

    # =====================================================
    # TIMER
    # =====================================================

    now = datetime.utcnow()

    if submission.started_at:

        elapsed = int(
            (now - submission.started_at).total_seconds()
        )

        remaining = (
            submission.remaining_seconds - elapsed
        )

        print("started_at =", submission.started_at)
        print("remaining_seconds =", submission.remaining_seconds)
        print("elapsed =", elapsed)
        print("remaining =", remaining)

        if remaining <= 0:

            if not submission.submitted_at:

                submission.completed = True

                submission.submitted_at = now

                db.session.commit()

            return redirect(
                url_for(
                    "dashboard",
                    theory_timeout=1
                )
            )

    else:

        remaining = (
            subject.theory_duration_minutes or 30
        ) * 60

    # =====================================================
    # LOAD ANSWERS
    # =====================================================

    try:

        answers = (
            json.loads(submission.answers_json)
            if submission.answers_json
            else {}
        )

    except:

        answers = {}

    # =====================================================
    # SAVE ANSWERS
    # =====================================================

    if request.method == "POST":

        print("POST RECEIVED")
        print("FORM DATA:", request.form)

        for q in questions:

            q_key = str(q.id)

            answers[q_key] = answers.get(q_key, {})

            answers[q_key]["text"] = request.form.get(
                f"answer_{q.id}"
            )

            file = request.files.get(
                f"file_{q.id}"
            )

            if file and file.filename:

                filename = secure_filename(file.filename)

                upload_dir = os.path.join(
                    app.root_path,
                    "static",
                    "uploads",
                    "theory"
                )

                os.makedirs(
                    upload_dir,
                    exist_ok=True
                )

                saved_filename = (
                    f"{user.id}_{subject.id}_{q.id}_{filename}"
                )

                file.save(
                    os.path.join(
                        upload_dir,
                        saved_filename
                    )
                )

                answers[q_key]["file_path"] = (
                    f"theory/{saved_filename}"
                )

        submission.answers_json = json.dumps(
            answers
        )

        # ==========================
        # FINAL SUBMIT
        # ==========================

        if request.form.get("submit") == "1":

            submission.submitted_at = datetime.utcnow()

            submission.completed = True

            submission.started_at = None

            submission.remaining_seconds = 0

            submission.can_retake = False

        db.session.commit()

        print(
            "SUBMIT FIELD VALUE =",
            request.form.get("submit")
        )

        if request.form.get("submit") == "1":

            flash(
                "Your theory exam has been submitted!",
                "success"
            )

            return redirect(
                url_for("dashboard")
            )

        flash(
            "Answers saved!",
            "info"
        )

        return redirect(
            url_for(
                "start_theory",
                subject_id=subject.id
            )
        )

    print(
        "REMAINING SENT TO TEMPLATE =",
        remaining
    )

    return render_template(
        "theory_exam.html",
        subject=subject,
        questions=questions,
        submission=submission,
        answers=answers,
        user=user,
        remaining=remaining
    )

#===============timeout route============
@app.route('/subject/<int:subject_id>/theory/timeout')
@login_required
def theory_timeout(subject_id):

    user = get_current_user()

##    submission = TheorySubmission.query.filter_by(
##        user_id=user.id, subject_id=subject_id ).first_or_404()
    submission = (
        filter_by_institution( TheorySubmission.query, TheorySubmission )
        .filter_by( user_id=user.id, subject_id=subject_id )
        .first_or_404() )    

    if not submission.submitted_at:
        submission.completed = True
        submission.submitted_at = datetime.utcnow()

        db.session.commit()

    flash(
        "⏰ Exam ended. Time limit reached.",
        "warning"
    )

    return redirect(url_for("dashboard"))

#=============Theory Auto-Save Route==============
@app.route( '/subject/<int:subject_id>/theory/autosave', methods=['POST'] )
@login_required
def theory_autosave(subject_id):

    user = get_current_user()

##    submission = TheorySubmission.query.filter_by(
##        user_id=user.id, subject_id=subject_id ).first_or_404()
    submission = (
        filter_by_institution( TheorySubmission.query, TheorySubmission )
        .filter_by( user_id=user.id, subject_id=subject_id )
        .first_or_404() )    

    try:
        answers = json.loads(
            submission.answers_json or "{}"
        )
    except:
        answers = {}

    for key, value in request.form.items():

        if key.startswith("answer_"):

            q_id = key.replace(
                "answer_",
                ""
            )

            answers[q_id] = answers.get(
                q_id,
                {}
            )

            answers[q_id]["text"] = value

    submission.answers_json = json.dumps(
        answers
    )

    db.session.commit()

    return {"success": True}

#-----Admin Scoring-----
@app.route('/admin/subject/<int:subject_id>/theory_submissions')
@login_required
def view_theory_submissions(subject_id):
    user = get_current_user()
    if user.role not in ['admin', 'subadmin']:
        abort(403)
##    submissions = TheorySubmission.query.filter_by(subject_id=subject_id)
##    .all()
    submissions = (
        filter_by_institution( TheorySubmission.query, TheorySubmission )
        .filter_by( subject_id=subject_id )
        .all() )    
    for sub in submissions:
        try:
            sub.answers = json.loads(sub.answers_json) if sub.answers_json else {}
        except Exception:
            sub.answers = {}    
    return render_template('view_theory_submissions.html',
                           submissions=submissions,
                           user=user)


# ================= ADMIN THEORY SUBMISSIONS =================
@app.route('/admin/theory_submissions', methods=['GET','POST'])
@login_required
@roles_required('global_admin', 'admin','subadmin')
def theory_submissions():
    user = get_current_user()
##    cleanup_deleted_submissions()

    # --- Subjects for filter ---
    subjects_query = filter_by_institution( Subject.query, Subject )
    if user.role == 'subadmin':
        allowed_ids = json.loads(user.allowed_subject_ids_json or "[]")
        subjects_query = subjects_query.filter(Subject.id.in_(allowed_ids))
    subjects = subjects_query.all()

    # --- FILTERS ---
    subject_id = request.args.get('subject_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    first_name = request.args.get('first_name', '').strip()
    last_name = request.args.get('last_name', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    submissions_q = ( filter_by_institution( TheorySubmission.query,
            TheorySubmission )
        .join( User, TheorySubmission.user)
        .filter(TheorySubmission.deleted == False))

    # --- SUBADMIN PROTECTION ---
    if user.role == 'subadmin' and allowed_ids:
        submissions_q = submissions_q.filter(TheorySubmission.subject_id.in_(allowed_ids))

    # --- APPLY FILTERS ---
    if subject_id:
        submissions_q = submissions_q.filter(TheorySubmission.subject_id == subject_id)

    if start_date:
        submissions_q = submissions_q.filter(
            TheorySubmission.submitted_at >= datetime.strptime(start_date, "%Y-%m-%d")
        )

    if end_date:
        submissions_q = submissions_q.filter(
            TheorySubmission.submitted_at <= datetime.strptime(end_date, "%Y-%m-%d")
        )

    if first_name:
        submissions_q = submissions_q.filter(User.first_name.ilike(f"%{first_name}%"))

    if last_name:
        submissions_q = submissions_q.filter(User.last_name.ilike(f"%{last_name}%"))

    submissions_q = submissions_q.order_by(desc(TheorySubmission.submitted_at))

    total_submissions = submissions_q.count()
    submissions = submissions_q.offset((page - 1) * per_page).limit(per_page).all()
    total_pages = math.ceil(total_submissions / per_page)


    # ================= LATEST 10 THEORY (NEW) =================
    latest_10_theory = submissions_q.order_by(
        desc(TheorySubmission.submitted_at)
    ).limit(10).all()
    #Ignore filters (Global latest 10 submissions) with subadmin restrictions
##    latest_10_theory = TheorySubmission.query\
##        .filter(TheorySubmission.deleted == False)
    latest_10_theory = (
        filter_by_institution( TheorySubmission.query, TheorySubmission )
        .filter( TheorySubmission.deleted == False ) )    

    if user.role == 'subadmin' and allowed_ids:
        latest_10_theory = latest_10_theory.filter(
            TheorySubmission.subject_id.in_(allowed_ids)
        )

    latest_10_theory = latest_10_theory.order_by(
        desc(TheorySubmission.submitted_at)
    ).limit(10).all()


    # --- ZIP DOWNLOAD ---
    if request.method == 'POST' and 'download_zip' in request.form:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zipf:
            for sub in submissions_q.all():
                answers = json.loads(sub.answers_json or "{}")
                for qid, ans in answers.items():
                    if ans.get('file_path'):
                        filepath = os.path.join(app.root_path, 'static', 'uploads', ans['file_path'])
                        if os.path.exists(filepath):
                            filename = f"{sub.user.first_name}_{sub.user.last_name}_{os.path.basename(filepath)}"
                            zipf.write(filepath, arcname=filename)
        zip_buffer.seek(0)
        return send_file(zip_buffer, mimetype='application/zip', download_name='submissions.zip', as_attachment=True)

    return render_template(
        'theory_submissions.html',
        subjects=subjects,
        submissions=submissions,
        selected_subject_id=subject_id,
        start_date=start_date,
        end_date=end_date,
        first_name=first_name,
        latest_10_theory=latest_10_theory,  # ✅ ADD THIS
        last_name=last_name,
        page=page,
        total_pages=total_pages,
        user=user
    )


@app.route('/delete_theory_submissions', methods=['POST'])
@login_required
@roles_required('global_admin', 'admin','subadmin')
def delete_theory_submissions():
    
    ids = request.form.getlist('submission_ids')

    for sid in ids:
##        sub = TheorySubmission.query.get(sid)
        sub = (
            filter_by_institution( TheorySubmission.query, TheorySubmission )
            .filter_by( id=sid )
            .first() )        
        if sub:
            sub.deleted = True
            sub.deleted_at = datetime.utcnow()

    db.session.commit()
    flash("Selected submissions moved to recycle bin", "success")
    return redirect(url_for('theory_submissions'))


def cleanup_deleted_submissions():
    cutoff = datetime.utcnow() - timedelta(days=RECYCLE_RETENTION_DAYS)

##    TheorySubmission.query.filter( TheorySubmission.deleted == True,
##        TheorySubmission.deleted_at <= cutoff
##    ).delete(synchronize_session=False)
    filter_by_institution( TheorySubmission.query, TheorySubmission
    ).filter(
        TheorySubmission.deleted == True,
        TheorySubmission.deleted_at <= cutoff
    ).delete( synchronize_session=False )    

    db.session.commit()


# ================= INLINE SCORING (UPDATED) =================
@app.route('/admin/theory_submission/<int:submission_id>/score',
           methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin', 'subadmin')
def score_theory(submission_id):

    u = get_current_user()

##    sub = TheorySubmission.query.get_or_404(submission_id)
    sub = get_record_in_scope( TheorySubmission, submission_id )
    # ---------- Display scoring page ----------
    if request.method == "GET":

        try:
            sub.answers = json.loads(sub.answers_json) if sub.answers_json else {}
        except Exception:
            sub.answers = {}

        return render_template(
            "score_theory.html",
            submission=sub
        )

    # ---------- Save score ----------
    score = int(request.form.get("score"))

    sub.score = score
    sub.marked_by = get_current_user().id

    result = ( filter_by_institution( Result.query, Result )
        .filter_by( user_id=sub.user_id, subject_id=sub.subject_id )
        .first() )

    if not result:
        result = Result( user_id=sub.user_id,
            institution_id=current_institution_id(),
            subject_id=sub.subject_id, score=0, total=0)
        db.session.add(result)

    result.theory_score = score

    db.session.commit()

    flash("Score saved!", "success")

    settings = InstitutionSettings.query.filter_by(
        institution_id=current_institution_id()
    ).first()

    auto_forward = (
        settings.auto_forward_scores
        if settings else False
    )
    app_id = get_setting("app_id", "NOT SET")

    if auto_forward and sub.user.email:
        try:
            msg = Message(
                subject="Theory Exam Result",
                recipients=[sub.user.email],
                body=f"""
App ID: {app_id}

Your theory score is: {score}

Regards
"""
            )

            load_admin_mail()
            queue_email(
                recipient=msg.recipients[0],
                subject=msg.subject,
                body=msg.body,
                institution_id=current_institution_id()
            )

        except Exception as e:
            app.logger.error(f"Theory auto mail failed: {e}")

    return redirect(
        url_for(
            "view_theory_submissions",
            subject_id=sub.subject_id,
            user=u
        )
    )


# (DASHBOARD/USERS/SUBJECTS/GROUPS 908) (Chat 2175) (License 2355)    
# (Questions 2719) (Exam 3466)(Report Card 4182) (DEBUG 5138)
# ================= Admin: Report Card Generation =================
# ================= Admin: Report Card Generation =================
# ================= Admin: Report Card Generation =================
# ================= Admin: Report Card Generation =================
# ================= Admin: Report Card Generation =================

# ================= CONFIG =================
UPLOAD_FOLDER = os.path.join(app.root_path, "static", "uploads/report_card")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["REPORT_CARD_UPLOAD"] = UPLOAD_FOLDER

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ================= HELPERS =================
def calculate_grade(p):
    if p >= 70: return "A", "Excellent"
    elif p >= 60: return "B", "Very Good"
    elif p >= 50: return "C", "Good"
    elif p >= 45: return "D", "Fair"
    elif p >= 40: return "E", "Pass"
    else: return "F", "Fail"


# ================= EDIT RESULT =================
@app.route('/admin/edit-result/<int:result_id>', methods=['GET','POST'])
@login_required
@roles_required('global_admin', 'admin','subadmin')
def edit_result(result_id):
    u = get_current_user()
##    result = Result.query.get_or_404(result_id)
    result = get_record_in_scope( Result, result_id )    

    if request.method == 'POST':
        result.theory_score = int(request.form.get('theory_score', 0))
        result.score = int(request.form['score'])
        result.total = int(request.form['total'])
        result.remark = request.form['remark']
        result.teacher_name = request.form['teacher_name']

        total_score = result.score + result.theory_score
        percentage = (total_score / result.total) * 100 if result.total else 0
        result.percentage = percentage

        grade, remark = calculate_grade(percentage)
        result.grade = grade
        result.remark = remark

        db.session.commit()
        flash("Result updated")

    return render_template("edit_result.html", result=result, user=u)


# ================= SCHOOL SETTINGS =================
@app.route('/admin/school-settings', methods=['GET','POST'])
@login_required
@roles_required('global_admin', 'admin')
def school_settings():
    u = get_current_user()
##    settings = SchoolSettings.query.first()
    settings = ( filter_by_institution( SchoolSettings.query, SchoolSettings )
        .first() )
    if settings is None:
        flash("School settings have not been configured for this institution.", "warning")
        return redirect(url_for("admin_settings"))    
    if request.method == 'POST':
        if not settings:
            settings = SchoolSettings(
                institution_id=current_institution_id())
        # =====================================
        # BASIC INFO
        # =====================================
        settings.name = request.form['name']
        settings.address = request.form['address']
        settings.phone = request.form['phone']
        settings.term = request.form['term']
        settings.session = request.form['session']
        settings.next_term_begins = request.form[ 'next_term_begins' ]
        
        # =====================================
        # REPORT CONFIGURATION
        # =====================================
        settings.academic_mode = request.form.get( "academic_mode",
            "secondary" )
        settings.grading_scale = request.form.get( "grading_scale",
            "percentage" )
        settings.default_report_design = request.form.get(
            "default_report_design", "standard.html" )        
        # =====================================
        # SCORE ALLOCATION
        # =====================================
        settings.ca1_percentage = float( request.form.get("ca1_percentage", 0) )
        settings.ca2_percentage = float( request.form.get("ca2_percentage", 0))
        settings.ca3_percentage = float( request.form.get("ca3_percentage", 0))
        settings.objective_percentage = float(
            request.form.get("objective_percentage", 0) )
        settings.theory_percentage = float(
            request.form.get("theory_percentage", 0) )
        total_percentage = (
            settings.ca1_percentage
            + settings.ca2_percentage
            + settings.ca3_percentage
            + settings.objective_percentage
            + settings.theory_percentage
        )
        if total_percentage != 100:
            flash( "Score allocation must total exactly 100%.", "danger" )
            return render_template(
                "school_settings.html",
                settings=settings,
                user=u
            )
        # =====================================
        # REPORT DISPLAY OPTIONS
        # =====================================
        settings.show_student_position = bool(
            request.form.get("show_student_position") )
        settings.show_class_average = bool(
            request.form.get("show_class_average") )
        settings.show_class_size = bool(
            request.form.get("show_class_size") )
        settings.show_gpa = bool( request.form.get("show_gpa")  )
        settings.show_cgpa = bool( request.form.get("show_cgpa") )
        settings.show_attendance = bool(
            request.form.get("show_attendance") )
        settings.show_psychomotor = bool(
            request.form.get("show_psychomotor") )
        settings.show_teacher_comment = bool(
            request.form.get("show_teacher_comment") )
        settings.show_principal_comment = bool(
            request.form.get("show_principal_comment") )
        settings.show_next_term = bool( request.form.get("show_next_term") )       
        # =====================================
        # FILE UPLOADS
        # =====================================
        for field in [
            'logo',
            'background',
            'principal_signature'
        ]:
            file = request.files.get(field)
            if file and allowed_file(file.filename):
                filename = secure_filename(
                    file.filename
                )
                path = os.path.join(
                    UPLOAD_FOLDER,
                    filename
                )
                file.save(path)
                setattr(
                    settings,
                    f"{field}_path",
                    "uploads/report_card/" + filename
                )
        db.session.add(settings)
        db.session.commit()
        flash("School settings saved")
    return render_template(
        'school_settings.html',
        settings=settings,
        user=u
    )


#====Allocate Percentages====
def scale_score(raw_score, raw_maximum, allocated_percentage):
    """
    Converts a raw score into its allocated contribution
    on the report card.
    """
    raw_score = raw_score or 0
    if raw_maximum <= 0:
        return 0
    return round( (raw_score / raw_maximum) * allocated_percentage, 2 )

#====Publish Route====
@app.route("/report-card/publish")
@login_required
@roles_required('global_admin', "admin")
def publish_report_cards():
    settings = (
        SchoolSettings.query
        .filter_by( institution_id=current_institution_id() )
        .first_or_404() )
    unfinished = (
        filter_by_institution( ReportCardRemark.query, ReportCardRemark )
        .filter(
            ReportCardRemark.term == settings.term,
            ReportCardRemark.session == settings.session,
            ReportCardRemark.report_batch ==
                f"{settings.session}-{settings.term}",
            (
                ReportCardRemark.teacher_locked == False
            ) |
            (
                ReportCardRemark.principal_locked == False
            ) )
        .count() )
    if unfinished:
        flash( f"{unfinished} report cards are still incomplete.", "danger" )
        return redirect(
            url_for("report_card_dashboard") )
    (
    filter_by_institution( ReportCardRemark.query, ReportCardRemark
    )
    .filter_by(
        term=settings.term,
        session=settings.session,
        report_batch=f"{settings.session}-{settings.term}"
    )
    .update(
        {
            ReportCardRemark.teacher_locked: True,
            ReportCardRemark.principal_locked: True
        },
        synchronize_session=False ) )    
    settings.report_cards_published = True
    settings.report_cards_published_at = datetime.utcnow()
    settings.report_cards_published_by = current_user.id
    settings.last_published_batch = ( f"{settings.session}-{settings.term}" )    
    db.session.commit()
    flash( "Report cards published successfully.", "success" )
    return redirect( url_for("report_card_dashboard") )    

#====UnPublish Route====
@app.route("/report-card/unpublish")
@login_required
@roles_required('global_admin', "admin")
def unpublish_report_cards():
    settings = (
        SchoolSettings.query
        .filter_by( institution_id=current_institution_id() )
        .first_or_404() )
    settings.report_cards_published = False
    settings.report_cards_published_at = None
    settings.report_cards_published_by = None
    db.session.commit()
    flash( "Report cards have been unpublished.",  "warning" )

    return redirect( url_for("report_card_dashboard") )   

#====student report card viewing route====
@app.route("/student-report/<token>")
def student_report(token):
    # Validate token
    link = get_valid_report_card_link(token)
    if not link:
        abort(404)
    session["report_card_institution"] = link.institution_id
    settings = (
        SchoolSettings.query
        .filter_by(
            institution_id=current_report_card_institution()
        )
        .first()
    )
    if not settings:
        abort(400)
    if not settings.report_cards_published:
        flash(
            "Report cards have not yet been published.",
             "warning"
        )
        return redirect(url_for("dashboard"))

    # continue generating report...


#====Institution Admin Report Card Dashboard====
@app.route("/report-card/dashboard")
@login_required
@roles_required('global_admin', "admin", "subadmin")
def report_card_dashboard():
    settings = (
        SchoolSettings.query
        .filter_by( institution_id=current_institution_id() )
        .first_or_404() )

    batch = f"{settings.session}-{settings.term}"
    remarks = (
        filter_by_institution( ReportCardRemark.query, ReportCardRemark )
        .filter_by( term=settings.term, session=settings.session,
            report_batch=batch )
        .all() )
    teacher_done = sum( 1 for r in remarks if r.teacher_completed_at )
    principal_done = sum( 1 for r in remarks if r.principal_completed_at )
    fully_completed = sum( 1 for r in remarks if r.teacher_completed_at
        and r.principal_completed_at )
    teacher_locked = sum( 1 for r in remarks if r.teacher_locked )
    principal_locked = sum( 1 for r in remarks if r.principal_locked )
    total_reports = len(remarks)
    return render_template(
        "report_card_dashboard.html",
        settings=settings,
        remarks=remarks,
        teacher_done=teacher_done,
        principal_done=principal_done,
        fully_completed=fully_completed,
        teacher_locked=teacher_locked,
        principal_locked=principal_locked,
        total_reports=total_reports
    )    


#=====Teacher Route=====
@app.route( "/report-card/teacher/<token>", methods=["GET","POST"] )
def teacher_report_card(token):
    link = get_valid_report_card_link(token)
    if not link:
        abort(404)
    if not link.role_type == "teacher":
        abort(403)
    session["report_card_institution"] = link.institution_id
    session["report_card_link_id"] = link.id
    session["report_card_role"] = link.role_type
    teacher_name = session.get("teacher_name")
    if request.method == "POST":
        teacher_name = request.form.get( "teacher_name", "" ).strip()
        if not teacher_name:
            flash( "Enter your name.", "warning" )
        else:
            session["teacher_name"] = teacher_name
            return redirect(
                url_for( "teacher_report_card", token=token ) )
    if not teacher_name:
        return render_template( "teacher_name.html", token=token )
    settings = (
        SchoolSettings.query
        .filter_by(
            institution_id=current_report_card_institution()
        ) .first() )
    if not settings:
        return "School settings not configured.", 400
    classes = (
        filter_by_institution( AcademicClass.query, AcademicClass )
        .order_by( AcademicClass.name )
        .all() )
    selected_class = request.args.get( "class_id", type=int )
    students = []
    if selected_class:
        students = ( filter_by_institution( User.query, User )
            .filter_by( role="student", deleted=False,
                academic_class_id=selected_class )
            .order_by( User.first_name, User.last_name )
            .all() )
        
    batch = f"{settings.session}-{settings.term}"
    remarks = (
        filter_by_institution( ReportCardRemark.query, ReportCardRemark )
        .filter_by(
            term=settings.term, session=settings.session, report_batch=batch )
        .all() )     
    status_map = {}
    for r in remarks:
        status_map[r.student_id] = {
            "teacher": bool(r.teacher_completed_at),
            "principal": bool(r.principal_completed_at)
        }
    completed_count = sum( 1 for r in remarks if r.teacher_completed_at )
    total_students = len(students)
    remaining = total_students - completed_count
        
    return render_template(
        "teacher_report_card.html",
        teacher_name=teacher_name,
        settings=settings,
        classes=classes,
        students=students,
        selected_class=selected_class,
        token=token,
        completed_count=completed_count,
        remaining=remaining,
        total_students=total_students,        
        status_map=status_map
    )


#====Student Entry Route====
@app.route( "/report-card/teacher/<token>/student/<int:student_id>",
    methods=["GET", "POST"] )
def teacher_report_card_student(token, student_id):
    # --------------------------
    # Validate Link
    # --------------------------
    link = get_valid_report_card_link(token)
    if not link:
        abort(404)
    if link.role_type != "teacher":
        abort(403)
    session["report_card_institution"] = link.institution_id
    # --------------------------
    # Teacher Name
    # --------------------------
    teacher_name = session.get("teacher_name")

    if not teacher_name:
        return redirect(
            url_for(
                "teacher_report_card",
                token=token ) )
    # --------------------------
    # School Settings
    # --------------------------
    settings = (
        SchoolSettings.query
        .filter_by( institution_id=current_report_card_institution() )
        .first() )
    if not settings:
        abort(400)
    # --------------------------
    # Student
    # --------------------------
    student = (
        filter_by_institution( User.query, User )
        .filter_by( id=student_id, role="student", deleted=False )
        .first_or_404() )
    # --------------------------
    # Previous / Next Navigation
    # --------------------------
    class_students = (
        filter_by_institution(User.query, User)
        .filter_by( role="student", deleted=False,
            academic_class_id=student.academic_class_id )
        .order_by( User.first_name, User.last_name )
        .all() )
    student_ids = [s.id for s in class_students]
    current_index = student_ids.index(student.id)
    previous_student = (
        class_students[current_index - 1]
        if current_index > 0 else None )
    next_student = (
        class_students[current_index + 1]
        if current_index < len(class_students) - 1 else None )
    
    remark = (
        filter_by_institution( ReportCardRemark.query, ReportCardRemark )
        .filter_by(
            student_id=student.id,
            term=settings.term,
            session=settings.session,
            report_batch=settings.session + "-" + settings.term
        ) .first() )
    if remark and remark.teacher_locked:
        flash( "This report has already been submitted.", "info" )
        return render_template(
            "teacher_report_card_student.html",
            student=student,
            remark=remark,
            teacher_name=teacher_name,
            settings=settings,
            token=token,
            previous_student=previous_student,
            next_student=next_student,
            locked=True )
    
    if not remark:
        remark = ReportCardRemark(
            institution_id=current_report_card_institution(),
            student_id=student.id,
            academic_class_id=student.academic_class_id,
            term=settings.term,
            session=settings.session,
            report_batch=f"{settings.session}-{settings.term}" )
        remark.class_teacher_comment = get_auto_teacher_comment(
            student, settings )        
        

        
    if request.method == "POST":
        remark.teacher_name = teacher_name
        remark.teacher_link_id = session.get( "report_card_link_id" )
        remark.days_open = int( request.form.get("days_open", 0) )
        remark.days_present = int( request.form.get("days_present", 0) )
        if remark.days_present > remark.days_open:
            flash( "Days present cannot exceed days opened.", "danger" )      
        remark.punctuality = request.form.get( "punctuality" )
        remark.honesty = request.form.get( "honesty" )
        remark.neatness = request.form.get( "neatness" )
        remark.leadership = request.form.get( "leadership" )
        remark.handwriting = request.form.get( "handwriting" )
        remark.sports = request.form.get( "sports" )
        remark.class_teacher_comment = request.form.get(
            "class_teacher_comment" )
        remark.teacher_completed_at = datetime.utcnow()
        ok, message = validate_teacher_report_completion(remark)
        if not ok:
            flash(message, "danger")
            return redirect(request.referrer or url_for(
                "teacher_report_cards",
                token=token
            ))        
        remark.teacher_locked = True
        remark.teacher_locked_at = datetime.utcnow()        
        remark.teacher_updated_at = datetime.utcnow()
        db.session.add(remark)
        db.session.commit()
        audit = ReportCardAudit(
            institution_id=current_report_card_institution(),
            report_card_remark_id=remark.id,
            student_id=student.id,
            role_type="teacher",
            actor_name=teacher_name,
            action="Submitted",
            details="Teacher submitted report card." )
        db.session.add(audit)
        db.session.commit()        
        flash( "Report saved successfully.", "success" )

        action = request.form.get("action", "save")
        if action == "next" and next_student:
            return redirect(
                url_for(
                    "teacher_report_card_student",
                    token=token,
                    student_id=next_student.id ) )

        return redirect(
            url_for(
                "teacher_report_card",
                token=token,
                class_id=student.academic_class_id
            )
        )
    return render_template(
        "teacher_report_card_student.html",
        student=student,
        remark=remark,
        settings=settings,
        teacher_name=teacher_name,
        token=token,
        previous_student=previous_student,
        next_student=next_student        
    )    

#=======Principal Entry Route========
@app.route( "/report-card/principal/<token>", methods=["GET", "POST"] )
def principal_report_card(token):
    # --------------------------
    # Validate Link
    # --------------------------
    link = get_valid_report_card_link(token)
    if not link:
        abort(404)
    if link.role_type != "principal":
        abort(403)
    session["report_card_institution"] = link.institution_id
    session["report_card_link_id"] = link.id
    session["report_card_role"] = link.role_type
    principal_name = session.get("principal_name")
    if request.method == "POST":
        principal_name = request.form.get( "principal_name", "" ).strip()
        if not principal_name:
            flash( "Enter your name.", "warning" )
        else:
            session["principal_name"] = principal_name
            return redirect(
                url_for( "principal_report_card", token=token ) )
    if not principal_name:
        return render_template( "principal_name.html", token=token )
    settings = (
        SchoolSettings.query
        .filter_by( institution_id=current_report_card_institution() )
        .first() )
    if not settings:
        abort(400)
    classes = (
        filter_by_institution( AcademicClass.query, AcademicClass )
        .order_by(AcademicClass.name)
        .all() )
    selected_class = request.args.get( "class_id", type=int )
    students = []
    if selected_class:
        students = (
            filter_by_institution( User.query, User )
            .filter_by( role="student", deleted=False,
                academic_class_id=selected_class )
            .order_by( User.first_name, User.last_name )
            .all() )
        
    batch = f"{settings.session}-{settings.term}"
    remarks = (
        filter_by_institution( ReportCardRemark.query, ReportCardRemark )
        .filter_by(
            term=settings.term, session=settings.session,
            report_batch=batch )
        .all() )
    
    status_map = {}
    for r in remarks:
        status_map[r.student_id] = {
            "teacher": bool(r.teacher_completed_at),
            "principal": bool(r.principal_completed_at) }        

    return render_template(
        "principal_report_card.html",
        token=token,
        settings=settings,
        principal_name=principal_name,
        classes=classes,
        students=students,
        selected_class=selected_class,
        status_map=status_map
    )



@app.route( "/report-card/principal/<token>/student/<int:student_id>",
    methods=["GET", "POST"] )
def principal_report_card_student(token, student_id):
    # --------------------------
    # Validate Link
    # --------------------------
    link = get_valid_report_card_link(token)
    if not link:
        abort(404)
    if link.role_type != "principal":
        abort(403)
    session["report_card_institution"] = link.institution_id
    principal_name = session.get("principal_name")
    if not principal_name:
        return redirect(
            url_for( "principal_report_card", token=token ) )
    settings = (
        SchoolSettings.query
        .filter_by( institution_id=current_report_card_institution() )
        .first() )
    if not settings:
        abort(400)
    student = ( filter_by_institution( User.query, User )
        .filter_by( id=student_id, role="student", deleted=False )
        .first_or_404() )
    # --------------------------
    # Previous / Next Navigation
    # --------------------------
    class_students = (
        filter_by_institution(User.query, User)
        .filter_by( role="student", deleted=False,
            academic_class_id=student.academic_class_id )
        .order_by( User.first_name, User.last_name )
        .all() )
    student_ids = [s.id for s in class_students]
    current_index = student_ids.index(student.id)
    previous_student = (
        class_students[current_index - 1]
        if current_index > 0 else None )
    next_student = (
        class_students[current_index + 1]
        if current_index < len(class_students) - 1 else None )
    
    batch = f"{settings.session}-{settings.term}"
    remark = (
        filter_by_institution( ReportCardRemark.query, ReportCardRemark )
        .filter_by(
            student_id=student.id,
            term=settings.term,
            session=settings.session,
            report_batch=batch
        ) .first() )
    if remark and remark.principal_locked:
        flash( "This report has already been submitted.", "info" )
        return render_template(
            "teacher_report_card_student.html",
            student=student,
            remark=remark,
            teacher_name=teacher_name,
            settings=settings,
            token=token,
            previous_student=previous_student,
            next_student=next_student,
            locked=True
        )
    
    if not remark:
        remark = ReportCardRemark(
            institution_id=current_report_card_institution(),
            student_id=student.id,
            academic_class_id=student.academic_class_id,
            term=settings.term,
            session=settings.session,
            report_batch=batch )
        remark.principal_comment = get_auto_principal_comment(
            student, settings )
    if request.method == "POST":
        remark.principal_name = principal_name
        remark.principal_link_id = session.get( "report_card_link_id" )
        remark.principal_comment = request.form.get( "principal_comment" )
        remark.principal_completed_at = datetime.utcnow()
        ok, message = validate_principal_report_completion(remark)
        if not ok:
            flash(message, "danger")
            return redirect(request.referrer or url_for(
                "principal_report_cards",
                token=token
            ))        
        remark.principal_locked = True
        remark.principal_locked_at = datetime.utcnow()        
        remark.principal_updated_at = datetime.utcnow()
        db.session.add(remark)
        db.session.commit()
        audit = ReportCardAudit(
            institution_id=current_report_card_institution(),
            report_card_remark_id=remark.id,
            student_id=student.id,
            role_type="principal",
            actor_name=principal_name,
            action="Submitted",
            details="Principal submitted report card." )
        db.session.add(audit)
        db.session.commit()        
        flash( "Principal report saved.", "success" )

        action = request.form.get("action", "save")
        if action == "next" and next_student:
            return redirect(
                url_for(
                    "principal_report_card_student",
                    token=token,
                    student_id=next_student.id ) )
        return redirect(
            url_for(
                "principal_report_card",
                token=token,
                class_id=student.academic_class_id ) )

    return render_template(
        "principal_report_card_student.html",
        student=student,
        remark=remark,
        principal_name=principal_name,
        settings=settings,
        token=token,
        previous_student=previous_student,
        next_student=next_student      
    )

#====Unlock Route (Admin) to unlock a teacher's section====
@app.route( "/report-card/unlock/teacher/<int:remark_id>" )
@login_required
@roles_required('global_admin', 'admin')
##@admin_required
def unlock_teacher_report(remark_id):
    remark = get_record_in_scope( ReportCardRemark, remark_id )
    if not remark:
        abort(404)
    remark.teacher_locked = False
    remark.teacher_unlocked_at = datetime.utcnow()
    remark.teacher_unlocked_by = current_user.id
    db.session.commit()
    audit = ReportCardAudit(
        institution_id=current_institution_id(),
        report_card_remark_id=remark.id,
        student_id=remark.student_id,
        role_type="admin",
        actor_name=current_user.full_name,
        action="Unlocked Teacher",
        details="Teacher report unlocked." )
    db.session.add(audit)
    db.session.commit()    

    flash( "Teacher report unlocked.", "success" )

    return redirect(request.referrer or url_for("report_card_dashboard"))

@app.route( "/report-card/unlock/teacher/<int:remark_id>" )
@login_required
@roles_required('global_admin', 'admin')
##@admin_required
def unlock_principal_report(remark_id):
    remark = get_record_in_scope( ReportCardRemark, remark_id )
    if not remark:
        abort(404)
    remark.principal_locked = False
    remark.principal_unlocked_at = datetime.utcnow()
    remark.principal_unlocked_by = current_user.id
    db.session.commit()
    audit = ReportCardAudit(
        institution_id=current_institution_id(),
        report_card_remark_id=remark.id,
        student_id=remark.student_id,
        role_type="admin",
        actor_name=current_user.full_name,
        action="Unlocked Principal",
        details="Principal report unlocked."
    )

    db.session.add(audit)
    db.session.commit()    

    flash( "Principal report unlocked.", "success" )

    return redirect(request.referrer or url_for("report_card_dashboard"))

#====History Page====
@app.route( "/report-card/history/<int:remark_id>" )
@login_required
@roles_required("admin", "subadmin")
def report_card_history(remark_id):
    history = (
        filter_by_institution( ReportCardAudit.query, ReportCardAudit )
        .filter_by( report_card_remark_id=remark_id )
        .order_by( ReportCardAudit.created_at.desc() )
        .all() )
    return render_template(
        "report_card_history.html",
        history=history )
    


def current_report_card_institution():
    if "report_card_institution" in session:
        return session["report_card_institution"]
    return current_institution_id()

def get_auto_teacher_comment(student, settings):
    average = calculate_student_average(
        student.id, settings.term, settings.session )
    rows = (
        filter_by_institution(
            ReportCardAutoComment.query, ReportCardAutoComment )
        .filter_by( role_type="teacher" )
        .all() )
    matches = [ r for r in rows
        if r.minimum_score <= average <= r.maximum_score ]
    if not matches:
        return ""

    return random.choice(matches).comment

#====Auto-generate the principal comment====
def get_auto_principal_comment(student, settings):
    average = calculate_student_average( student.id,
        settings.term, settings.session )
    rows = (
        filter_by_institution( ReportCardAutoComment.query,
            ReportCardAutoComment )
        .filter_by( role_type="principal" )
        .all() )
    matches = [
        r for r in rows if r.minimum_score <= average <= r.maximum_score ]
    if not matches:
        return ""

    return random.choice(matches).comment

#====Report locking Validation Helpers====
def validate_teacher_report_completion(report):
    """
    Validate that the teacher has completed all required
    sections before locking the report.
    Returns:
        (True, "")
        (False, "Reason")
    """
    required_fields = {
        "Teacher Name": report.teacher_name,
        "Days School Open": report.days_open,
        "Days Present": report.days_present,
        "Punctuality": report.punctuality,
        "Honesty": report.honesty,
        "Neatness": report.neatness,
        "Leadership": report.leadership,
        "Handwriting": report.handwriting,
        "Sports": report.sports,
        "Teacher Comment": report.class_teacher_comment,
    }

    for label, value in required_fields.items():
        if value is None:
            return False, f"{label} is required."
        if isinstance(value, str) and not value.strip():
            return False, f"{label} is required."
    if report.days_present > report.days_open:
        return False, "Days Present cannot exceed Days School Open."
    return True, ""

def validate_principal_report_completion(report):
    """
    Validate that the principal has completed all required
    sections before locking.
    """
    if not report.principal_comment or not report.principal_comment.strip():
        return False, "Principal comment is required."
    return True, ""



#=============Report Card Entry===============
@app.route( "/classes/<int:class_id>/report_cards" )
@roles_required("global_admin", "admin", "subadmin")
def class_report_cards(class_id):
    academic_class = get_record_in_scope( AcademicClass, class_id )
    students = get_students_in_academic_class( academic_class.id )

    return render_template( "class_report_cards.html",
        academic_class=academic_class, students=students )

# ================= REPORT CARD =================
from flask import request, render_template, send_file
from io import BytesIO
import zipfile, json
from xhtml2pdf import pisa


# ---------- HELPER: POSITION ----------
def get_position(student, students):
    scores = []

    for s in students:
##        results = Result.query.filter_by(user_id=s.id).all()
        results = ( filter_by_institution( Result.query, Result )
            .filter_by( user_id=s.id )
            .all() )        
        total = sum([(r.score or 0) + (getattr(r, 'theory_score', 0) or 0) for r in results])
        scores.append((s.id, total))

    scores.sort(key=lambda x: x[1], reverse=True)

    for i, (sid, _) in enumerate(scores, start=1):
        if sid == student.id:
            return i
    return "-"


# ---------- PDF GENERATOR ----------
def generate_report_card_pdf(
    student,
    results,
    settings,
    design,
    students,
    attendance,
    grand_total,
    average,
    gpa,
    cgpa,
    teacher_name,
    remark,
    grading_scale,
    grading_scale_name,
    academic_mode):

    position = get_position(student, students)
    class_size = len(students)

    html = render_template(
        f"report_card_designs/{design}",
        student=student,
        results=results,
        settings=settings,
        position=position,
        class_size=class_size,
        attendance=attendance,
        grand_total=grand_total,
        average=average,
        gpa=gpa,
        cgpa=cgpa,
        grading_scale=grading_scale,
        grading_scale_name=grading_scale_name,
        academic_mode=academic_mode,
        teacher_name=teacher_name,
        remark=remark )

    filename = f"{student.username}_report.pdf"
    pdf_path = os.path.join(app.config["REPORT_CARD_UPLOAD"], filename)

    with open(pdf_path, "wb") as f:
        pisa.CreatePDF(html, dest=f)

    return pdf_path


# ================= REPORT CARD ROUTE =================
@app.route('/admin/report-card', methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin', 'subadmin')
def report_card():
    user = get_current_user()
##    settings = SchoolSettings.query.first()
    settings = ( filter_by_institution( SchoolSettings.query, SchoolSettings )
        .first() )
    if settings is None:
        flash("School settings have not been configured for this institution.", "warning")
        return redirect(url_for("admin_settings"))    
    attendance = {"opened": "", "present": "", "punctual": ""}
    search = request.args.get("search", "").strip()
    page = request.args.get('page', 1, type=int)
    per_page = 100
    students_query = ( filter_by_institution( User.query, User )
        .filter_by( role='student', deleted=False ) )
    if search:
        students_query = students_query.filter(
            (User.first_name.ilike(f"%{search}%")) |
            (User.last_name.ilike(f"%{search}%")) |
            (User.username.ilike(f"%{search}%"))
        )
    if user.role == 'subadmin':
        allowed_subject_ids = json.loads(user.allowed_subject_ids_json or "[]")
        if allowed_subject_ids:
            students_query = students_query.join(Result).filter(
                Result.subject_id.in_(allowed_subject_ids)
            ).distinct()
    students_paginated = students_query.order_by(
        User.first_name, User.last_name
    ).paginate(page=page, per_page=per_page, error_out=False)
    students = students_paginated.items
    start_index = (students_paginated.page - 1) * per_page
    designs_folder = os.path.join(app.root_path, "templates/report_card_designs")
    designs = [f for f in os.listdir(designs_folder) if f.endswith(".html")]
##    exam_labels = [x[0] for x in db.session.query(Subject.label)
##                   .filter(Subject.result_type == "exam").distinct().all()
##                   if x[0]]
##    ca_labels = [x[0] for x in db.session.query(Subject.label)
##                 .filter(Subject.result_type == "ca").distinct().all() if x[0]]
    exam_labels = [ x[0]
        for x in (
            filter_by_institution( db.session.query(Subject.label), Subject )
            .filter( Subject.result_type == "exam" )
            .distinct()
            .all() )
        if x[0] ]
    ca_labels = [ x[0]
        for x in (
            filter_by_institution( db.session.query(Subject.label), Subject )
            .filter( Subject.result_type == "ca" )
            .distinct()
            .all() )
        if x[0] ]    
    filtered_all_ids = [str(u.id) for u in students_query.with_entities(User.id).all()]
    total_students_count = len(filtered_all_ids)
    
##    academic_mode = settings.academic_mode
##    grading_scale_names = {
##        "percentage": "Percentage (100%)",
##        "gpa4": "GPA (4.0)",
##        "cgpa5": "CGPA (5.0)" }
##    grading_scale_name = grading_scale_names.get(
##        grading_scale, "Percentage (100%)" )
##    term = settings.term
##    report_design = settings.default_report_design###
    
    use_auto_comments = bool( request.form.get( "use_auto_comments" ) )
    # ================= POST =================
    if request.method == 'POST':
        academic_mode = settings.academic_mode
        grading_scale = settings.grading_scale
        grading_scale_names = {
            "percentage": "Percentage (100%)",
            "gpa4": "GPA (4.0)",
            "cgpa5": "CGPA (5.0)"
        }
        grading_scale_name = grading_scale_names.get(
            grading_scale, "Percentage (100%)" )
        term = settings.term
        report_design = settings.default_report_design
        
        exam_label = request.form.get("exam_label")
        ca_labels_selected = request.form.getlist("ca_labels")

        start_date = datetime.strptime(request.form.get("start_date"), "%Y-%m-%d")
        end_date = datetime.strptime(request.form.get("end_date"), "%Y-%m-%d") + timedelta(days=1)

        design = ( request.form.get("design")
            or settings.default_report_design )

        select_all = request.form.get("select_all", "0")
        student_ids_raw = request.form.get("student_ids", "[]")

        try:
            student_ids = json.loads(student_ids_raw)
        except:
            student_ids = []

        if select_all == "1":
            selected_students = students_query.all()
        else:
            if not student_ids:
                flash("Select at least one student")
                return redirect(url_for("report_card", search=search))

            selected_students = (
                filter_by_institution(User.query, User)
                .filter( User.id.in_(student_ids), User.deleted == False )
                .all() )

        pdf_files = []
        gpa = 0
        cgpa = 0
        for student in selected_students:
            student_results = ( filter_by_institution( Result.query, Result )
                .filter(
                    Result.user_id == student.id,
                    Result.taken_at >= start_date,
                    Result.taken_at < end_date,
                    Result.deleted == False )
                .order_by(Result.taken_at)
                .all() )
            remark = (
                filter_by_institution(StudentRemark.query, StudentRemark)
                .filter_by(student_id=student.id)
                .first()
            )
            teacher_name = ""
            if remark:
                teacher_name = remark.teacher_name or ""
            principal_name = ""
            if remark:
                principal_name = remark.principal_name or ""
                            
            attendance = {
                "opened": remark.days_open if remark else 0,
                "present": remark.days_present if remark else 0
            }
                

            if not student_results:
                continue

            subjects_map = {}
            total_score = 0

            for r in student_results:
##                subj = Subject.query.get(r.subject_id)
                subj = (
                    filter_by_institution( Subject.query, Subject )
                    .filter_by( id=r.subject_id )
                    .first() )                
                if not subj:
                    continue

                subject_name = subj.name

                if subject_name not in subjects_map:
                    subjects_map[subject_name] = {
                        "ca1": 0, "ca2": 0, "ca3": 0,
                        "exam_obj": 0, "exam_theory": 0
                    }

                label_upper = (subj.label or "").upper()

                if subj.result_type == "ca":

                    if ca_labels_selected and subj.label not in ca_labels_selected:
                        continue

                    if "1ST" in label_upper:
                        converted_ca = scale_score( r.score, 20,
                            settings.ca1_percentage )
                    elif "2ND" in label_upper:
                        converted_ca = scale_score( r.score, 20,
                            settings.ca2_percentage )
                    elif "3RD" in label_upper:
                        converted_ca = scale_score( r.score, 20,
                            settings.ca3_percentage )
                    else:
                        converted_ca = 0
                        
                    if "1ST" in label_upper:
                        subjects_map[subject_name]["ca1"] = converted_ca
                    elif "2ND" in label_upper:
                        subjects_map[subject_name]["ca2"] = converted_ca
                    elif "3RD" in label_upper:
                        subjects_map[subject_name]["ca3"] = converted_ca

                elif subj.result_type == "exam":

                    if exam_label and subj.label != exam_label:
                        continue

                    converted_exam_obj = scale_score( r.score, 50,
                        settings.objective_percentage
                                                      )
                    subjects_map[subject_name]["exam_obj"] = converted_exam_obj
                    subjects_map[subject_name]["exam_theory"] = scale_score(
                        r.theory_score, 50, settings.theory_percentage )
                    
            final_results = []

            for subject_name, vals in subjects_map.items():
                combined_ca = round( vals["ca1"] + vals["ca2"] + vals["ca3"], 2)
                subject_total = round( combined_ca + vals["exam_obj"] +
                    vals["exam_theory"], 2 )

                grade, remark = calculate_grade(subject_total)

                final_results.append({
                    "subject": subject_name,
                    "ca": combined_ca,
                    "exam_obj": vals["exam_obj"],
                    "exam_theory": vals["exam_theory"],
                    "total": subject_total,
                    "grade": grade,
                    "remark": remark })

                total_score += subject_total
            if not final_results:
                continue
            total_subjects = len(final_results)
            grand_total = round(total_score, 2)
            average = round(grand_total / total_subjects, 2)
            gpa = average#total_score / total_subjects if total_subjects else 0
            cgpa = average   # temporary until cumulative GPA is implemented                           

            print(student.username, len(final_results))
            pdf = generate_report_card_pdf( student=student,
                results=final_results, settings=settings,
                design=design, students=selected_students,
                attendance=attendance, grand_total=grand_total,
                average=average, gpa=gpa, cgpa=cgpa,
                teacher_name=teacher_name, remark=remark,
                grading_scale=grading_scale,
                grading_scale_name=grading_scale_name,
                academic_mode=academic_mode )


 ###           
            pdf_files.append(pdf)

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for f in pdf_files:
                zipf.write(f, os.path.basename(f))

        zip_buffer.seek(0)
        print("Generated PDFs:", len(pdf_files))
        print(student.username, len(student_results))

        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name="report_cards.zip"
        )

    return render_template(
        "admin_report_card.html",
        students=students,
        designs=designs,
        user=user,
        attendance=attendance,
        students_paginated=students_paginated,
        start_index=start_index,
        search=search,
        total_students_count=total_students_count,
        filtered_all_ids=filtered_all_ids,
        exam_labels=exam_labels,
        ca_labels=ca_labels
    )
#######################



#------Report Card Preview-------
@app.route('/admin/report-card/preview/<design>')
@login_required
@roles_required('global_admin', 'admin', 'subadmin')
def preview_report_card_design(design):
##    settings = SchoolSettings.query.first()
    settings = ( filter_by_institution( SchoolSettings.query, SchoolSettings )
        .first() )
    if settings is None:
        flash("School settings have not been configured for this institution.", "warning")
        return redirect(url_for("admin_settings"))    

##    # OPTIONAL SAFETY
##    allowed_designs = [
##        "blue.html",
##        "classic.html",
##        "gold.html",
##        "minimal.html",
##        "modern.html",
##        "standard.html"
##    ]
##
##    if design not in allowed_designs:
##        abort(404)

    # =========================================
    # DEMO STUDENT
    # =========================================
    student = {
        "first_name": "John",
        "last_name": "Doe",
        "username": "ADM001",
        "class_name": "JSS1",
        "age": 14
    }

    # =========================================
    # DEMO RESULTS
    # =========================================
    results = [

        {
            "subject": "Mathematics",
            "ca": 24,
            "exam_obj": 30,
            "exam_theory": 20,
            "total": 74,
            "grade": "A",
            "remark": "Excellent"
        },

        {
            "subject": "English",
            "ca": 24,
            "exam_obj": 28,
            "exam_theory": 18,
            "total": 70,
            "grade": "A",
            "remark": "Very Good"
        }

    ]

    # =========================================
    # TOTALS
    # =========================================
    grand_total = sum(r["total"] for r in results)

    average = round(
        grand_total / len(results),
        2
    ) if results else 0

    # =========================================
    # RENDER
    # =========================================
    return render_template(
        f"report_card_designs/{design}",
        settings=settings,
        student=student,
        results=results,
        grand_total=grand_total,
        average=average,
        position=1,
        class_size=40,
        gpa=3.5,
        cgpa=3.2,
        remark = None,
        teacher_name = "Demo Teacher",
        grading_scale = "percentage",
        grading_scale_name = "Percentage (100%)",
        academic_mode = "secondary"        
    )

# =========================================
# SEND REPORT CARDS PAGE
# =========================================
@app.route('/admin/send-report-cards', methods=['GET', 'POST'])
@login_required
@roles_required('global_admin', 'admin', 'subadmin')
def send_report_cards():
    user = get_current_user()
    students = ( filter_by_institution( User.query, User )
        .filter_by( role='student', deleted=False )
        .all() )
    if request.method == "POST":
        student_ids = request.form.getlist("student_ids")
        manual_emails = request.form.get("manual_emails", "")

        emails = []

        # selected students
##        selected_students = User.query.filter(User.id.in_(student_ids),
##                                              deleted=False).all()
        selected_students = (
            filter_by_institution( User.query, User )
            .filter( User.id.in_(student_ids), User.deleted == False )
            .all() )        

        for s in selected_students:
            if s.email:
                emails.append(s.email)

        # manual emails
        if manual_emails:
            emails.extend([
                e.strip() for e in manual_emails.split(",") if e.strip()
            ])

        app_id = get_setting("app_id", "NOT SET")

        for email in emails:

            try:
                msg = Message(
                    subject="Report Card",
                    recipients=[email],
                    body=f"""
                    App ID: {app_id}

                    Your report card has been generated and is
                    attached/available.

                    Regards,
                    Exam System
                    """
                        )

##                load_admin_mail()
####                mail.send(msg)
##                queue_email(
##                    recipient=msg.recipients[0],
##                    subject=msg.subject,
##                    body=msg.body
##                )
                queue_email(
                    recipient=msg.recipients[0],
                    subject=msg.subject,
                    body=msg.body,
                    mail_profile="admin",
                    institution_id=current_institution_id()
                )                

            except Exception as e:
                app.logger.error(f"Report card email failed: {e}")

        flash(f"Report cards sent to {len(emails)} recipients", "success")
        return redirect(url_for("send_report_cards"))

    return render_template(
        "send_report_cards.html",
        students=students,
        user=user
    )

#========DEBUG; Print the currently logged-in user ========
@app.route('/debug_user')
def debug_user():
    uid = session.get('uid')

    user = User.query.get(uid) if uid else None

    return {
        "session_uid": uid,
        "user_exists": bool(user),
        "username": user.username if user else None,
        "role": user.role if user else None
    }

#--------Add column to database table --------
#Run:   http://127.0.0.1:5000/fix-show-answers-column
@app.route('/fix-show-answers-column')
def fix_show_answers_column():

    from sqlalchemy import text

    try:

        db.session.execute(text("""

            ALTER TABLE subject
            ADD COLUMN show_answers_to_students BOOLEAN DEFAULT 0

        """))

        db.session.commit()

        return "✅ Column added successfully."

    except Exception as e:

        return f"Error: {e}"
#Run:   http://127.0.0.1:5000/fix-series-group-column
@app.route('/fix-series-group-column')
def fix_series_group_column():

    from sqlalchemy import text

    try:

        db.session.execute(text("""

            ALTER TABLE subject
            ADD COLUMN series_group VARCHAR

        """))

        db.session.commit()

        return "✅ series_group column added successfully."

    except Exception as e:

        return f"Error: {e}"

#Run:   http://127.0.0.1:5000//fix-subject-recycle-columns
@app.route('/fix-subject-recycle-columns')
def fix_subject_recycle_columns():
    from sqlalchemy import text
    try:
        # =====================================
        # deleted column
        # =====================================
        try:
            db.session.execute(text("""
                ALTER TABLE subject
                ADD COLUMN deleted BOOLEAN DEFAULT 0
            """))
            db.session.commit()
            print("deleted column added")
        except Exception as e:
            print("deleted column:", e)
        # =====================================
        # deleted_at column
        # =====================================
        try:
            db.session.execute(text("""
                ALTER TABLE subject
                ADD COLUMN deleted_at DATETIME
            """))
            db.session.commit()
            print("deleted_at column added")
        except Exception as e:
            print("deleted_at column:", e)
        return (
            "✅ Subject recycle bin columns "
            "checked successfully."
        )
    except Exception as e:
        return f"Error: {e}"

#Run:   http://127.0.0.1:5000//fix-group-recycle-columns
@app.route('/fix-group-recycle-columns')
def fix_group_recycle_columns():
    from sqlalchemy import text
    try:
        # =====================================
        # deleted column
        # =====================================
        try:
            db.session.execute(text("""
                ALTER TABLE subject_group
                ADD COLUMN deleted BOOLEAN DEFAULT 0
            """))
            db.session.commit()
            print("deleted column added")
        except Exception as e:
            print("deleted column:", e)
        # =====================================
        # deleted_at column
        # =====================================
        try:
            db.session.execute(text("""
                ALTER TABLE subject_group
                ADD COLUMN deleted_at DATETIME
            """))
            db.session.commit()
            print("deleted_at column added")
        except Exception as e:
            print("deleted_at column:", e)
        return "✅ SubjectGroup recycle bin columns checked successfully."
    except Exception as e:
        return f"Error: {e}"

# =====================================
# Data Migrations
# =====================================
#===========SQLite → PostgreSQL=============
@app.route( "/admin/import_sqlite_to_postgres", methods=["POST"] )
@login_required
@roles_required("global_admin")
def import_sqlite_to_postgres():
    sync_database_schema()
    uploaded = request.files.get( "sqlite_file" )
    if not uploaded:
        flash(
            "Please select a SQLite database file.",
            "danger"
        )
        return redirect( url_for("admin_settings") )
    temp_path = os.path.join(
        app.instance_path,
        "migration_import.db"
    )
    uploaded.save(temp_path)
    replace_existing = (
        request.form.get(
            "replace_existing"
        ) == "on"
    )
    report = migrate_sqlite_to_postgres( temp_path, replace_existing )
    flash( report, "success" )
    return redirect( url_for("admin_settings") )

#==========PostgreSQL → Export===============
@app.route( "/admin/export_postgres_backup", methods=["POST"] )
@login_required
@roles_required("global_admin")
def export_postgres_backup():
    sync_database_schema()
    export_folder = os.path.join(app.instance_path, "exports")
    os.makedirs(
        export_folder,
        exist_ok=True
    )
    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )
    backup_file = os.path.join(
        export_folder,
        f"postgres_backup_{timestamp}.sql"
    )
    create_postgres_backup( backup_file )
    return send_file(
        backup_file,
        as_attachment=True,
        download_name=os.path.basename(
            backup_file
        )
    )


#===============Verify databases==============
@app.route( "/admin/verify_databases", methods=["POST"] )
@login_required
@roles_required("global_admin")
def verify_databases():
    sync_database_schema()
    report = verify_database_integrity()
    session["verify_report"] = report
    return redirect(
        url_for("admin_settings")
    )

#===============Clear Verify report==============
@app.route( "/admin/clear_verify_report", methods=["POST"] )
@login_required
@roles_required("global_admin")
def clear_verify_report():
    session.pop(
        "verify_report",
        None
    )
    return redirect(
        url_for("admin_settings")
    )

def migrate_sqlite_to_postgres(
    sqlite_path,
    replace_existing=False
):
    sync_database_schema()
    sqlite_conn = sqlite3.connect(sqlite_path)
    sqlite_conn.row_factory = sqlite3.Row
    sqlite_cur = sqlite_conn.cursor()
    inspector = inspect(db.engine)
    # =========================================================
    # SAFE DEPENDENCY ORDER
    # =========================================================
    ordered_tables = [
        # -----------------------------
        # ROOT TABLES
        # -----------------------------
        "app_settings",
        "app_config",
        "school_settings",
        "email_queue",
        "auto_license_config",
        "auto_license_request",

        "subject_group",
        "user",

        # -----------------------------
        # DEPEND ON USER / GROUP
        # -----------------------------
        "subject",
        "license",
        "chat_message",
        "student_remark",
        "student_term_record",
        "auto_comment",

        # -----------------------------
        # DEPEND ON SUBJECT
        # -----------------------------
        "question",
        "theory_question",

        # -----------------------------
        # DEPEND ON USER + SUBJECT
        # -----------------------------
        "exam_session",
        "theory_submission",

        # -----------------------------
        # DEPEND ON SESSION
        # -----------------------------
        "result"
    ]
    db_tables = inspector.get_table_names()
    for t in db_tables:
        if (
            t not in ordered_tables
            and t != "alembic_version"
        ):
            ordered_tables.append(t)
    tables = ordered_tables
    total_tables = 0
    total_rows = 0
    failed_rows = 0
    # =========================================================
    # PRINT ONLY FIRST ERROR PER TABLE
    # =========================================================
    printed_error_tables = set()
    # =========================================================
    # REPLACE EXISTING DATA
    # =========================================================
    if replace_existing:
        print("Replacing existing PostgreSQL data...")
        for table in reversed(tables):
            if table == "alembic_version":
                continue
            try:
                db.session.execute(
                    text(
                        f'''
                        TRUNCATE TABLE "{table}"
                        RESTART IDENTITY
                        CASCADE
                        '''
                    )
                )
            except Exception as e:
                print(
                    f"TRUNCATE failed {table}: {e}"
                )
                db.session.rollback()
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
    # =========================================================
    # DISABLE FK CHECKS TEMPORARILY
    # =========================================================
    try:
        db.session.execute(
            text(
                "SET session_replication_role = 'replica';"
            )
        )
        db.session.commit()
        print(
            "Foreign-key checks temporarily disabled."
        )
    except Exception as e:
        print(
            f"Warning: could not disable FK checks: {e}"
        )
        db.session.rollback()
    # =========================================================
    # IMPORT TABLES
    # =========================================================
    for table in tables:
        if table == "alembic_version":
            continue
        try:
            sqlite_cur.execute(
                f'SELECT * FROM "{table}"'
            )
        except Exception:
            print(
                f"SQLite table missing: {table}"
            )
            continue
        rows = sqlite_cur.fetchall()
        if not rows:
            continue
        total_tables += 1
        # =====================================================
        # DETECT BOOLEAN COLUMNS FROM POSTGRES
        # =====================================================
        boolean_columns = set()
        try:
            pg_columns = inspector.get_columns(table)
            for col in pg_columns:
                if str(col["type"]).upper() == "BOOLEAN":
                    boolean_columns.add(col["name"])
        except Exception as e:
            print(
                f"Could not inspect columns for {table}: {e}"
            )
        columns = rows[0].keys()
        column_list = ", ".join(
            f'"{c}"'
            for c in columns
        )
        placeholders = ", ".join(
            f":{c}"
            for c in columns
        )
        insert_sql = text(
            f'''
            INSERT INTO "{table}"
            ({column_list})
            VALUES ({placeholders})
            '''
        )
        copied = 0
        # =====================================================
        # ROW-BY-ROW IMPORT
        # =====================================================
        for row in rows:
            try:
                row_data = dict(row)
                # =========================================
                # SQLITE INTEGER -> POSTGRES BOOLEAN
                # =========================================
                for col in boolean_columns:
                    if col not in row_data:
                        continue
                    if row_data[col] == 1:
                        row_data[col] = True
                    elif row_data[col] == 0:
                        row_data[col] = False
                db.session.execute( insert_sql, row_data )
                db.session.commit()
                copied += 1
            except Exception as e:
                db.session.rollback()
                failed_rows += 1
                if table not in printed_error_tables:
                    printed_error_tables.add(table)
                    print("=" * 80)
                    print(f"FIRST ERROR IN TABLE: {table}")
                    print(f"ROW: {dict(row)}")
                    print(f"ERROR: {repr(e)}")
                    print("=" * 80)
                continue
        total_rows += copied
        print(
            f"{table}: {copied} rows copied"
        )
    sqlite_conn.close()
    # =========================================================
    # RESTORE FK CHECKS
    # =========================================================
    try:
        db.session.execute(
            text(
                "SET session_replication_role = 'origin';"
            )
        )
        db.session.commit()
        print(
            "Foreign-key checks restored."
        )
    except Exception as e:
        print(
            f"Warning: could not restore FK checks: {e}"
        )
        db.session.rollback()
    # =========================================================
    # RESET POSTGRES SEQUENCES
    # =========================================================
    reset_postgres_sequences()
    print(
        f"Migration completed. "
        f"Tables={total_tables}, "
        f"Rows={total_rows}, "
        f"Failed={failed_rows}"
    )
    return (
        f"Migration completed. "
        f"Tables: {total_tables}, "
        f"Rows: {total_rows}, "
        f"Failed rows: {failed_rows}. "
        f"PostgreSQL sequences reset and "
        f"foreign-key mode restored."
    )

@app.route( "/admin/restore_postgres_backup", methods=["POST"] )
@login_required
@roles_required("global_admin")
def restore_postgres_backup():
    uploaded = request.files.get( "backup_file" )
    if not uploaded:
        flash(
            "No backup file selected.",
            "danger"
        )
        return redirect( url_for("admin_settings") )
    temp_file = os.path.join(
        tempfile.gettempdir(),
        secure_filename(
            uploaded.filename
        )
    )
    uploaded.save( temp_file )
    try:
        restore_postgres_database( temp_file )
        flash(
            "Backup restored successfully.",
            "success"
        )
    except Exception as e:
        flash(
            f"Restore failed: {e}",
            "danger"
        )
    return redirect( url_for("admin_settings") )

@app.route( "/admin/import_postgres_backup", methods=["POST"] )
@login_required
@roles_required("global_admin")
def import_postgres_backup():
    uploaded = request.files.get( "backup_file" )
    if not uploaded:
        flash(
            "Select a backup file.",
            "danger"
        )
        return redirect( url_for("admin_settings") )
    temp_file = os.path.join(
        tempfile.gettempdir(),
        secure_filename(
            uploaded.filename
        )
    )
    uploaded.save( temp_file )
    try:
        restore_postgres_database( temp_file )
        flash(
            "PostgreSQL backup imported successfully.",
            "success"
        )
    except Exception as e:
        flash(
            str(e),
            "danger"
        )
    return redirect(
        url_for("admin_settings")
    )

def restore_postgres_database( backup_file ):
    db_url = app.config[ "SQLALCHEMY_DATABASE_URI" ]
    result = subprocess.run(
        [
            "psql",
            db_url,
            "-f",
            backup_file
        ],
        capture_output=True,
        text=True
    )
    if result.returncode != 0:
        raise Exception(
            result.stderr
        )


def create_postgres_backup( backup_file ):
    db_url = app.config[
        "SQLALCHEMY_DATABASE_URI"
    ]
    result = subprocess.run(
        [
            "pg_dump",
            db_url,
            "-f",
            backup_file
        ],
        capture_output=True,
        text=True
    )
    if result.returncode != 0:
        raise Exception(
            result.stderr
        )
    return backup_file

def reset_postgres_sequences():
    inspector = inspect( db.engine )
    tables = inspector.get_table_names()
    for table in tables:
        if table == "alembic_version":
            continue
        columns = inspector.get_columns( table )
        id_column = None
        for col in columns:
            if col["name"] == "id":
                id_column = "id"
                break
        if not id_column:
            continue
        try:
            db.session.execute(
                text(
                    f"""
                    SELECT setval(
                        pg_get_serial_sequence(
                            '"{table}"',
                            'id'
                        ),
                        COALESCE(
                            (
                                SELECT MAX(id)
                                FROM "{table}"
                            ),
                            1
                        ),
                        true
                    )
                    """
                )
            )
            print( f"Sequence reset: {table}" )
        except Exception as e:
            print(
                f"Sequence reset failed for "
                f"{table}: {e}"
            )
    db.session.commit()

#==========Produces database useful summary=============
def verify_database_integrity():
    inspector = inspect( db.engine )
    tables = inspector.get_table_names()
    report = []
    total_rows = 0
    for table in tables:
        if table == "alembic_version":
            continue
        try:
            count = db.session.execute(
                text(
                    f'''
                    SELECT COUNT(*)
                    FROM "{table}"
                    '''
                )
            ).scalar()
            total_rows += count
            report.append(
                f"{table}={count}"
            )
        except Exception as e:
            report.append( f"{table}=ERROR" )
    return (
        f"Verified "
        f"{len(report)} tables. "
        f"Total rows={total_rows}. "
        + " | ".join(report)
    )

#=========Reactivate affected accounts===========
@app.route("/dev/reactivate_accounts")
@login_required
@roles_required("global_admin")
def dev_reactivate_accounts():

    # ---------------------------------------------------
    # Global Admins
    # ---------------------------------------------------
    User.query.filter_by(
        role="global_admin"
    ).update({
        User.approved: True,
        User.approval_status: "Approved",
        User.account_status: "active",
        User.is_global: True
    }, synchronize_session=False)

    # ---------------------------------------------------
    # Institution Admins
    # ---------------------------------------------------
    User.query.filter(
        User.is_institution_admin == True
    ).update({
        User.approved: True,
        User.approval_status: "Approved",
        User.account_status: "active",
        User.role: "admin"
    }, synchronize_session=False)

    # ---------------------------------------------------
    # Students
    # ---------------------------------------------------
    User.query.filter_by(
        role="student"
    ).update({
        User.approved: True,
        User.approval_status: "Approved",
        User.account_status: "active"
    }, synchronize_session=False)

    # ---------------------------------------------------
    # Subadmins
    # ---------------------------------------------------
    User.query.filter_by(
        role="subadmin"
    ).update({
        User.approved: True,
        User.approval_status: "Approved",
        User.account_status: "active"
    }, synchronize_session=False)

    db.session.commit()

    flash(
        "All accounts reactivated successfully.",
        "success"
    )

    return redirect(url_for("dashboard"))
### =========================================================
### TEMPORARY - FIX INSTITUTION ADMIN FLAG
### RUN ONCE THEN DELETE THIS ROUTE
### =========================================================
##@app.route("/temp/fix-institution-admin-flag")
##@login_required
##@roles_required("admin")
##def temp_fix_institution_admin_flag():
##    current = get_current_user()
##    if not current.is_global:
##        abort(403)
##    updated = 0
##    users = User.query.all()
##    for user in users:
##        if (
##            user.role == "institution_admin"
##            or (
##                user.role == "admin"
##                and not user.is_global
##                and user.institution_id is not None ) ):
##            if not user.is_institution_admin:
##                user.is_institution_admin = True
##                updated += 1
##    db.session.commit()
##    flash(
##        f"{updated} Institution Administrator account(s) updated successfully.",
##        "success"
##    )
##    return redirect(url_for("approval_center"))

# ---------- Main ----------
if __name__ == "__main__":
    with app.app_context():
        sync_database_schema()
        ensure_free_institution()
##        migrate_legacy_records_to_free_institution()
        migrate_null_records_to_free()
        restore_default_licenser()
        db.session.commit()
        print("✅ Licenser account verified.")
        ensure_admin_created()
        sync_student_academic_classes()
        sync_student_class_history()
    # Start Flask-SocketIO server
    socketio.run(app, host="0.0.0.0", port=5000, debug=True)



"""
📄 CSV Format
    Save as questions.csv:
        question,a,b,c,d,answer
        What is the capital of France?,Paris,London,Berlin,Madrid,A
        2 + 2 equals?,1,2,3,4,D
        Which planet is known as the Red Planet?,Earth,Venus,Mars,Jupiter,C
    First row = headers, and answer must be A, B, C, or D.

📄 XLSX Format (Excel)
    Sheet name: Questions

        question	                a	b	c	d	answer
        What is the capital of France?	Paris	London	Berlin	Madrid	A
        2 + 2 equals?	                1	2	3	4	D
        Which planet is the Red Planet?	Earth	Venus	Mars	Jupiter	C

        Columns must be exactly: question, a, b, c, d, answer, and Answer is a
        single letter (A–D).
Generated Licenses
365days
580deaf07e5b357f21cb3f0af5668a01
4391f07137b90e8a7be3a39126551f41
1day
984306e67df358a9d32bc1594d9c6772
"""

